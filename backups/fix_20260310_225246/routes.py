# ============================================================================
# AMPOULEX ERP - COMPLETE ROUTES FILE
# All Errors Fixed + Production Module + Barcodes
# ============================================================================

from datetime import datetime, timedelta
import os
import json
import random
import qrcode.constants

from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_user, logout_user, login_required, current_user
from flask_socketio import emit
from email.mime.multipart import MIMEMultipart  # Fix: "MIMEMultipart" is not defined
#from email.mime.text import MIMEText  # Fix: "MIMEText" is not defined

from app import db, socketio
from app.models import (
    User, Product, Customer, Inquiry, Order, OrderItem,
    Employee, Expense, Accounting, CompanySetting, CustomerProductPrice,
    TaxSetting, StockAdjustment, Report, InquiryItem,
    Attendance, Timesheet, LeaveRequest, PayrollPayment,
    RawMaterial, BOMItem, RawMaterialUsage, MaterialUsage, ProductionBatch,
    StockAlert, OrderApproval, BatchRawMaterial,
    Supplier, PurchaseOrder, PurchaseOrderItem, GoodsReceipt, GoodsReceiptItem,
    Warehouse, WarehouseStock, StockTransfer, StockTransferItem, MaterialBatch, SupplierInvoice)
from app.utils.tax_calculator import calculate_monthly_tax_deduction

main_bp = Blueprint('main', __name__)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

# ============================================================================
# CUSTOMER AUTO-CREATION HELPERS (NEW - CRITICAL!)
# ============================================================================

# ============================================================================
# CUSTOMER AUTO-CREATION HELPER (ADD THIS)
# ============================================================================

def find_or_create_customer(customer_name=None, business_name=None, email=None, phone=None):
    """
    Find existing customer or create new one automatically.
    Prevents duplicates by checking email, phone, and business name.
    Returns: (customer, is_new)
    """
    # Try to find by email first
    if email and email.strip():
        customer = Customer.query.filter_by(email=email.strip().lower()).first()
        if customer:
            return customer, False
    
    # Try to find by phone
    if phone and phone.strip():
        phone_clean = ''.join(filter(str.isdigit, phone))
        customers_by_phone = Customer.query.filter(Customer.phone.like(f'%{phone_clean}%')).all()
        if customers_by_phone:
            return customers_by_phone[0], False
    
    # Try to find by business name
    if business_name and business_name.strip():
        customer = Customer.query.filter_by(business_name=business_name.strip()).first()
        if customer:
            return customer, False
    
    # Try to find by customer name
    if customer_name and customer_name.strip():
        customer = Customer.query.filter_by(name=customer_name.strip()).first()
        if customer:
            return customer, False
    
    # No existing customer found - create new one
    new_customer = Customer(
        name=customer_name or 'N/A',
        business_name=business_name or 'N/A',
        email=email.strip().lower() if email else '',
        phone=phone or '',
        role='Customer',
        payment_terms='cash',
        credit_limit=0,
        is_active=True
    )
    db.session.add(new_customer)
    db.session.commit()
    
    return new_customer, True

# ============================================================================
# EMAIL NOTIFICATION HELPERS (PHASE 4)
# ============================================================================

def send_email(subject, recipients, body, html=None):
    """Send email notification"""
    try:
        from flask_mail import Mail, Message
        
        mail = Mail(current_app)
        msg = Message(subject, recipients=recipients if isinstance(recipients, list) else [recipients])
        msg.body = body
        if html:
            msg.html = html
        
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

def notify_order_created(order):
    """Send email notification when order is created"""
    subject = f'New Order Created - {order.order_number}'
    
    body = f"""
    Dear Team,
    
    A new order has been created:
    
    Order Number: {order.order_number}
    Customer: {order.customer_name_snapshot}
    Total Amount: PKR {order.total_amount:.2f}
    Status: {order.status}
    
    Please process this order promptly.
    
    Best regards,
    Ampoulex
    """
    
    # Send to admin email
    send_email(subject, 'ampoulex.hr@gmail.com', body)

def notify_low_stock(product):
    """Send email notification for low stock products"""
    subject = f'Low Stock Alert - {product.name}'
    
    body = f"""
    Dear Inventory Team,
    
    The following product is running low on stock:
    
    Product: {product.name}
    Color: {product.color}
    Current Stock: {product.stock}
    Reorder Level: 10,000
    
    Please arrange for restocking immediately.
    
    Best regards,
    Ampoulex
    """
    
    send_email(subject, 'ampoulex.hr@gmail.com', body)

def notify_payment_received(order):
    """Send email notification when payment is received"""
    if not order.customer_email_snapshot or order.customer_email_snapshot == 'N/A':
        return
    
    subject = f'Payment Received - Order {order.order_number}'
    
    body = f"""
    Dear {order.customer_name_snapshot},
    
    Thank you for your payment!
    
    Order Number: {order.order_number}
    Amount Paid: PKR {order.paid_amount:.2f}
    Payment Date: {order.payment_date.strftime('%Y-%m-%d') if order.payment_date else 'N/A'}
    
    Your order is being processed and will be shipped soon.
    
    Best regards,
    Ampoulex Team
    Malik Arshad Farmhouse, Malik Akram Street, Darbar-e-Kareemi Stop, G.T. Road
    Wah Cantt, The. Taxila, Dist. RWP, Punjab, Pakistan
    Tel: 0340-5336238 | 0331-9980908
    """
    
    send_email(subject, order.customer_email_snapshot, body)

def notify_order_shipped(order):
    """Send email notification when order is shipped"""
    if not order.customer_email_snapshot or order.customer_email_snapshot == 'N/A':
        return
    
    subject = f'Your Order Has Been Shipped - {order.order_number}'
    
    body = f"""
    Dear {order.customer_name_snapshot},
    
    Great news! Your order has been shipped.
    
    Order Number: {order.order_number}
    Total Amount: PKR {order.total_amount:.2f}
    
    You will receive tracking information shortly.
    
    Thank you for your business!
    
    Best regards,
    Ampoulex Team
    Malik Arshad Farmhouse, Malik Akram Street, Darbar-e-Kareemi Stop, G.T. Road
    Wah Cantt, The. Taxila, Dist. RWP, Punjab, Pakistan
    Tel: 0340-5336238 | 0331-9980908
    """
    
    send_email(subject, order.customer_email_snapshot, body)

def broadcast_update(event_name, data):
    """Broadcast real-time updates to all connected admin clients"""
    try:
        socketio.emit(event_name, data, namespace='/admin')
    except Exception as e:
        print(f"Socket.IO error: {e}")

def notify_new_inquiry(inquiry):
    """Notify when new inquiry is created"""
    broadcast_update('new_inquiry', {
        'id': inquiry.id,
        'number': inquiry.inquiry_number,
        'customer': inquiry.customer_name,
        'business': inquiry.business_name,
        'quantity': inquiry.quantity,
        'status': inquiry.status,
        'created_at': inquiry.created_at.isoformat()
    })

def notify_new_order(order):
    """Notify when new order is created"""
    broadcast_update('new_order', {
        'id': order.id,
        'number': order.order_number,
        'customer': order.customer_name_snapshot,
        'total': float(order.total_amount),
        'status': order.status,
        'payment_status': order.payment_status,
        'created_at': order.created_at.isoformat()
    })

def notify_low_stock_alert(product):
    """Notify when product stock is low"""
    broadcast_update('low_stock_alert', {
        'product_id': product.id,
        'product_name': product.name,
        'current_stock': product.stock,
        'threshold': 10000
    })

def group_products_by_base():
    """Group products by base name with variants"""
    products = Product.query.filter_by(is_deleted=False).all()
    grouped = {}
    
    for product in products:
        base_name = product.name.split(' - ')[0] if ' - ' in product.name else product.name
        
        if base_name not in grouped:
            grouped[base_name] = {
                'name': base_name,
                'specification': product.specification,
                'product_type': product.product_type,
                'material_type': product.material_type,
                'usp_type': product.usp_type,
                'shape_type': product.shape_type,
                'dimensions': product.dimensions,
                'use_case': product.use_case,
                'variants': []
            }
        
        grouped[base_name]['variants'].append({
            'id': product.id,
            'color': product.color,
            'price': integer(product.base_price),
            'stock': product.stock,
            'product_type': product.product_type
        })
    
    return list(grouped.values())

def notify_attendance_marked(attendance):
    """Notify when attendance is marked"""
    broadcast_update('attendance_updated', {
        'employee': attendance.employee.name,
        'date': attendance.date.isoformat(),
        'status': attendance.status,
        'hours': float(attendance.hours_worked)
    })

def check_and_create_stock_alerts():
    """Check all products and create alerts for low stock"""
    low_stock_products = Product.query.filter(
        Product.stock < 10000,
        Product.is_deleted == False
    ).all()
    
    for product in low_stock_products:
        existing_alert = StockAlert.query.filter_by(
            product_id=product.id,
            is_resolved=False
        ).first()
        
        if not existing_alert:
            alert = StockAlert(
                product_id=product.id,
                threshold=10000,
                current_stock=product.stock
            )
            db.session.add(alert)
            notify_low_stock_alert(product)
    
    db.session.commit()

def send_email_report(recipient_email, report_type, report_data):
    """Send email report (stub - configure SMTP settings)"""
    try:
        # Configure your SMTP settings here
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', 587))
        smtp_user = os.getenv('SMTP_USER', '')
        smtp_password = os.getenv('SMTP_PASSWORD', '')
        
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = recipient_email
        msg['Subject'] = f'Ampoulex {report_type} Report'
        
        body = f"""
        Ampoulex {report_type} Report
        Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}
        
        {report_data}
        
        This is an automated report from Ampoulex.
        """
        msg.attach(MIMEText(body, 'plain'))
        
        # Uncomment to enable email sending
        # with smtplib.SMTP(smtp_server, smtp_port) as server:
        #     server.starttls()
        #     server.login(smtp_user, smtp_password)
        #     server.send_message(msg)
        
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

# ============================================================================
# AUTHENTICATION
# ============================================================================

@main_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('main.dashboard'))
        flash('Invalid username or password', 'error')
    
    return render_template('auth/login.html')

@main_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))

# ============================================================================
# DASHBOARD (PHASE 2 ENHANCED)
# ============================================================================

@main_bp.route('/')
@main_bp.route('/dashboard')
@login_required
def dashboard():
    check_and_create_stock_alerts()
    
    total_revenue = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing'])
    ).scalar() or 0
    
    total_orders = Order.query.count()
    active_inquiries = Inquiry.query.filter(
        Inquiry.status.in_(['new', 'followup']), 
        Inquiry.is_deleted == False
    ).count()
    total_customers = Customer.query.filter_by(is_active=True, is_deleted=False).count()
    
    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(5).all()
    recent_inquiries = Inquiry.query.filter_by(is_deleted=False).order_by(Inquiry.created_at.desc()).limit(5).all()
    
    return render_template('dashboard.html', 
                         total_revenue=total_revenue, 
                         total_orders=total_orders,
                         active_inquiries=active_inquiries, 
                         total_customers=total_customers,
                         recent_orders=recent_orders, 
                         recent_inquiries=recent_inquiries)
# ============================================================================
# PRODUCTS
# ============================================================================

@main_bp.route('/products')
@login_required
def products():
    products = Product.query.filter_by(is_deleted=False).all()
    low_stock_products = [p for p in products if p.stock < 10000]
    return render_template('products/index.html', products=products, low_stock_products=low_stock_products)

@main_bp.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    if request.method == 'POST':
        try:
            product = Product(
                name=request.form.get('name', ''),
                specification=request.form.get('specification', ''),
                base_price=integer(request.form.get('base_price', 0)),
                stock=int(request.form.get('stock', 0)),
                color=request.form.get('color', ''),
                product_type=request.form.get('product_type', 'product'),
                material_type=request.form.get('material_type', ''),
                usp_type=request.form.get('usp_type', ''),
                shape_type=request.form.get('shape_type', ''),
                dimensions=request.form.get('dimensions', ''),
                use_case=request.form.get('use_case', '')
            )
            db.session.add(product)
            db.session.commit()
            
            adj_type = request.form.get('adjustment_type', '')
            adj_qty = request.form.get('adjustment_qty', '0')
            if adj_type and adj_qty:
                adjustment = StockAdjustment(
                    product_id=product.id, 
                    adjustment_type=adj_type,
                    quantity=int(adj_qty), 
                    reason=request.form.get('adjustment_reason', 'Initial stock'),
                    adjusted_by=current_user.username
                )
                db.session.add(adjustment)
                
                if adj_type == 'add':
                    product.stock += int(adj_qty)
                elif adj_type == 'remove':
                    product.stock -= int(adj_qty)
                elif adj_type == 'correction':
                    product.stock = int(adj_qty)
                
                db.session.commit()
            
            if product.stock < 10000:
                notify_low_stock_alert(product)
            
            flash('Product added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.products'))
    
    return render_template('products/form.html', product=None)

@main_bp.route('/products/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            product.name = request.form.get('name', '')
            product.specification = request.form.get('specification', '')
            product.base_price = Integer(request.form.get('base_price', 0))
            product.stock = int(request.form.get('stock', 0))
            product.color = request.form.get('color', '')
            product.product_type = request.form.get('product_type', 'product')
            product.material_type = request.form.get('material_type', '')
            product.usp_type = request.form.get('usp_type', '')
            product.shape_type = request.form.get('shape_type', '')
            product.dimensions = request.form.get('dimensions', '')
            product.use_case = request.form.get('use_case', '')
            
            db.session.commit()
            
            if product.stock < 10000:
                notify_low_stock_alert(product)
            
            flash('Product updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.products'))
    
    return render_template('products/form.html', product=product)

@main_bp.route('/products/delete/<int:id>')
@login_required
def delete_product(id):
    try:
        product = Product.query.get_or_404(id)
        product.is_deleted = True
        db.session.commit()
        flash('Product archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.products'))

@main_bp.route('/products/stock-adjust/<int:id>', methods=['POST'])
@login_required
def adjust_stock(id):
    product = Product.query.get_or_404(id)
    
    try:
        adj_type = request.form.get('adjustment_type', '')
        qty = int(request.form.get('quantity', 0))
        reason = request.form.get('reason', '')
        
        adjustment = StockAdjustment(
            product_id=product.id, 
            adjustment_type=adj_type, 
            quantity=qty,
            reason=reason, 
            adjusted_by=current_user.username
        )
        db.session.add(adjustment)
        
        if adj_type == 'add':
            product.stock += qty
        elif adj_type == 'remove':
            product.stock -= qty
        elif adj_type == 'correction':
            product.stock = qty
        
        db.session.commit()
        
        if product.stock < 10000:
            notify_low_stock_alert(product)
        
        flash('Stock adjusted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.products'))

# ============================================================================
# INQUIRIES
# ============================================================================

@main_bp.route('/inquiries')
@login_required
def inquiries():
    status = request.args.get('status', 'all')
    show_deleted = request.args.get('show_deleted', 'false') == 'true'
    
    query = Inquiry.query
    if not show_deleted:
        query = query.filter_by(is_deleted=False)
    if status != 'all':
        query = query.filter_by(status=status)
    
    inquiries = query.order_by(Inquiry.created_at.desc()).all()
    return render_template('inquiries/index.html', 
                         inquiries=inquiries, 
                         current_status=status, 
                         show_deleted=show_deleted)

@main_bp.route('/inquiries/add', methods=['GET', 'POST'])
@login_required
def add_inquiry():
    if request.method == 'POST':
        try:
            inquiry_number = f"INQ-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            product_id = request.form.get('product_id')
            
            inquiry = Inquiry(
                inquiry_number=inquiry_number, 
                customer_name=request.form.get('customer_name', ''),
                business_name=request.form.get('business_name', ''), 
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''), 
                product_id=int(product_id) if product_id else None,
                quantity=int(request.form.get('quantity', 0)), 
                notes=request.form.get('notes', ''), 
                status='new'
            )
            db.session.add(inquiry)
            db.session.commit()
            
            notify_new_inquiry(inquiry)
            flash('Inquiry created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.inquiries'))
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('inquiries/form.html', inquiry=None, products=products)

@main_bp.route('/inquiries/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_inquiry(id):
    inquiry = Inquiry.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Update basic info
            inquiry.customer_name = request.form.get('customer_name', '')
            inquiry.business_name = request.form.get('business_name', '')
            inquiry.email = request.form.get('email', '')
            inquiry.phone = request.form.get('phone', '')
            inquiry.status = request.form.get('status', 'new')
            inquiry.notes = request.form.get('notes', '')
            
            # Update quantities for each item
            for item in inquiry.inquiry_items:
                qty_field = f'qty_{item.id}'
                if qty_field in request.form:
                    new_qty = int(request.form.get(qty_field, item.quantity))
                    if new_qty >= 0:
                        item.quantity = new_qty
            
            # Recalculate total quantity
            inquiry.quantity = sum(item.quantity for item in inquiry.inquiry_items)
            
            db.session.commit()
            flash('Inquiry updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.edit_inquiry', id=id))
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('inquiries/edit.html', inquiry=inquiry, products=products)

@main_bp.route('/inquiries/delete/<int:id>')
@login_required
def delete_inquiry(id):
    try:
        inquiry = Inquiry.query.get_or_404(id)
        inquiry.is_deleted = True
        inquiry.deleted_at = datetime.utcnow()
        db.session.commit()
        flash('Inquiry archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.inquiries'))

@main_bp.route('/inquiries/process-invoice/<int:id>', methods=['GET', 'POST'])
@login_required
def process_invoice(id):
    """Convert inquiry to order with stock checking and quantity adjustment"""
    inquiry = Inquiry.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            order_number = f"ORD-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            
            # AUTO-CREATE OR FIND CUSTOMER
            customer, is_new = find_or_create_customer(
                customer_name=inquiry.customer_name,
                business_name=inquiry.business_name,
                email=inquiry.email,
                phone=inquiry.phone
            )
            
            if is_new:
                flash(f'New customer created: {customer.name}', 'info')
            
            # Calculate totals with adjusted quantities
            total_amount = 0
            tax_amount = 0
            stock_warnings = []
            
            # First pass: calculate totals and check stock
            for item in inquiry.inquiry_items:
                # Get adjusted quantity from form
                qty_field = f'qty_{item.id}'
                adjusted_qty = int(request.form.get(qty_field, item.quantity))
                
                # Check stock availability
                if item.product:
                    if adjusted_qty > item.product.stock:
                        stock_warnings.append(f"{item.product.name}: Requested {adjusted_qty}, Available {item.product.stock}. Adjusted to {item.product.stock}.")
                        adjusted_qty = item.product.stock  # Auto-adjust to available stock
                
                unit_price = Integer(request.form.get(f'price_{item.id}', item.product.base_price if item.product else 0))
                tax_rate = float(request.form.get(f'tax_{item.id}', 17))
                
                item_subtotal = adjusted_qty * unit_price
                item_tax = item_subtotal * (tax_rate / 100)
                total_amount += item_subtotal
                tax_amount += item_tax
            
            # Create order
            order = Order(
                order_number=order_number,
                customer_id=customer.id,
                inquiry_id=inquiry.id,
                total_amount=total_amount,
                tax_amount=tax_amount,
                status='pending',
                payment_status='unpaid',
                payment_method=None,
                paid_amount=0,
                customer_name_snapshot=customer.name,
                customer_business_snapshot=customer.business_name,
                customer_phone_snapshot=customer.phone,
                customer_email_snapshot=customer.email
            )
            db.session.add(order)
            db.session.flush()
            
            # Second pass: create order items with adjusted quantities
            for item in inquiry.inquiry_items:
                qty_field = f'qty_{item.id}'
                adjusted_qty = int(request.form.get(qty_field, item.quantity))
                
                # Check stock again and adjust
                if item.product and adjusted_qty > item.product.stock:
                    adjusted_qty = item.product.stock
                
                unit_price = float(request.form.get(f'price_{item.id}', item.product.base_price if item.product else 0))
                
                order_item = OrderItem(
                    order_id=order.id,
                    product_id=item.product_id,
                    quantity=adjusted_qty,
                    unit_price=unit_price,
                    subtotal=adjusted_qty * unit_price
                )
                db.session.add(order_item)
            
            # Update inquiry status
            inquiry.status = 'converted'
            
            # Update customer balance
            customer.current_balance += total_amount
            
            db.session.commit()
            
            # Show stock warnings if any
            if stock_warnings:
                for warning in stock_warnings:
                    flash(f'⚠️ {warning}', 'warning')
            
            notify_new_order(order)
            flash(f'Invoice processed! Order {order_number} created for {customer.name}', 'success')
            return redirect(url_for('main.view_order', id=order.id))
            
        except Exception as e:
            db.session.rollback()
            print(f"ERROR: {str(e)}")
            flash(f'Error processing invoice: {str(e)}', 'error')
    
    # GET request - show the invoice form with stock warnings
    stock_issues = []
    for item in inquiry.inquiry_items:
        if item.product and item.quantity > item.product.stock:
            stock_issues.append({
                'product': item.product.name,
                'requested': item.quantity,
                'available': item.product.stock
            })
    
    return render_template('inquiries/invoice.html', 
                         inquiry=inquiry, 
                         stock_issues=stock_issues)
@main_bp.route('/orders')
@login_required
def orders():
    status = request.args.get('status', 'all')
    query = Order.query
    if status != 'all':
        query = query.filter_by(status=status)
    orders = query.order_by(Order.created_at.desc()).all()
    return render_template('orders/index.html', orders=orders, current_status=status)

@main_bp.route('/orders/add', methods=['GET', 'POST'])
@login_required
def add_order():
    if request.method == 'POST':
        try:
            order_number = f"ORD-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            unit_price = float(request.form.get('unit_price', 0))
            quantity = int(request.form.get('quantity', 0))
            subtotal = unit_price * quantity
            tax_amount = subtotal * 0.17
            total_amount = subtotal + tax_amount
            
            customer_id = request.form.get('customer_id')
            customer = None
            
            # If customer selected, use it
            if customer_id:
                customer = Customer.query.get(int(customer_id))
            else:
                # AUTO-CREATE CUSTOMER FROM ORDER INFO
                customer_name = request.form.get('customer_name', '')
                business_name = request.form.get('business_name', '')
                email = request.form.get('email', '')
                phone = request.form.get('phone', '')
                
                if customer_name or business_name:
                    customer, is_new = find_or_create_customer(
                        customer_name=customer_name,
                        business_name=business_name,
                        email=email,
                        phone=phone
                    )
                    customer_id = customer.id
                    
                    if is_new:
                        flash(f'New customer created: {customer.name}', 'info')
            
            order = Order(
                order_number=order_number,
                customer_id=int(customer_id) if customer_id else None,
                total_amount=total_amount, 
                tax_amount=tax_amount, 
                status='pending',
                payment_status='unpaid', 
                payment_method=request.form.get('payment_method', ''),
                customer_name_snapshot=request.form.get('customer_name', ''),
                customer_business_snapshot=request.form.get('business_name', ''),
                customer_phone_snapshot=request.form.get('phone', ''),
                customer_email_snapshot=request.form.get('email', '')
            )
            db.session.add(order)
            db.session.flush()
            
            product_id = request.form.get('product_id')
            if product_id:
                order_item = OrderItem(
                    order_id=order.id,
                    product_id=int(product_id),
                    quantity=quantity,
                    unit_price=unit_price,
                    subtotal=subtotal
                )
                db.session.add(order_item)
            
            db.session.commit()
            
            # Update customer balance
            if customer:
                customer.current_balance += total_amount
                db.session.commit()
            
            notify_new_order(order)
            flash('Order created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.orders'))
    
    customers = Customer.query.filter_by(is_active=True).all()
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('orders/form.html', customers=customers, products=products)

@main_bp.route('/orders/view/<int:id>')
@login_required
def view_order(id):
    order = Order.query.get_or_404(id)
    
    # Get approval request if exists
    approval_request = OrderApproval.query.filter_by(order_id=id, status='pending').first()
    
    return render_template('orders/view.html', order=order, approval_request=approval_request)

@main_bp.route('/orders/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_order(id):
    order = Order.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            old_payment_status = order.payment_status
            
            order.status = request.form.get('status', order.status)
            order.payment_status = request.form.get('payment_status', order.payment_status)
            order.payment_method = request.form.get('payment_method', '')
            
            if order.payment_status == 'paid':
                order.paid_amount = order.total_amount
                order.payment_date = datetime.utcnow()
                
                # Send payment received email if status changed
                if old_payment_status != 'paid':
                    notify_payment_received(order)
                    
            elif order.payment_status == 'partial':
                paid_amount = float(request.form.get('paid_amount', 0))
                order.paid_amount = paid_amount
                order.payment_date = datetime.utcnow()
            
            db.session.commit()
            
            # Send order status update email
            if order.status == 'completed':
                notify_order_shipped(order)
            
            flash('Order updated successfully!', 'success')
            return redirect(url_for('main.view_order', id=order.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating order: {str(e)}', 'error')
    
    return render_template('orders/edit.html', order=order)

@main_bp.route('/orders/invoice/<int:id>')
@login_required
def order_invoice(id):
    order = Order.query.get_or_404(id)
    logo_setting = CompanySetting.query.filter_by(key='company_logo').first()
    logo_path = logo_setting.value if logo_setting else 'logos/ampoulex-logo.png'
    return render_template('orders/invoice.html', order=order, logo_path=logo_path)

@main_bp.route('/orders/delete/<int:id>')
@login_required
def delete_order(id):
    try:
        order = Order.query.get_or_404(id)
        db.session.delete(order)
        db.session.commit()
        flash('Order deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.orders'))

@main_bp.route('/customers/merge', methods=['GET', 'POST'])
@login_required
def merge_customers():
    """Merge duplicate customers"""
    if request.method == 'POST':
        try:
            primary_id = request.form.get('primary_customer_id')
            merge_ids = request.form.getlist('merge_customer_id[]')
            
            primary_customer = Customer.query.get_or_404(primary_id)
            
            for merge_id in merge_ids:
                merge_customer = Customer.query.get(int(merge_id))
                if merge_customer and merge_customer.id != primary_id:
                    # Transfer orders
                    for order in merge_customer.orders:
                        order.customer_id = primary_id
                    
                    # Transfer balances
                    primary_customer.current_balance += merge_customer.current_balance
                    
                    # Mark as merged/inactive
                    merge_customer.is_active = False
                    merge_customer.is_deleted = True
                    
                    db.session.commit()
            
            flash(f'Merged {len(merge_ids)} customer(s) into {primary_customer.name}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error merging customers: {str(e)}', 'error')
        
        return redirect(url_for('main.customers'))
    
    # Show potential duplicates
    potential_duplicates = db.session.query(
        Customer.business_name,
        db.func.count(Customer.id).label('count')
    ).filter(
        Customer.business_name != None,
        Customer.business_name != '',
        Customer.is_deleted == False
    ).group_by(Customer.business_name).having(
        db.func.count(Customer.id) > 1
    ).all()
    
    return render_template('customers/merge.html', potential_duplicates=potential_duplicates)

# ORDER APPROVAL WORKFLOW (Phase 2 Feature)
@main_bp.route('/orders/request-approval/<int:id>')
@login_required
def request_order_approval(id):
    order = Order.query.get_or_404(id)
    
    # Check if approval already requested
    existing = OrderApproval.query.filter_by(order_id=id, status='pending').first()
    if existing:
        flash('Approval already requested', 'info')
        return redirect(url_for('main.view_order', id=id))
    
    # Create approval request
    approval = OrderApproval(
        order_id=order.id,
        requested_by=current_user.username,
        approval_threshold=50000  # Configurable threshold
    )
    db.session.add(approval)
    db.session.commit()
    
    # Update order status
    order.status = 'pending_approval'
    db.session.commit()
    
    # Notify managers via real-time
    broadcast_update('approval_requested', {
        'order_id': order.id,
        'order_number': order.order_number,
        'amount': float(order.total_amount),
        'requested_by': current_user.username
    })
    
    flash('Approval requested. Waiting for manager approval.', 'info')
    return redirect(url_for('main.view_order', id=id))

@main_bp.route('/orders/approve/<int:id>', methods=['POST'])
@login_required
def approve_order(id):
    # Check if user is manager/admin
    if current_user.role not in ['admin', 'manager']:
        flash('Unauthorized', 'error')
        return redirect(url_for('main.orders'))
    
    approval = OrderApproval.query.get_or_404(id)
    approval.status = 'approved'
    approval.approved_by = current_user.username
    approval.approved_at = datetime.utcnow()
    
    # Update order status
    approval.order.status = 'pending'
    
    db.session.commit()
    flash('Order approved!', 'success')
    return redirect(url_for('main.view_order', id=approval.order.id))

@main_bp.route('/orders/reject/<int:id>', methods=['POST'])
@login_required
def reject_order(id):
    if current_user.role not in ['admin', 'manager']:
        flash('Unauthorized', 'error')
        return redirect(url_for('main.orders'))
    
    approval = OrderApproval.query.get_or_404(id)
    approval.status = 'rejected'
    approval.approved_by = current_user.username
    approval.approved_at = datetime.utcnow()
    approval.rejection_reason = request.form.get('reason', '')
    
    # Update order status
    approval.order.status = 'cancelled'
    
    db.session.commit()
    flash('Order rejected', 'warning')
    return redirect(url_for('main.view_order', id=approval.order.id))

# ============================================================================
# CUSTOMERS
# ============================================================================

@main_bp.route('/customers')
@login_required
def customers():
    show_inactive = request.args.get('show_inactive', 'false') == 'true'
    query = Customer.query
    if not show_inactive:
        query = query.filter_by(is_active=True, is_deleted=False)
    customers = query.all()
    
    for customer in customers:
        total_orders = sum(o.total_amount for o in customer.orders if o.status != 'cancelled')
        total_paid = sum(o.paid_amount for o in customer.orders)
        customer.pending_balance = total_orders - total_paid
    
    return render_template('customers/index.html', customers=customers, show_inactive=show_inactive)

@main_bp.route('/customers/add', methods=['GET', 'POST'])
@login_required
def add_customer():
    if request.method == 'POST':
        try:
            customer = Customer(
                name=request.form.get('name', ''), 
                business_name=request.form.get('business_name', ''),
                role=request.form.get('role', ''), 
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''), 
                address=request.form.get('address', ''),
                payment_terms=request.form.get('payment_terms', 'cash'),
                credit_limit=float(request.form.get('credit_limit', 0))
            )
            db.session.add(customer)
            db.session.commit()
            flash('Customer added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.customers'))
    
    return render_template('customers/form.html', customer=None)

@main_bp.route('/customers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            customer.name = request.form.get('name', '')
            customer.business_name = request.form.get('business_name', '')
            customer.role = request.form.get('role', '')
            customer.email = request.form.get('email', '')
            customer.phone = request.form.get('phone', '')
            customer.address = request.form.get('address', '')
            customer.payment_terms = request.form.get('payment_terms', 'cash')
            customer.is_active = request.form.get('is_active') == 'on'
            customer.credit_limit = float(request.form.get('credit_limit', 0))
            
            db.session.commit()
            flash('Customer updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.customers'))
    
    return render_template('customers/form.html', customer=customer)

@main_bp.route('/customers/delete/<int:id>')
@login_required
def delete_customer(id):
    try:
        customer = Customer.query.get_or_404(id)
        customer.is_deleted = True
        customer.is_active = False
        db.session.commit()
        flash('Customer archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.customers'))

@main_bp.route('/customers/pricing/<int:id>', methods=['GET', 'POST'])
@login_required
def customer_pricing(id):
    customer = Customer.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            product_ids = request.form.getlist('product_id[]')
            prices = request.form.getlist('agreed_price[]')
            terms = request.form.getlist('payment_terms[]')
            
            for i, pid in enumerate(product_ids):
                existing = CustomerProductPrice.query.filter_by(
                    customer_id=customer.id, 
                    product_id=int(pid)
                ).first()
                
                if existing:
                    existing.agreed_price = float(prices[i]) if prices[i] else 0
                    existing.payment_terms = terms[i]
                else:
                    cpp = CustomerProductPrice(
                        customer_id=customer.id, 
                        product_id=int(pid),
                        agreed_price=float(prices[i]) if prices[i] else 0, 
                        payment_terms=terms[i]
                    )
                    db.session.add(cpp)
            
            db.session.commit()
            flash('Pricing updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.customers'))
    
    products = Product.query.filter_by(is_deleted=False).all()
    custom_prices = {cp.product_id: cp for cp in customer.custom_prices}
    return render_template('customers/pricing.html', 
                         customer=customer, 
                         products=products, 
                         custom_prices=custom_prices)

# ============================================================================
# PAYROLL/EMPLOYEES
# ============================================================================

@main_bp.route('/payroll')
@login_required
def payroll():
    employees = Employee.query.all()
    return render_template('payroll/index.html', employees=employees)

@main_bp.route('/employees/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if request.method == 'POST':
        try:
            annual_salary = float(request.form.get('base_salary', 0)) * 12
            monthly_tax, _ = calculate_monthly_tax_deduction(annual_salary)
            
            employee = Employee(
                name=request.form.get('name', ''), 
                role=request.form.get('role', ''),
                department=request.form.get('department', ''), 
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''), 
                cnic=request.form.get('cnic', ''),
                address=request.form.get('address', ''), 
                emergency_contact=request.form.get('emergency_contact', ''),
                emergency_phone=request.form.get('emergency_phone', ''),
                base_salary=float(request.form.get('base_salary', 0)),
                payment_frequency=request.form.get('payment_frequency', 'monthly'),
                bank_account=request.form.get('bank_account', ''), 
                bank_name=request.form.get('bank_name', ''),
                outstanding_loan=float(request.form.get('outstanding_loan', 0))
            )
            db.session.add(employee)
            db.session.commit()
            flash('Employee added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.payroll'))
    
    return render_template('payroll/form.html', employee=None)

@main_bp.route('/employees/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_employee(id):
    employee = Employee.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            employee.name = request.form.get('name', '')
            employee.role = request.form.get('role', '')
            employee.department = request.form.get('department', '')
            employee.email = request.form.get('email', '')
            employee.phone = request.form.get('phone', '')
            employee.cnic = request.form.get('cnic', '')
            employee.address = request.form.get('address', '')
            employee.emergency_contact = request.form.get('emergency_contact', '')
            employee.emergency_phone = request.form.get('emergency_phone', '')
            employee.base_salary = float(request.form.get('base_salary', 0))
            employee.payment_frequency = request.form.get('payment_frequency', 'monthly')
            employee.bank_account = request.form.get('bank_account', '')
            employee.bank_name = request.form.get('bank_name', '')
            employee.outstanding_loan = float(request.form.get('outstanding_loan', 0))
            employee.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash('Employee updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.payroll'))
    
    return render_template('payroll/form.html', employee=employee)

@main_bp.route('/employees/delete/<int:id>')
@login_required
def delete_employee(id):
    try:
        employee = Employee.query.get_or_404(id)
        employee.is_active = False
        db.session.commit()
        flash('Employee deactivated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.payroll'))

# ============================================================================
# EXPENSES
# ============================================================================

@main_bp.route('/expenses')
@login_required
def expenses():
    expenses = Expense.query.order_by(Expense.date.desc()).all()
    total_expenses = db.session.query(db.func.sum(Expense.amount)).scalar() or 0
    pending_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.status == 'pending'
    ).scalar() or 0
    paid_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.status == 'paid'
    ).scalar() or 0
    
    return render_template('expenses/index.html', 
                         expenses=expenses, 
                         total_expenses=total_expenses,
                         pending_expenses=pending_expenses, 
                         paid_expenses=paid_expenses)

@main_bp.route('/expenses/add', methods=['GET', 'POST'])
@login_required
def add_expense():
    if request.method == 'POST':
        try:
            date_str = request.form.get('date')
            if date_str:
                expense_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                expense_date = datetime.utcnow().date()
            
            expense = Expense(
                category=request.form.get('category', ''), 
                description=request.form.get('description', ''),
                amount=float(request.form.get('amount', 0)),
                date=expense_date,
                status=request.form.get('status', 'pending'),
                paid_by=request.form.get('paid_by', ''), 
                payment_method=request.form.get('payment_method', ''),
                reference_number=request.form.get('reference_number', ''), 
                approved_by=request.form.get('approved_by', '')
            )
            db.session.add(expense)
            db.session.commit()
            flash('Expense added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.expenses'))
    
    return render_template('expenses/form.html', expense=None)

@main_bp.route('/expenses/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_expense(id):
    expense = Expense.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            expense.category = request.form.get('category', '')
            expense.description = request.form.get('description', '')
            expense.amount = float(request.form.get('amount', 0))
            
            date_str = request.form.get('date')
            if date_str:
                expense.date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            expense.status = request.form.get('status', 'pending')
            expense.paid_by = request.form.get('paid_by', '')
            expense.payment_method = request.form.get('payment_method', '')
            expense.reference_number = request.form.get('reference_number', '')
            expense.approved_by = request.form.get('approved_by', '')
            
            db.session.commit()
            flash('Expense updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.expenses'))
    
    return render_template('expenses/form.html', expense=expense)

@main_bp.route('/expenses/delete/<int:id>')
@login_required
def delete_expense(id):
    try:
        expense = Expense.query.get_or_404(id)
        db.session.delete(expense)
        db.session.commit()
        flash('Expense deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.expenses'))

# ============================================================================
# ACCOUNTING
# ============================================================================

@main_bp.route('/accounting')
@login_required
def accounting():
    show_inactive = request.args.get('show_inactive', 'false') == 'true'
    
    petty_cash = Accounting.query.filter_by(account_type='petty_cash').all()
    bank = Accounting.query.filter_by(account_type='bank').all()
    receivables = Accounting.query.filter_by(account_type='receivable').all()
    payables = Accounting.query.filter_by(account_type='payable').all()
    
    petty_cash_balance = sum(a.amount if a.transaction_type == 'debit' else -a.amount for a in petty_cash)
    bank_balance = sum(a.amount if a.transaction_type == 'debit' else -a.amount for a in bank)
    receivables_balance = sum(a.amount if a.transaction_type == 'debit' else -a.amount for a in receivables)
    payables_balance = sum(a.amount if a.transaction_type == 'debit' else -a.amount for a in payables)
    
    employee_loans = Employee.query.filter(Employee.outstanding_loan > 0).all()
    
    return render_template('accounting/index.html', 
                         petty_cash_balance=petty_cash_balance,
                         bank_balance=bank_balance, 
                         receivables_balance=receivables_balance,
                         payables_balance=payables_balance, 
                         employee_loans=employee_loans)

# ============================================================================
# ANALYSIS
# ============================================================================

@main_bp.route('/analysis')
@login_required
def analysis():
    total_revenue = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing'])
    ).scalar() or 0
    
    total_orders = Order.query.count()
    total_customers = Customer.query.filter_by(is_active=True).count()
    total_products = Product.query.filter_by(is_deleted=False).count()
    
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    top_products = Product.query.filter_by(is_deleted=False).order_by(Product.stock.desc()).limit(5).all()
    
    return render_template('analysis/index.html', 
                         total_revenue=total_revenue, 
                         total_orders=total_orders,
                         total_customers=total_customers, 
                         total_products=total_products,
                         avg_order_value=avg_order_value, 
                         top_products=top_products,
                         customer_retention=68, 
                         inventory_turnover=75, 
                         growth_rate=12.5)

# ============================================================================
# REPORT GENERATION (EXCEL + PDF)
# ============================================================================

@main_bp.route('/reports')
@login_required
def reports_dashboard():
    """Reports dashboard with available reports"""
    # Get report statistics
    total_reports = Report.query.count()
    recent_reports = Report.query.order_by(Report.generated_at.desc()).limit(10).all()
    
    return render_template('reports/dashboard.html', 
                         total_reports=total_reports,
                         recent_reports=recent_reports)

@main_bp.route('/reports/generate/<module>/excel', methods=['POST'])
@login_required
def generate_excel_report(module):
    """Generate Excel reports - TABLES ONLY, NO CHARTS"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        date_from = request.form.get('date_from', '')
        date_to = request.form.get('date_to', '')
        
        # Set default dates
        if not date_from:
            date_from = '2000-01-01'
        if not date_to:
            date_to = '2099-12-31'
        
        wb = Workbook()
        ws = wb.active
        if not ws:
            flash('Error creating workbook', 'error')
            return redirect(url_for(f'main.{module}'))
        
        ws.title = f"{module.title()} Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"AmpouleX - {module.title()} Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} by {current_user.username}"
        ws['A3'] = f"Period: {date_from} to {date_to}"
        
        # Data based on module - ENHANCED COLUMNS
        row = 5
        headers = []
        data = []
        
        if module == 'inquiries':
            headers = ['Inquiry #', 'Customer', 'Business', 'Email', 'Phone', 'Product', 'Quantity', 'Status', 'Created Date', 'Converted Date']
            query = Inquiry.query.filter(
                Inquiry.created_at >= date_from,
                Inquiry.created_at <= date_to
            )
            for inq in query.all():
                data.append([
                    inq.inquiry_number,
                    inq.customer_name,
                    inq.business_name or '-',
                    inq.email or '-',
                    inq.phone or '-',
                    inq.product.name if inq.product else 'N/A',
                    inq.quantity,
                    inq.status,
                    inq.created_at.strftime('%Y-%m-%d'),
                    inq.converted_at.strftime('%Y-%m-%d') if inq.converted_at else '-'
                ])
        
        elif module == 'orders':
            headers = ['Order #', 'Customer', 'Business', 'Email', 'Phone', 'Products', 'Quantity', 'Subtotal', 'Tax', 'Total (PKR)', 'Status', 'Payment', 'Payment Date', 'Order Date']
            query = Order.query.filter(
                Order.created_at >= date_from,
                Order.created_at <= date_to
            )
            for order in query.all():
                products = ', '.join([
                    f"{item.product.name} x{item.quantity}"
                    for item in order.order_items
                ]) if order.order_items else 'N/A'
                total_qty = sum(item.quantity for item in order.order_items) if order.order_items else 0
                subtotal = order.total_amount - order.tax_amount
                data.append([
                    order.order_number,
                    order.customer_name_snapshot or (order.customer.name if order.customer else 'N/A'),
                    order.customer_business_snapshot or '-',
                    order.customer_email_snapshot or '-',
                    order.customer_phone_snapshot or '-',
                    products,
                    total_qty,
                    f"{subtotal:.2f}",
                    f"{order.tax_amount:.2f}",
                    f"{order.total_amount:.2f}",
                    order.status,
                    order.payment_status,
                    order.payment_date.strftime('%Y-%m-%d') if order.payment_date else '-',
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif module == 'customers':
            headers = ['Name', 'Business', 'Role', 'Email', 'Phone', 'Address', 'Payment Terms', 'Credit Limit (PKR)', 'Current Balance (PKR)', 'Total Orders', 'Total Spent (PKR)', 'Status']
            for customer in Customer.query.all():
                total_orders = len(customer.orders)
                total_spent = sum(o.total_amount for o in customer.orders)
                data.append([
                    customer.name,
                    customer.business_name or '-',
                    customer.role or '-',
                    customer.email or '-',
                    customer.phone or '-',
                    customer.address or '-',
                    customer.payment_terms,
                    f"{customer.credit_limit:.2f}",
                    f"{customer.current_balance:.2f}",
                    total_orders,
                    f"{total_spent:.2f}",
                    'Active' if customer.is_active else 'Inactive'
                ])
        
        elif module == 'products':
            headers = ['ID', 'Name', 'Specification', 'Type', 'Color', 'Material', 'USP Type', 'Shape', 'Dimensions', 'Price (PKR)', 'Stock', 'Stock Value (PKR)', 'Status']
            for product in Product.query.filter_by(is_deleted=False).all():
                stock_value = product.stock * product.base_price
                stock_status = 'Low Stock' if product.stock < 10000 else ('Out of Stock' if product.stock == 0 else 'In Stock')
                data.append([
                    product.id,
                    product.name,
                    product.specification or '-',
                    product.product_type,
                    product.color,
                    product.material_type or '-',
                    product.usp_type or '-',
                    product.shape_type or '-',
                    product.dimensions or '-',
                    f"{product.base_price:.2f}",
                    product.stock,
                    f"{stock_value:.2f}",
                    stock_status
                ])
        
        elif module == 'expenses':
            headers = ['Date', 'Category', 'Description', 'Amount (PKR)', 'Paid By', 'Payment Method', 'Reference #', 'Approved By', 'Status']
            query = Expense.query.filter(
                Expense.date >= date_from,
                Expense.date <= date_to
            )
            for exp in query.all():
                data.append([
                    exp.date.strftime('%Y-%m-%d'),
                    exp.category,
                    exp.description,
                    f"{exp.amount:.2f}",
                    exp.paid_by or '-',
                    exp.payment_method or '-',
                    exp.reference_number or '-',
                    exp.approved_by or '-',
                    exp.status
                ])
        
        elif module == 'payroll':
            headers = ['Employee', 'Role', 'Department', 'CNIC', 'Email', 'Phone', 'Base Salary (PKR)', 'Bank Name', 'Bank Account', 'Outstanding Loan (PKR)', 'Status']
            for emp in Employee.query.all():
                data.append([
                    emp.name,
                    emp.role or '-',
                    emp.department or '-',
                    emp.cnic or '-',
                    emp.email or '-',
                    emp.phone or '-',
                    f"{emp.base_salary:.2f}",
                    emp.bank_name or '-',
                    emp.bank_account or '-',
                    f"{emp.outstanding_loan:.2f}",
                    'Active' if emp.is_active else 'Inactive'
                ])
        
        elif module == 'sales_analysis':
            group_by = request.form.get('group_by', 'product')
            if group_by == 'product':
                headers = ['Product', 'Quantity Sold', 'Revenue (PKR)', 'Orders', '% of Total']
            elif group_by == 'customer':
                headers = ['Customer', 'Orders', 'Revenue (PKR)', 'Avg Order (PKR)', '% of Total']
            elif group_by == 'date':
                headers = ['Date', 'Orders', 'Revenue (PKR)', '% of Total']
            
            # Get data from session or recalculate
            # (Implementation depends on how you pass data)
            data = []  # Populate based on sales_analysis_report logic
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row += 1
        
        # Write data
        for record in data:
            for col, value in enumerate(record, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
        
        # Auto-adjust column widths
        if ws.columns:
            for column in ws.columns:
                if column and len(column) > 0:
                    first_cell = column[0]
                    if hasattr(first_cell, 'column_letter'):
                        column_letter = first_cell.column_letter
                        max_length = 0
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = f"{module}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        filepath = os.path.join(reports_dir, filename)
        wb.save(filepath)
        
        # Save report record
        report = Report(
            report_type='excel',
            module=module,
            generated_by=current_user.username,
            parameters=json.dumps({'date_from': date_from, 'date_to': date_to}),
            file_path=filepath
        )
        db.session.add(report)
        db.session.commit()
        
        flash(f'Excel report generated: {filename}', 'success')
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')
        return redirect(url_for(f'main.{module}'))

@main_bp.route('/reports/generate/<module>/pdf', methods=['POST'])
@login_required
def generate_pdf_report(module):
    """Generate PDF reports - TABLES ONLY, NO CHARTS"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        
        date_from = request.form.get('date_from', '')
        date_to = request.form.get('date_to', '')
        
        # Set default dates
        if not date_from:
            date_from = '2000-01-01'
        if not date_to:
            date_to = '2099-12-31'
        
        # Create PDF
        filename = f"{module}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
        reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        filepath = os.path.join(reports_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12
        )
        
        # Title
        elements.append(Paragraph(f"AmpouleX - {module.title()} Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Metadata
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} by {current_user.username}", styles['Normal']))
        elements.append(Paragraph(f"Period: {date_from} to {date_to}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Data based on module - ENHANCED COLUMNS
        data = []
        headers = []
        
        if module == 'inquiries':
            headers = ['Inquiry #', 'Customer', 'Business', 'Product', 'Qty', 'Status', 'Date']
            query = Inquiry.query.filter(
                Inquiry.created_at >= date_from,
                Inquiry.created_at <= date_to
            )
            for inq in query.all():
                data.append([
                    inq.inquiry_number,
                    inq.customer_name,
                    inq.business_name or '-',
                    inq.product.name if inq.product else 'N/A',
                    str(inq.quantity),
                    inq.status,
                    inq.created_at.strftime('%Y-%m-%d')
                ])
        
        elif module == 'orders':
            headers = ['Order #', 'Customer', 'Total (PKR)', 'Status', 'Payment', 'Date']
            query = Order.query.filter(
                Order.created_at >= date_from,
                Order.created_at <= date_to
            )
            for order in query.all():
                data.append([
                    order.order_number,
                    order.customer_name_snapshot or (order.customer.name if order.customer else 'N/A'),
                    f"PKR {order.total_amount:.2f}",
                    order.status,
                    order.payment_status,
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif module == 'customers':
            headers = ['Name', 'Business', 'Email', 'Phone', 'Balance (PKR)', 'Status']
            for customer in Customer.query.all():
                data.append([
                    customer.name,
                    customer.business_name or '-',
                    customer.email or '-',
                    customer.phone or '-',
                    f"PKR {customer.current_balance:.2f}",
                    'Active' if customer.is_active else 'Inactive'
                ])
        
        elif module == 'products':
            headers = ['Name', 'Specification', 'Color', 'Price (PKR)', 'Stock', 'Status']
            for product in Product.query.filter_by(is_deleted=False).all():
                stock_status = 'Low' if product.stock < 10000 else 'OK'
                data.append([
                    product.name,
                    product.specification or '-',
                    product.color,
                    f"PKR {product.base_price:.2f}",
                    str(product.stock),
                    stock_status
                ])
        
        elif module == 'expenses':
            headers = ['Date', 'Category', 'Description', 'Amount (PKR)', 'Status']
            query = Expense.query.filter(
                Expense.date >= date_from,
                Expense.date <= date_to
            )
            for exp in query.all():
                data.append([
                    exp.date.strftime('%Y-%m-%d'),
                    exp.category,
                    exp.description,
                    f"PKR {exp.amount:.2f}",
                    exp.status
                ])
        
        elif module == 'payroll':
            headers = ['Employee', 'Role', 'Department', 'Salary (PKR)', 'Loan (PKR)', 'Status']
            for emp in Employee.query.all():
                data.append([
                    emp.name,
                    emp.role or '-',
                    emp.department or '-',
                    f"PKR {emp.base_salary:.2f}",
                    f"PKR {emp.outstanding_loan:.2f}",
                    'Active' if emp.is_active else 'Inactive'
                ])
        
        # Create table
        if headers and data:
            table_data = [headers] + data
            table = Table(table_data, colWidths=[1.2*inch] * len(headers))
            
            # Table style
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Save report record
        report = Report(
            report_type='pdf',
            module=module,
            generated_by=current_user.username,
            parameters=json.dumps({'date_from': date_from, 'date_to': date_to}),
            file_path=filepath
        )
        db.session.add(report)
        db.session.commit()
        
        flash(f'PDF report generated: {filename}', 'success')
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF report: {str(e)}', 'error')
        return redirect(url_for(f'main.{module}'))


# SETTINGS
# ============================================================================

@main_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        try:
            settings_data = {
                'company_name': request.form.get('company_name', ''),
                'company_ntn': request.form.get('company_ntn', ''),
                'company_gst': request.form.get('company_gst', ''),
                'company_secp': request.form.get('company_secp', ''),
                'company_drap': request.form.get('company_drap', ''),
                'company_iso': request.form.get('company_iso', ''),
                'company_address': request.form.get('company_address', ''),
                'company_phone': request.form.get('company_phone', ''),
                'company_email': request.form.get('company_email', ''),
                'company_website': request.form.get('company_website', ''),
                'invoice_footer': request.form.get('invoice_footer', '')
            }
            
            for key, value in settings_data.items():
                setting = CompanySetting.query.filter_by(key=key).first()
                if setting:
                    setting.value = value
                else:
                    setting = CompanySetting(key=key, value=value)
                    db.session.add(setting)
            
            db.session.commit()
            flash('Settings updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.settings'))
    
    settings = {s.key: s.value for s in CompanySetting.query.all()}
    return render_template('settings/index.html', settings=settings)

@main_bp.route('/settings/upload-logo', methods=['POST'])
@login_required
def upload_logo():
    if 'logo' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('main.settings'))
    
    file = request.files['logo']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('main.settings'))
    
    if file:
        filename = "company_logo.png"
        filepath = os.path.join('static', 'logos', filename)
        file.save(filepath)
        
        setting = CompanySetting.query.filter_by(key='company_logo').first()
        if setting:
            setting.value = filepath
        else:
            setting = CompanySetting(key='company_logo', value=filepath)
            db.session.add(setting)
        
        db.session.commit()
        flash('Logo uploaded successfully!', 'success')
    
    return redirect(url_for('main.settings'))

# ============================================================================
# CUSTOMER-FACING WEBSITE
# ============================================================================

@main_bp.route('/customer-site')
def customer_site():
    grouped_products = group_products_by_base()
    return render_template('customer-site.html', products=grouped_products)

@main_bp.route('/submit-inquiry', methods=['POST'])
def submit_inquiry():
    try:
        inquiry_number = f"INQ-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        product_ids = request.form.getlist('product_ids')
        
        customer_name = request.form.get('customer_name', '')
        business_name = request.form.get('business_name', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        
        # AUTO-CREATE CUSTOMER FROM WEBSITE INQUIRY (NEW!)
        customer, is_new = find_or_create_customer(
            customer_name=customer_name,
            business_name=business_name,
            email=email,
            phone=phone
        )
        
        inquiry = Inquiry(
            inquiry_number=inquiry_number,
            customer_name=customer_name,
            business_name=business_name,
            email=email,
            phone=phone,
            product_id=None,
            quantity=0,
            notes=request.form.get('notes', ''),
            status='new'
        )
        db.session.add(inquiry)
        db.session.flush()
        
        total_quantity = 0
        products_list = []
        
        for pid in product_ids:
            qty = request.form.get(f'qty_{pid}', '1')
            qty = int(qty) if qty and qty.isdigit() and int(qty) > 0 else 1
            total_quantity += qty
            
            item = InquiryItem(
                inquiry_id=inquiry.id,
                product_id=int(pid),
                quantity=qty
            )
            db.session.add(item)
            
            product = Product.query.get(int(pid))
            if product:
                products_list.append(f"{product.name} ({product.color})")
        
        inquiry.quantity = total_quantity
        db.session.commit()
        
        try:
            socketio.emit('new_inquiry', {
                'id': inquiry.id,
                'number': inquiry.inquiry_number,
                'customer': customer_name,
                'business': business_name,
                'products': ', '.join(products_list),
                'quantity': total_quantity,
                'status': inquiry.status,
                'created_at': inquiry.created_at.isoformat(),
                'customer_id': customer.id,  # Include customer ID
                'is_new_customer': is_new
            }, namespace='/admin')
        except Exception as e:
            print(f"Socket.IO error: {e}")
        
        flash(f'Thank you! Your inquiry #{inquiry_number} has been submitted. We will contact you within 24 hours.', 'success')
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {str(e)}")
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.customer_site'))
# ============================================================================
# API ENDPOINTS (PHASE 2)
# ============================================================================

@main_bp.route('/api/inquiry-count')
@login_required
def inquiry_count():
    count = Inquiry.query.filter(
        Inquiry.status.in_(['new', 'followup']), 
        Inquiry.is_deleted == False
    ).count()
    return jsonify({'count': count})

@main_bp.route('/api/revenue-data')
@login_required
def revenue_data():
    six_months_ago = datetime.utcnow() - timedelta(days=180)
    monthly_revenue = db.session.query(
        db.func.strftime('%Y-%m', Order.created_at),
        db.func.sum(Order.total_amount)
    ).filter(
        Order.created_at >= six_months_ago, 
        Order.status.in_(['completed', 'processing'])
    ).group_by(db.func.strftime('%Y-%m', Order.created_at)).all()
    
    return jsonify({
        'labels': [r[0] for r in monthly_revenue], 
        'data': [float(r[1]) for r in monthly_revenue]
    })

@main_bp.route('/api/low-stock-count')
@login_required
def low_stock_count():
    count = Product.query.filter(
        Product.stock < 10000,
        Product.is_deleted == False
    ).count()
    return jsonify({'count': count})

@main_bp.route('/api/overdue-orders-count')
@login_required
def overdue_orders_count():
    count = Order.query.filter(
        Order.payment_status == 'unpaid',
        Order.created_at < datetime.utcnow() - timedelta(days=30)
    ).count()
    return jsonify({'count': count})

@main_bp.route('/api/pending-orders-count')
@login_required
def pending_orders_count():
    count = Order.query.filter_by(status='pending').count()
    return jsonify({'count': count})

@main_bp.route('/api/order-status-data')
@login_required
def order_status_data():
    status_counts = db.session.query(
        Order.status, 
        db.func.count(Order.id)
    ).group_by(Order.status).all()
    
    labels = [s[0] or 'Unknown' for s in status_counts]
    data = [s[1] for s in status_counts]
    
    return jsonify({
        'labels': labels,
        'data': data
    })

# ============================================================================
# ATTENDANCE TRACKING
# ============================================================================

@main_bp.route('/attendance')
@login_required
def attendance():
    date = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
    employees = Employee.query.filter_by(is_active=True).all()
    attendance_records = {a.employee_id: a for a in Attendance.query.filter_by(date=date).all()}
    
    return render_template('payroll/attendance.html',
                         employees=employees,
                         attendance_records=attendance_records,
                         selected_date=date)

@main_bp.route('/attendance/mark', methods=['POST'])
@login_required
def mark_attendance():
    try:
        employee_id = int(request.form.get('employee_id', 0))
        date_str = request.form.get('date', '')
        date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.utcnow().date()
        status = request.form.get('status', 'present')
        
        check_in_str = request.form.get('check_in', '')
        check_out_str = request.form.get('check_out', '')
        check_in = datetime.strptime(check_in_str, '%H:%M') if check_in_str else None
        check_out = datetime.strptime(check_out_str, '%H:%M') if check_out_str else None
        
        hours_worked = 0
        overtime_hours = 0
        
        if check_in and check_out:
            check_in_dt = datetime.combine(date, check_in.time())
            check_out_dt = datetime.combine(date, check_out.time())
            hours_worked = (check_out_dt - check_in_dt).total_seconds() / 3600
            if hours_worked > 6:
                hours_worked -= 1
            overtime_hours = max(0, hours_worked - 8)
            hours_worked = min(hours_worked, 8)
        
        attendance = Attendance.query.filter_by(employee_id=employee_id, date=date).first()
        if not attendance:
            attendance = Attendance(employee_id=employee_id, date=date)
            db.session.add(attendance)
        
        attendance.status = status
        attendance.check_in = check_in
        attendance.check_out = check_out
        attendance.hours_worked = hours_worked
        attendance.overtime_hours = overtime_hours
        attendance.notes = request.form.get('notes', '')
        
        db.session.commit()
        notify_attendance_marked(attendance)
        flash('Attendance marked successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.attendance', date=date_str))

@main_bp.route('/attendance/history')
@login_required
def attendance_history():
    employee_id = request.args.get('employee_id', type=int)
    date_from = request.args.get('date_from', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.utcnow().strftime('%Y-%m-%d'))
    
    query = Attendance.query.filter(
        Attendance.date >= date_from,
        Attendance.date <= date_to
    )
    
    if employee_id:
        query = query.filter_by(employee_id=employee_id)
    
    attendance_records = query.order_by(Attendance.date.desc()).all()
    employees = Employee.query.filter_by(is_active=True).all()
    
    return render_template('payroll/attendance_history.html',
                         attendance_records=attendance_records,
                         employees=employees,
                         selected_employee=employee_id,
                         date_from=date_from,
                         date_to=date_to)

# ============================================================================
# TIMESHEET MANAGEMENT
# ============================================================================

@main_bp.route('/timesheets')
@login_required
def timesheets():
    status = request.args.get('status', 'all')
    query = Timesheet.query
    if status != 'all':
        query = query.filter_by(status=status)
    timesheets = query.order_by(Timesheet.created_at.desc()).all()
    return render_template('payroll/timesheets.html', timesheets=timesheets, current_status=status)

@main_bp.route('/timesheets/generate', methods=['GET', 'POST'])
@login_required
def generate_timesheet():
    if request.method == 'POST':
        try:
            employee_id = int(request.form.get('employee_id', 0))
            
            period_start_str = request.form.get('period_start', '')
            period_end_str = request.form.get('period_end', '')
            
            period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date() if period_start_str else datetime.utcnow().date()
            period_end = datetime.strptime(period_end_str, '%Y-%m-%d').date() if period_end_str else datetime.utcnow().date()
            
            period_type = request.form.get('period_type', 'weekly')
            
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee_id,
                Attendance.date >= period_start,
                Attendance.date <= period_end,
                Attendance.is_approved == True
            ).all()
            
            regular_hours = sum(a.hours_worked for a in attendance_records if a.status == 'present')
            overtime_hours = sum(a.overtime_hours for a in attendance_records)
            leave_days = sum(1 for a in attendance_records if a.status == 'leave')
            sick_days = sum(1 for a in attendance_records if a.status == 'sick')
            absent_days = sum(1 for a in attendance_records if a.status == 'absent')
            
            employee = Employee.query.get(employee_id)
            
            hourly_rate = employee.base_salary / 176 if employee else 0
            overtime_rate = hourly_rate * 1.5
            
            regular_pay = regular_hours * hourly_rate
            overtime_pay = overtime_hours * overtime_rate
            
            timesheet = Timesheet(
                employee_id=employee_id,
                period_start=period_start,
                period_end=period_end,
                period_type=period_type,
                regular_hours=regular_hours,
                overtime_hours=overtime_hours,
                total_hours=regular_hours + overtime_hours,
                leave_days=leave_days,
                sick_days=sick_days,
                absent_days=absent_days,
                hourly_rate=hourly_rate,
                overtime_rate=overtime_rate,
                regular_pay=regular_pay,
                overtime_pay=overtime_pay,
                deductions=0,
                net_pay=regular_pay + overtime_pay
            )
            db.session.add(timesheet)
            db.session.commit()
            
            flash('Timesheet generated successfully!', 'success')
            return redirect(url_for('main.timesheets'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    employees = Employee.query.filter_by(is_active=True).all()
    return render_template('payroll/timesheet_form.html', employees=employees, timesheet=None)

@main_bp.route('/timesheets/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_timesheet(id):
    timesheet = Timesheet.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            timesheet.regular_hours = float(request.form.get('regular_hours', 0))
            timesheet.overtime_hours = float(request.form.get('overtime_hours', 0))
            timesheet.total_hours = timesheet.regular_hours + timesheet.overtime_hours
            timesheet.leave_days = float(request.form.get('leave_days', 0))
            timesheet.sick_days = float(request.form.get('sick_days', 0))
            timesheet.notes = request.form.get('notes', '')
            
            employee = Employee.query.get(timesheet.employee_id)
            hourly_rate = employee.base_salary / 176 if employee else 0
            overtime_rate = hourly_rate * 1.5
            
            timesheet.hourly_rate = hourly_rate
            timesheet.overtime_rate = overtime_rate
            timesheet.regular_pay = timesheet.regular_hours * hourly_rate
            timesheet.overtime_pay = timesheet.overtime_hours * overtime_rate
            timesheet.net_pay = timesheet.regular_pay + timesheet.overtime_pay
            
            db.session.commit()
            flash('Timesheet updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.timesheets'))
    
    return render_template('payroll/timesheet_form.html', employees=Employee.query.all(), timesheet=timesheet)

@main_bp.route('/timesheets/approve/<int:id>')
@login_required
def approve_timesheet(id):
    timesheet = Timesheet.query.get_or_404(id)
    timesheet.status = 'approved'
    timesheet.approved_by = current_user.username
    timesheet.approved_at = datetime.utcnow()
    db.session.commit()
    flash('Timesheet approved!', 'success')
    return redirect(url_for('main.timesheets'))

@main_bp.route('/timesheets/delete/<int:id>')
@login_required
def delete_timesheet(id):
    timesheet = Timesheet.query.get_or_404(id)
    db.session.delete(timesheet)
    db.session.commit()
    flash('Timesheet deleted!', 'success')
    return redirect(url_for('main.timesheets'))

# ============================================================================
# LEAVE MANAGEMENT
# ============================================================================

@main_bp.route('/leave-requests')
@login_required
def leave_requests():
    status = request.args.get('status', 'all')
    query = LeaveRequest.query
    if status != 'all':
        query = query.filter_by(status=status)
    leave_requests_list = query.order_by(LeaveRequest.created_at.desc()).all()
    return render_template('payroll/leave_requests.html', leave_requests=leave_requests_list, current_status=status)

@main_bp.route('/leave-requests/add', methods=['GET', 'POST'])
@login_required
def add_leave_request():
    if request.method == 'POST':
        try:
            employee_id = int(request.form.get('employee_id', 0))
            
            start_date_str = request.form.get('start_date', '')
            end_date_str = request.form.get('end_date', '')
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else datetime.utcnow().date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else datetime.utcnow().date()
            
            total_days = (end_date - start_date).days + 1
            
            leave_request = LeaveRequest(
                employee_id=employee_id,
                leave_type=request.form.get('leave_type', ''),
                start_date=start_date,
                end_date=end_date,
                total_days=total_days,
                reason=request.form.get('reason', '')
            )
            db.session.add(leave_request)
            db.session.commit()
            flash('Leave request submitted!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.leave_requests'))
    
    employees = Employee.query.filter_by(is_active=True).all()
    return render_template('payroll/leave_form.html', employees=employees, leave_request=None)

@main_bp.route('/leave-requests/approve/<int:id>')
@login_required
def approve_leave(id):
    leave_request = LeaveRequest.query.get_or_404(id)
    leave_request.status = 'approved'
    leave_request.approved_by = current_user.username
    leave_request.approved_at = datetime.utcnow()
    db.session.commit()
    flash('Leave request approved!', 'success')
    return redirect(url_for('main.leave_requests'))

@main_bp.route('/leave-requests/reject/<int:id>', methods=['POST'])
@login_required
def reject_leave(id):
    leave_request = LeaveRequest.query.get_or_404(id)
    leave_request.status = 'rejected'
    leave_request.rejection_reason = request.form.get('reason', '')
    leave_request.approved_by = current_user.username
    leave_request.approved_at = datetime.utcnow()
    db.session.commit()
    flash('Leave request rejected!', 'warning')
    return redirect(url_for('main.leave_requests'))

# ============================================================================
# PAYROLL PAYMENTS
# ============================================================================

@main_bp.route('/payroll-payments')
@login_required
def payroll_payments():
    status = request.args.get('status', 'all')
    query = PayrollPayment.query
    if status != 'all':
        query = query.filter_by(status=status)
    payments = query.order_by(PayrollPayment.created_at.desc()).all()
    return render_template('payroll/payments.html', payments=payments, current_status=status)

@main_bp.route('/payroll-payments/process', methods=['GET', 'POST'])
@login_required
def process_payment():
    if request.method == 'POST':
        try:
            timesheet_id = int(request.form.get('timesheet_id', 0))
            timesheet = Timesheet.query.get_or_404(timesheet_id)
            employee = Employee.query.get(timesheet.employee_id)
            
            annual_salary = employee.base_salary * 12
            monthly_tax, _ = calculate_monthly_tax_deduction(annual_salary)
            
            loan_deduction = min(employee.outstanding_loan, timesheet.net_pay * 0.1)
            total_deductions = monthly_tax + loan_deduction
            net_pay = timesheet.net_pay - total_deductions
            
            payment = PayrollPayment(
                employee_id=employee.id,
                timesheet_id=timesheet_id,
                period_start=timesheet.period_start,
                period_end=timesheet.period_end,
                base_salary=employee.base_salary,
                hourly_wages=timesheet.regular_pay,
                overtime_pay=timesheet.overtime_pay,
                bonuses=float(request.form.get('bonuses', 0)),
                allowances=float(request.form.get('allowances', 0)),
                total_earnings=timesheet.net_pay + float(request.form.get('bonuses', 0)) + float(request.form.get('allowances', 0)),
                income_tax=monthly_tax,
                loan_deduction=loan_deduction,
                other_deductions=float(request.form.get('other_deductions', 0)),
                total_deductions=total_deductions + float(request.form.get('other_deductions', 0)),
                net_pay=net_pay,
                payment_method=request.form.get('payment_method', 'bank_transfer'),
                payment_date=datetime.utcnow().date(),
                bank_name=employee.bank_name,
                bank_account=employee.bank_account,
                status='processed',
                processed_by=current_user.username,
                processed_at=datetime.utcnow()
            )
            db.session.add(payment)
            
            if loan_deduction > 0:
                employee.outstanding_loan -= loan_deduction
            
            timesheet.status = 'paid'
            timesheet.paid_at = datetime.utcnow()
            
            db.session.commit()
            flash('Payment processed successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.payroll_payments'))
    
    timesheets = Timesheet.query.filter_by(status='approved').all()
    return render_template('payroll/payment_form.html', timesheets=timesheets)



# ============================================================================
# FINANCIAL STATEMENTS (PHASE 4)
# ============================================================================

@main_bp.route('/financials/profit-loss')
@login_required
def profit_loss():
    """Generate Profit & Loss Statement"""
    # Get date range from query params
    start_date = request.args.get('start_date', datetime.utcnow().replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    # Calculate Revenue
    revenue_query = Order.query.filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= start_date,
        Order.created_at <= end_date
    )
    total_revenue = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= start_date,
        Order.created_at <= end_date
    ).scalar() or 0
    
    # Calculate Cost of Goods Sold (COGS) - Estimated 60% of revenue for ampoules
    cogs = total_revenue * 0.60
    
    # Calculate Gross Profit
    gross_profit = total_revenue - cogs
    
    # Calculate Operating Expenses
    expenses_query = Expense.query.filter(
        Expense.date >= start_date,
        Expense.date <= end_date
    )
    total_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.date >= start_date,
        Expense.date <= end_date
    ).scalar() or 0
    
    # Calculate Operating Profit
    operating_profit = gross_profit - total_expenses
    
    # Calculate Payroll Expenses
    payroll_query = PayrollPayment.query.filter(
        PayrollPayment.payment_date >= start_date,
        PayrollPayment.payment_date <= end_date
    )
    total_payroll = db.session.query(db.func.sum(PayrollPayment.net_pay)).filter(
        PayrollPayment.payment_date >= start_date,
        PayrollPayment.payment_date <= end_date
    ).scalar() or 0
    
    # Calculate Net Profit
    net_profit = operating_profit - total_payroll
    
    # Expense breakdown by category
    expenses_by_category = db.session.query(
        Expense.category,
        db.func.sum(Expense.amount)
    ).filter(
        Expense.date >= start_date,
        Expense.date <= end_date
    ).group_by(Expense.category).all()
    
    # Monthly comparison (last 6 months)
    six_months = []
    for i in range(6):
        month_end = datetime.utcnow() - timedelta(days=30*i)
        month_start = month_end - timedelta(days=30)
        month_revenue = db.session.query(db.func.sum(Order.total_amount)).filter(
            Order.status.in_(['completed', 'processing']),
            Order.created_at >= month_start,
            Order.created_at <= month_end
        ).scalar() or 0
        six_months.append({
            'month': month_end.strftime('%b %Y'),
            'revenue': month_revenue
        })
    
    return render_template('financials/profit_loss.html',
                         start_date=start_date,
                         end_date=end_date,
                         total_revenue=total_revenue,
                         cogs=cogs,
                         gross_profit=gross_profit,
                         total_expenses=total_expenses,
                         operating_profit=operating_profit,
                         total_payroll=total_payroll,
                         net_profit=net_profit,
                         expenses_by_category=expenses_by_category,
                         six_months=six_months)

@main_bp.route('/financials/balance-sheet')
@login_required
def balance_sheet():
    """Generate Balance Sheet"""
    # ASSETS
    # Current Assets
    cash_in_bank = Accounting.query.filter_by(account_type='bank').all()
    bank_balance = sum(a.amount if a.transaction_type == 'debit' else -a.amount for a in cash_in_bank)
    
    petty_cash = Accounting.query.filter_by(account_type='petty_cash').all()
    petty_cash_balance = sum(a.amount if a.transaction_type == 'debit' else -a.amount for a in petty_cash)
    
    # Accounts Receivable (unpaid orders)
    receivables = Order.query.filter_by(payment_status='unpaid').all()
    accounts_receivable = sum(o.total_amount for o in receivables)
    
    # Inventory Value (products * base_price * 0.6 for cost)
    products = Product.query.filter_by(is_deleted=False).all()
    inventory_value = sum(p.stock * p.base_price * 0.6 for p in products)
    
    # Raw Materials Inventory
    raw_materials = RawMaterial.query.filter_by(is_active=True).all()
    raw_material_value = sum(m.current_stock * m.cost_per_unit for m in raw_materials)
    
    # Total Current Assets
    total_current_assets = bank_balance + petty_cash_balance + accounts_receivable + inventory_value + raw_material_value
    
    # Fixed Assets (simplified - would need fixed assets table)
    fixed_assets = 0  # Add when fixed assets module is implemented
    
    # Total Assets
    total_assets = total_current_assets + fixed_assets
    
    # LIABILITIES
    # Current Liabilities
    payables = Accounting.query.filter_by(account_type='payable').all()
    accounts_payable = sum(a.amount if a.transaction_type == 'credit' else -a.amount for a in payables)
    
    # Employee Loans Outstanding
    employee_loans = Employee.query.filter(Employee.outstanding_loan > 0).all()
    total_employee_loans = sum(e.outstanding_loan for e in employee_loans)
    
    # Tax Payable (GST)
    tax_payable = total_revenue * 0.17 if 'total_revenue' in locals() else 0
    
    # Total Current Liabilities
    total_current_liabilities = accounts_payable + total_employee_loans + tax_payable
    
    # Long-term Liabilities
    long_term_liabilities = 0  # Add when loans module is implemented
    
    # Total Liabilities
    total_liabilities = total_current_liabilities + long_term_liabilities
    
    # EQUITY
    # Retained Earnings (simplified calculation)
    equity = total_assets - total_liabilities
    
    return render_template('financials/balance_sheet.html',
                         bank_balance=bank_balance,
                         petty_cash_balance=petty_cash_balance,
                         accounts_receivable=accounts_receivable,
                         inventory_value=inventory_value,
                         raw_material_value=raw_material_value,
                         total_current_assets=total_current_assets,
                         fixed_assets=fixed_assets,
                         total_assets=total_assets,
                         accounts_payable=accounts_payable,
                         total_employee_loans=total_employee_loans,
                         tax_payable=tax_payable,
                         total_current_liabilities=total_current_liabilities,
                         long_term_liabilities=long_term_liabilities,
                         total_liabilities=total_liabilities,
                         equity=equity)

@main_bp.route('/financials/cash-flow')
@login_required
def cash_flow():
    """Generate Cash Flow Statement"""
    start_date = request.args.get('start_date', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    # OPERATING ACTIVITIES
    # Cash from Sales (paid orders)
    cash_from_sales = db.session.query(db.func.sum(Order.paid_amount)).filter(
        Order.payment_status == 'paid',
        Order.payment_date >= start_date,
        Order.payment_date <= end_date
    ).scalar() or 0
    
    # Cash from Expenses
    cash_from_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.status == 'paid',
        Expense.date >= start_date,
        Expense.date <= end_date
    ).scalar() or 0
    
    # Cash from Payroll
    cash_from_payroll = db.session.query(db.func.sum(PayrollPayment.net_pay)).filter(
        PayrollPayment.payment_date >= start_date,
        PayrollPayment.payment_date <= end_date
    ).scalar() or 0
    
    net_cash_from_operations = cash_from_sales - cash_from_expenses - cash_from_payroll
    
    # INVESTING ACTIVITIES
    # Equipment purchases (would need fixed assets table)
    cash_from_investing = 0
    
    net_cash_from_investing = -cash_from_investing
    
    # FINANCING ACTIVITIES
    # Loan disbursements
    cash_from_loans = 0  # Would need loans table
    
    # Loan repayments
    cash_to_loans = db.session.query(db.func.sum(PayrollPayment.loan_deduction)).filter(
        PayrollPayment.payment_date >= start_date,
        PayrollPayment.payment_date <= end_date
    ).scalar() or 0
    
    net_cash_from_financing = cash_from_loans - cash_to_loans
    
    # Net Change in Cash
    net_change_in_cash = net_cash_from_operations + net_cash_from_investing + net_cash_from_financing
    
    # Opening Cash Balance
    opening_cash = bank_balance + petty_cash_balance if 'bank_balance' in locals() else 0
    
    # Closing Cash Balance
    closing_cash = opening_cash + net_change_in_cash
    
    return render_template('financials/cash_flow.html',
                         start_date=start_date,
                         end_date=end_date,
                         cash_from_sales=cash_from_sales,
                         cash_from_expenses=cash_from_expenses,
                         cash_from_payroll=cash_from_payroll,
                         net_cash_from_operations=net_cash_from_operations,
                         cash_from_investing=cash_from_investing,
                         net_cash_from_investing=net_cash_from_investing,
                         cash_from_loans=cash_from_loans,
                         cash_to_loans=cash_to_loans,
                         net_cash_from_financing=net_cash_from_financing,
                         net_change_in_cash=net_change_in_cash,
                         opening_cash=opening_cash,
                         closing_cash=closing_cash)
# ============================================================================
# PRODUCTION/MANUFACTURING (Phase 3)
# ============================================================================

@main_bp.route('/production')
@login_required
def production():
    """Production dashboard"""
    active_batches = ProductionBatch.query.filter_by(status='in_progress').all()
    planned_batches = ProductionBatch.query.filter_by(status='planned').all()
    completed_today = ProductionBatch.query.filter(
        ProductionBatch.status == 'completed',
        db.func.date(ProductionBatch.end_date) == datetime.utcnow().date()
    ).all()
    
    today_production = sum(b.actual_quantity or 0 for b in completed_today)
    
    low_stock_materials = RawMaterial.query.filter(
        RawMaterial.current_stock <= RawMaterial.reorder_level,
        RawMaterial.is_active == True
    ).all()
    
    return render_template('production/index.html',
                         active_batches=active_batches,
                         planned_batches=planned_batches,
                         completed_today=completed_today,
                         today_production=today_production,
                         low_stock_materials=low_stock_materials)

@main_bp.route('/production/batches')
@login_required
def production_batches():
    """List all production batches"""
    status = request.args.get('status', 'all')
    query = ProductionBatch.query
    
    if status != 'all':
        query = query.filter_by(status=status)
    
    batches = query.order_by(ProductionBatch.created_at.desc()).all()
    return render_template('production/batches.html', batches=batches, current_status=status)

@main_bp.route('/production/batch/add', methods=['GET', 'POST'])
@login_required
def add_production_batch():
    """Create new production batch"""
    if request.method == 'POST':
        try:
            batch_number = f"BATCH-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            product_id = request.form.get('product_id')
            planned_quantity = int(request.form.get('planned_quantity', 0))
            
            batch = ProductionBatch(
                batch_number=batch_number,
                product_id=int(product_id),
                planned_quantity=planned_quantity,
                status='planned',
                created_by=current_user.username,
                notes=request.form.get('notes', '')
            )
            db.session.add(batch)
            db.session.commit()
            
            flash(f'Production batch {batch_number} created successfully!', 'success')
            return redirect(url_for('main.production_batches'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('production/batch_form.html', batch=None, products=products)

@main_bp.route('/production/batch/<int:id>')
@login_required
def view_production_batch(id):
    """View production batch details"""
    batch = ProductionBatch.query.get_or_404(id)
    return render_template('production/batch_view.html', batch=batch)

@main_bp.route('/production/batch/<int:id>/start', methods=['POST'])
@login_required
def start_production_batch(id):
    """Start production batch"""
    batch = ProductionBatch.query.get_or_404(id)
    batch.status = 'in_progress'
    batch.start_date = datetime.utcnow()
    db.session.commit()
    flash('Production batch started!', 'success')
    return redirect(url_for('main.view_production_batch', id=batch.id))

@main_bp.route('/production/batch/<int:id>/complete', methods=['GET', 'POST'])
@login_required
def complete_production_batch(id):
    """Complete production batch"""
    batch = ProductionBatch.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            actual_quantity = int(request.form.get('actual_quantity', 0))
            rejected_quantity = int(request.form.get('rejected_quantity', 0))
            
            batch.actual_quantity = actual_quantity
            batch.rejected_quantity = rejected_quantity
            
            if batch.planned_quantity > 0:
                batch.yield_percentage = (actual_quantity / batch.planned_quantity) * 100
            
            batch.status = 'completed'
            batch.end_date = datetime.utcnow()
            
            # Update product stock
            product = Product.query.get(batch.product_id)
            if product:
                product.stock += actual_quantity
            
            db.session.commit()
            flash('Production batch completed!', 'success')
            return redirect(url_for('main.production_batches'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('production/batch_complete.html', batch=batch)

@main_bp.route('/production/batch/<int:id>/add-material', methods=['POST'])
@login_required
def add_batch_material(id):
    """Add raw material to production batch"""
    batch = ProductionBatch.query.get_or_404(id)
    
    try:
        material = BatchRawMaterial(
            batch_id=batch.id,
            material_name=request.form.get('material_name', ''),
            material_type=request.form.get('material_type', ''),
            quantity_used=float(request.form.get('quantity_used', 0)),
            unit=request.form.get('unit', ''),
            cost_per_unit=float(request.form.get('cost_per_unit', 0)),
            total_cost=float(request.form.get('total_cost', 0)),
            supplier=request.form.get('supplier', ''),
            batch_code=request.form.get('batch_code', '')
        )
        db.session.add(material)
        db.session.commit()
        flash('Raw material added to batch!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.view_production_batch', id=batch.id))

@main_bp.route('/production/batch/<int:id>/quality-check', methods=['GET', 'POST'])
@login_required
def add_quality_check(id):
    """Add quality check for production batch"""
    batch = ProductionBatch.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            check = QualityCheck(
                batch_id=batch.id,
                check_type=request.form.get('check_type', ''),
                check_result=request.form.get('check_result', ''),
                checked_quantity=int(request.form.get('checked_quantity', 0)),
                passed_quantity=int(request.form.get('passed_quantity', 0)),
                failed_quantity=int(request.form.get('failed_quantity', 0)),
                failure_reason=request.form.get('failure_reason', ''),
                checked_by=current_user.username,
                notes=request.form.get('notes', '')
            )
            db.session.add(check)
            db.session.commit()
            flash('Quality check added!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.view_production_batch', id=batch.id))
    
    return render_template('production/quality_check.html', batch=batch)


@main_bp.route('/production/raw-materials/<int:id>/usage', methods=['POST'])
@login_required
def record_material_usage(id):
    """Record raw material usage"""
    material = RawMaterial.query.get_or_404(id)
    
    try:
        quantity = float(request.form.get('quantity', 0))
        batch_id = request.form.get('batch_id')
        
        usage = RawMaterialUsage(
            material_id=material.id,
            batch_id=int(batch_id) if batch_id else None,
            quantity_used=quantity,
            unit=material.unit,
            used_by=current_user.username,
            notes=request.form.get('notes', '')
        )
        db.session.add(usage)
        
        material.current_stock -= quantity
        
        db.session.commit()
        flash('Material usage recorded!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.raw_materials'))

@main_bp.route('/production/reports')
@login_required
def production_reports():
    """Production reports and analytics"""
    total_batches = ProductionBatch.query.count()
    completed_batches = ProductionBatch.query.filter_by(status='completed').count()
    avg_yield = db.session.query(db.func.avg(ProductionBatch.yield_percentage)).filter(
        ProductionBatch.status == 'completed'
    ).scalar() or 0
    
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    production_by_product = db.session.query(
        Product.name,
        db.func.sum(ProductionBatch.actual_quantity)
    ).join(ProductionBatch).filter(
        ProductionBatch.status == 'completed',
        ProductionBatch.end_date >= thirty_days_ago
    ).group_by(Product.name).all()
    
    return render_template('production/reports.html',
                         total_batches=total_batches,
                         completed_batches=completed_batches,
                         avg_yield=avg_yield,
                         production_by_product=production_by_product)


@main_bp.route('/analytics/dashboard')
@login_required
def analytics_dashboard():
    """Advanced analytics dashboard with KPIs and trends"""
    
    # Time periods
    today = datetime.utcnow().date()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)
    last_90_days = today - timedelta(days=90)
    
    # Revenue Metrics
    revenue_today = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        db.func.date(Order.created_at) == today
    ).scalar() or 0
    
    revenue_7_days = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= last_7_days
    ).scalar() or 0
    
    revenue_30_days = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= last_30_days
    ).scalar() or 0
    revenue_90_days = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= last_90_days
    ).scalar() or 0
    
    # Order Metrics
    orders_today = Order.query.filter(
        db.func.date(Order.created_at) == today
    ).count()
    
    orders_7_days = Order.query.filter(
        Order.created_at >= last_7_days
    ).count()
    
    orders_30_days = Order.query.filter(
        Order.created_at >= last_30_days
    ).count()
    
    # Average Order Value
    avg_order_value = revenue_30_days / orders_30_days if orders_30_days > 0 else 0
    
    # Customer Metrics
    new_customers_30_days = Customer.query.filter(
        Customer.created_at >= last_30_days
    ).count()
    
    total_customers = Customer.query.filter_by(is_active=True, is_deleted=False).count()
    
    # Product Performance (Top 10)
    top_products = db.session.query(
        Product.name,
        Product.color,
        db.func.sum(OrderItem.quantity).label('total_sold'),
        db.func.sum(OrderItem.subtotal).label('total_revenue')
    ).join(OrderItem).join(Order).filter(
        Order.created_at >= last_30_days,
        Order.status.in_(['completed', 'processing'])
    ).group_by(Product.id).order_by(db.func.sum(OrderItem.quantity).desc()).limit(10).all()
    
    # Monthly Revenue Trend (Last 12 months)
    monthly_revenue = db.session.query(
        db.func.strftime('%Y-%m', Order.created_at),
        db.func.sum(Order.total_amount)
    ).filter(
        Order.created_at >= today - timedelta(days=365),
        Order.status.in_(['completed', 'processing'])
    ).group_by(db.func.strftime('%Y-%m', Order.created_at)).all()
    
    # Order Status Distribution
    order_status_dist = db.session.query(
        Order.status,
        db.func.count(Order.id)
    ).group_by(Order.status).all()
    
    # Payment Status Distribution
    payment_status_dist = db.session.query(
        Order.payment_status,
        db.func.count(Order.id)
    ).group_by(Order.payment_status).all()
    
    # Customer Retention Rate (simplified)
    returning_customers = db.session.query(Order.customer_id).filter(
        Order.created_at >= last_90_days
    ).group_by(Order.customer_id).having(
        db.func.count(Order.id) > 1
    ).count()
    
    customer_retention_rate = (returning_customers / total_customers * 100) if total_customers > 0 else 0
    
    # Inventory Turnover
    total_inventory_value = db.session.query(
        db.func.sum(Product.stock * Product.base_price)
    ).filter(Product.is_deleted == False).scalar() or 0
    
    inventory_turnover = (revenue_30_days / total_inventory_value * 12) if total_inventory_value > 0 else 0
    
    # Growth Rate (Month over Month)
    last_month_revenue = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.created_at >= last_30_days,
        Order.created_at < last_7_days,
        Order.status.in_(['completed', 'processing'])
    ).scalar() or 0
    
    growth_rate = ((revenue_7_days - last_month_revenue) / last_month_revenue * 100) if last_month_revenue > 0 else 0
    
    return render_template('analytics/dashboard.html',
                         revenue_today=revenue_today,
                         revenue_7_days=revenue_7_days,
                         revenue_30_days=revenue_30_days,
                         revenue_90_days=revenue_90_days,
                         orders_today=orders_today,
                         orders_7_days=orders_7_days,
                         orders_30_days=orders_30_days,
                         avg_order_value=avg_order_value,
                         new_customers_30_days=new_customers_30_days,
                         total_customers=total_customers,
                         top_products=top_products,
                         monthly_revenue=monthly_revenue,
                         order_status_dist=order_status_dist,
                         payment_status_dist=payment_status_dist,
                         customer_retention_rate=customer_retention_rate,
                         inventory_turnover=inventory_turnover,
                         growth_rate=growth_rate)


# ============================================================================
# MATERIAL CONSUMPTION REPORT (ADD THIS)
# ============================================================================

@main_bp.route('/reports/material-consumption')
@login_required
def material_consumption_report():
    """Material consumption report"""
    start_date = request.args.get('start_date', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    consumption = db.session.query(
        RawMaterial.name,
        RawMaterial.unit,
        db.func.sum(MaterialUsage.quantity_used).label('total_used'),
        db.func.count(MaterialUsage.id).label('usage_count')
    ).join(MaterialUsage).filter(
        MaterialUsage.usage_date >= start_date,
        MaterialUsage.usage_date <= end_date
    ).group_by(RawMaterial.id).all()
    
    return render_template('reports/material_consumption.html',
                         consumption=consumption,
                         start_date=start_date,
                         end_date=end_date)
# ============================================================================
# BARCODE/QR CODE GENERATION (Phase 3)
# ============================================================================

@main_bp.route('/products/<int:id>/barcode')
@login_required
def generate_product_barcode(id):
    """Generate barcode for product"""
    try:
        import barcode
        from barcode.writer import ImageWriter
        
        product = Product.query.get_or_404(id)
        
        barcode_number = f"890{product.id:010d}"
        barcode_class = barcode.get('ean13', barcode_number, writer=ImageWriter())
        
        barcode_dir = os.path.join('static', 'barcodes')
        os.makedirs(barcode_dir, exist_ok=True)
        
        filename = f"product_{product.id}_barcode"
        filepath = os.path.join(barcode_dir, filename)
        
        saved_path = barcode_class.save(filepath)
        
        return send_file(f"{saved_path}.png", mimetype='image/png')
        
    except ImportError:
        flash('Barcode library not installed. Run: pip install python-barcode', 'error')
        return redirect(url_for('main.products'))
    except Exception as e:
        flash(f'Error generating barcode: {str(e)}', 'error')
        return redirect(url_for('main.products'))

@main_bp.route('/products/<int:id>/qrcode')
@login_required
def generate_product_qrcode(id):
    """Generate QR code for product"""
    try:
        import qrcode
        
        product = Product.query.get_or_404(id)
        
        qr_data = f"Product: {product.name}\nID: {product.id}\nPrice: PKR {product.base_price}\nStock: {product.stock}"
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        qrcode_dir = os.path.join('static', 'qrcodes')
        os.makedirs(qrcode_dir, exist_ok=True)
        
        filename = f"product_{product.id}_qrcode.png"
        filepath = os.path.join(qrcode_dir, filename)
        
        img.save(filepath)
        
        return send_file(filepath, mimetype='image/png')
        
    except ImportError:
        flash('QR code library not installed. Run: pip install qrcode[pil]', 'error')
        return redirect(url_for('main.products'))
    except Exception as e:
        flash(f'Error generating QR code: {str(e)}', 'error')
        return redirect(url_for('main.products'))

@main_bp.route('/products/<int:id>/labels')
@login_required
def generate_product_labels(id):
    """Generate printable product labels with barcode"""
    product = Product.query.get_or_404(id)
    return render_template('products/labels.html', product=product)

# ============================================================================
# PRODUCTION/MANUFACTURING ROUTES (ADD THESE BEFORE SOCKET.IO HANDLERS)
# ============================================================================

@main_bp.route('/production/raw-materials')
@login_required
def raw_materials():
    """List all raw materials"""
    materials = RawMaterial.query.filter_by(is_active=True).all()
    low_stock = [m for m in materials if m.current_stock <= m.reorder_level]
    return render_template('materials/index.html', materials=materials, low_stock=low_stock)

@main_bp.route('/raw-materials/add', methods=['GET', 'POST'])
@login_required
def add_raw_material():
    """Add new raw material"""
    if request.method == 'POST':
        try:
            material = RawMaterial(
                name=request.form.get('name', ''),
                material_type=request.form.get('material_type', ''),
                specification=request.form.get('specification', ''),
                current_stock=float(request.form.get('current_stock', 0)),
                unit=request.form.get('unit', ''),
                reorder_level=float(request.form.get('reorder_level', 0)),
                cost_per_unit=float(request.form.get('cost_per_unit', 0)),
                supplier=request.form.get('supplier', '')
            )
            db.session.add(material)
            db.session.commit()
            flash('Raw material added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.raw_materials'))
    
    return render_template('materials/form.html', material=None)

@main_bp.route('/products/<int:id>/bom', methods=['GET', 'POST'])
@login_required
def manage_product_bom(id):
    """Manage Bill of Materials for a product"""
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Clear existing BOM items
            BOMItem.query.filter_by(product_id=product.id).delete()
            
            # Add new BOM items
            material_ids = request.form.getlist('material_id[]')
            quantities = request.form.getlist('quantity_required[]')
            waste_percentages = request.form.getlist('waste_percentage[]')
            
            for i, mid in enumerate(material_ids):
                if mid and quantities[i]:
                    bom_item = BOMItem(
                        product_id=product.id,
                        raw_material_id=int(mid),
                        quantity_required=float(quantities[i]),
                        unit='meters',
                        waste_percentage=float(waste_percentages[i]) if waste_percentages[i] else 0
                    )
                    db.session.add(bom_item)
            
            db.session.commit()
            flash('BOM updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.manage_product_bom', id=product.id))
    
    materials = RawMaterial.query.filter_by(is_active=True).all()
    bom_items = BOMItem.query.filter_by(product_id=product.id).all()
    return render_template('products/bom.html', product=product, materials=materials, bom_items=bom_items)


# BULK IMPORTS
@main_bp.route('/import/<module>', methods=['GET', 'POST'])
@login_required
def import_data(module):
    """Bulk import data from Excel"""
    if request.method == 'POST':
        try:
            file = request.files['file']
            if not file:
                flash('No file selected', 'error')
                return redirect(request.url)
            
            # Read Excel file
            import pandas as pd
            df = pd.read_excel(file)
            
            # Process based on module
            if module == 'products':
                for index, row in df.iterrows():
                    product = Product(
                        name=row['name'],
                        specification=row.get('specification', ''),
                        base_price=integer(row['base_price']),
                        stock=int(row['stock']),
                        color=row['color'],
                        product_type=row.get('product_type', 'product'),
                        material_type=row.get('material_type', ''),
                        usp_type=row.get('usp_type', ''),
                        shape_type=row.get('shape_type', ''),
                        dimensions=row.get('dimensions', '')
                    )
                    db.session.add(product)
            
            elif module == 'customers':
                for index, row in df.iterrows():
                    customer = Customer(
                        name=row['name'],
                        business_name=row.get('business_name', ''),
                        role=row.get('role', ''),
                        email=row.get('email', ''),
                        phone=row['phone'],
                        address=row.get('address', ''),
                        payment_terms=row.get('payment_terms', 'cash'),
                        credit_limit=float(row.get('credit_limit', 0))
                    )
                    db.session.add(customer)
            
            # Add more modules as needed...
            
            db.session.commit()
            flash(f'{len(df)} records imported successfully!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing data: {str(e)}', 'error')
        
        return redirect(url_for(f'main.{module}'))
    
    return render_template(f'import/{module}.html')


# ============================================================================
# ENHANCED REPORTING (PHASE 4) - COMPLETE ROUTES
# ============================================================================

@main_bp.route('/reports/sales-analysis', methods=['GET'])
@login_required
def sales_analysis_report():
    """Advanced sales analysis with multiple dimensions"""
    date_from = request.args.get('date_from', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.utcnow().strftime('%Y-%m-%d'))
    group_by = request.args.get('group_by', 'product')  # product, customer, date
    
    # Base query
    orders = Order.query.filter(
        Order.created_at >= date_from,
        Order.created_at <= date_to,
        Order.status.in_(['completed', 'processing'])
    ).all()
    
    # Calculate metrics
    total_revenue = sum(o.total_amount for o in orders)
    total_orders = len(orders)
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # Group data
    report_data = []
    
    if group_by == 'product':
        # Sales by product
        product_sales = {}
        for order in orders:
            for item in order.order_items:
                if item.product_id not in product_sales:
                    product_sales[item.product_id] = {
                        'name': item.product.name if item.product else 'N/A',
                        'quantity': 0,
                        'revenue': 0,
                        'orders': 0
                    }
                product_sales[item.product_id]['quantity'] += item.quantity
                product_sales[item.product_id]['revenue'] += item.subtotal
                product_sales[item.product_id]['orders'] += 1
        
        report_data = sorted(product_sales.values(), key=lambda x: x['revenue'], reverse=True)
        
    elif group_by == 'customer':
        # Sales by customer
        customer_sales = {}
        for order in orders:
            cust_key = order.customer_name_snapshot or 'Walk-in Customer'
            if cust_key not in customer_sales:
                customer_sales[cust_key] = {
                    'name': cust_key,
                    'orders': 0,
                    'revenue': 0,
                    'avg_order': 0
                }
            customer_sales[cust_key]['orders'] += 1
            customer_sales[cust_key]['revenue'] += order.total_amount
        
        for cust in customer_sales.values():
            cust['avg_order'] = cust['revenue'] / cust['orders'] if cust['orders'] > 0 else 0
        
        report_data = sorted(customer_sales.values(), key=lambda x: x['revenue'], reverse=True)
    
    elif group_by == 'date':
        # Sales by date (daily trend)
        from collections import defaultdict
        daily_sales = defaultdict(lambda: {'orders': 0, 'revenue': 0})
        
        for order in orders:
            date_key = order.created_at.strftime('%Y-%m-%d')
            daily_sales[date_key]['orders'] += 1
            daily_sales[date_key]['revenue'] += order.total_amount
        
        report_data = [
            {'date': k, 'orders': v['orders'], 'revenue': v['revenue']}
            for k, v in sorted(daily_sales.items())
        ]
    
    return render_template('reports/sales_analysis.html',
                         report_data=report_data,
                         total_revenue=total_revenue,
                         total_orders=total_orders,
                         avg_order_value=avg_order_value,
                         date_from=date_from,
                         date_to=date_to,
                         group_by=group_by)


@main_bp.route('/reports/inventory-valuation', methods=['GET'])
@login_required
def inventory_valuation_report():
    """Inventory valuation with stock analysis"""
    products = Product.query.filter_by(is_deleted=False).all()
    
    # Calculate metrics
    total_items = len(products)
    total_stock_value = sum(p.stock * p.base_price for p in products)
    low_stock_items = [p for p in products if p.stock < 10000]
    out_of_stock = [p for p in products if p.stock == 0]
    
    # Categorize by stock level
    fast_moving = [p for p in products if p.stock < 5000]  # Low stock = fast moving
    slow_moving = [p for p in products if p.stock > 50000]  # High stock = slow moving
    
    return render_template('reports/inventory_valuation.html',
                         products=products,
                         total_items=total_items,
                         total_stock_value=total_stock_value,
                         low_stock_items=low_stock_items,
                         out_of_stock=out_of_stock,
                         fast_moving=fast_moving,
                         slow_moving=slow_moving)


@main_bp.route('/reports/customer-purchase-history/<int:customer_id>', methods=['GET'])
@login_required
def customer_purchase_history(customer_id):
    """Detailed customer purchase history"""
    customer = Customer.query.get_or_404(customer_id)
    
    orders = Order.query.filter_by(customer_id=customer_id).order_by(Order.created_at.desc()).all()
    
    # Calculate metrics
    total_orders = len(orders)
    total_spent = sum(o.total_amount for o in orders)
    total_paid = sum(o.paid_amount for o in orders)
    avg_order_value = total_spent / total_orders if total_orders > 0 else 0
    
    # Get most purchased products
    product_purchases = {}
    for order in orders:
        for item in order.order_items:
            if item.product_id not in product_purchases:
                product_purchases[item.product_id] = {
                    'name': item.product.name if item.product else 'N/A',
                    'quantity': 0,
                    'revenue': 0
                }
            product_purchases[item.product_id]['quantity'] += item.quantity
            product_purchases[item.product_id]['revenue'] += item.subtotal
    
    top_products = sorted(product_purchases.values(), key=lambda x: x['revenue'], reverse=True)[:5]
    
    return render_template('reports/customer_purchase_history.html',
                         customer=customer,
                         orders=orders,
                         total_orders=total_orders,
                         total_spent=total_spent,
                         total_paid=total_paid,
                         avg_order_value=avg_order_value,
                         top_products=top_products)


@main_bp.route('/reports/production-efficiency', methods=['GET'])
@login_required
def production_efficiency_report():
    """Production batch efficiency analysis"""
    date_from = request.args.get('date_from', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.utcnow().strftime('%Y-%m-%d'))
    
    # Check if ProductionBatch model exists
    try:
        batches = ProductionBatch.query.filter(
            ProductionBatch.start_date >= date_from,
            ProductionBatch.end_date <= date_to,
            ProductionBatch.status == 'completed'
        ).all()
    except:
        batches = []
    
    # Calculate metrics
    total_batches = len(batches)
    avg_yield = sum(b.yield_percentage for b in batches) / total_batches if total_batches > 0 else 0
    total_produced = sum(b.actual_quantity or 0 for b in batches)
    total_planned = sum(b.planned_quantity for b in batches)
    total_rejected = sum(b.rejected_quantity or 0 for b in batches)
    
    return render_template('reports/production_efficiency.html',
                         batches=batches,
                         total_batches=total_batches,
                         avg_yield=avg_yield,
                         total_produced=total_produced,
                         total_planned=total_planned,
                         total_rejected=total_rejected,
                         date_from=date_from,
                         date_to=date_to)


# ============================================================================
# SUPPLY CHAIN & PURCHASE MANAGEMENT (PHASE 5)
# ============================================================================

@main_bp.route('/suppliers')
@login_required
def suppliers():
    """List all suppliers"""
    suppliers = Supplier.query.filter_by(is_active=True).all()
    return render_template('suppliers/index.html', suppliers=suppliers)

@main_bp.route('/suppliers/add', methods=['GET', 'POST'])
@login_required
def add_supplier():
    """Add new supplier"""
    if request.method == 'POST':
        try:
            supplier = Supplier(
                name=request.form.get('name', ''),
                contact_person=request.form.get('contact_person', ''),
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''),
                address=request.form.get('address', ''),
                city=request.form.get('city', ''),
                country=request.form.get('country', 'Pakistan'),
                cnic=request.form.get('cnic', ''),
                ntn=request.form.get('ntn', ''),
                payment_terms=request.form.get('payment_terms', '30 days'),
                credit_limit=float(request.form.get('credit_limit', 0))
            )
            db.session.add(supplier)
            db.session.commit()
            flash('Supplier added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.suppliers'))
    
    return render_template('suppliers/form.html', supplier=None)

@main_bp.route('/suppliers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_supplier(id):
    """Edit supplier"""
    supplier = Supplier.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            supplier.name = request.form.get('name', '')
            supplier.contact_person = request.form.get('contact_person', '')
            supplier.email = request.form.get('email', '')
            supplier.phone = request.form.get('phone', '')
            supplier.address = request.form.get('address', '')
            supplier.city = request.form.get('city', '')
            supplier.country = request.form.get('country', 'Pakistan')
            supplier.cnic = request.form.get('cnic', '')
            supplier.ntn = request.form.get('ntn', '')
            supplier.payment_terms = request.form.get('payment_terms', '30 days')
            supplier.credit_limit = float(request.form.get('credit_limit', 0))
            supplier.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash('Supplier updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.suppliers'))
    
    return render_template('suppliers/form.html', supplier=supplier)

@main_bp.route('/suppliers/delete/<int:id>')
@login_required
def delete_supplier(id):
    """Delete supplier"""
    try:
        supplier = Supplier.query.get_or_404(id)
        supplier.is_active = False
        db.session.commit()
        flash('Supplier archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.suppliers'))

@main_bp.route('/suppliers/<int:id>/purchase-history')
@login_required
def supplier_purchase_history(id):
    """Supplier purchase history"""
    supplier = Supplier.query.get_or_404(id)
    purchase_orders = PurchaseOrder.query.filter_by(supplier_id=id).order_by(PurchaseOrder.order_date.desc()).all()
    
    total_orders = len(purchase_orders)
    total_spent = sum(po.grand_total for po in purchase_orders)
    avg_order_value = total_spent / total_orders if total_orders > 0 else 0
    
    return render_template('suppliers/purchase_history.html',
                         supplier=supplier,
                         purchase_orders=purchase_orders,
                         total_orders=total_orders,
                         total_spent=total_spent,
                         avg_order_value=avg_order_value)

@main_bp.route('/purchase-orders')
@login_required
def purchase_orders():
    """List all purchase orders"""
    status = request.args.get('status', 'all')
    query = PurchaseOrder.query
    if status != 'all':
        query = query.filter_by(status=status)
    purchase_orders = query.order_by(PurchaseOrder.order_date.desc()).all()
    return render_template('purchase_orders/index.html', purchase_orders=purchase_orders, current_status=status)

@main_bp.route('/purchase-orders/add', methods=['GET', 'POST'])
@login_required
def add_purchase_order():
    """Create new purchase order"""
    if request.method == 'POST':
        try:
            po_number = f"PO-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            supplier_id = request.form.get('supplier_id')
            
            po = PurchaseOrder(
                po_number=po_number,
                supplier_id=int(supplier_id),
                expected_date=datetime.strptime(request.form.get('expected_date', ''), '%Y-%m-%d') if request.form.get('expected_date') else None,
                status='draft',
                notes=request.form.get('notes', ''),
                created_by=current_user.username
            )
            db.session.add(po)
            db.session.flush()
            
            # Add items
            material_ids = request.form.getlist('material_id[]')
            quantities = request.form.getlist('quantity[]')
            unit_prices = request.form.getlist('unit_price[]')
            tax_rates = request.form.getlist('tax_rate[]')
            
            total_amount = 0
            tax_amount = 0
            
            for i, mid in enumerate(material_ids):
                if mid and quantities[i]:
                    qty = float(quantities[i])
                    price = float(unit_prices[i])
                    tax_rate = float(tax_rates[i]) if tax_rates[i] else 17
                    subtotal = qty * price
                    tax = subtotal * (tax_rate / 100)
                    total = subtotal + tax
                    
                    item = PurchaseOrderItem(
                        po_id=po.id,
                        material_id=int(mid),
                        quantity_ordered=qty,
                        unit_price=price,
                        subtotal=subtotal,
                        tax_rate=tax_rate,
                        tax_amount=tax,
                        total=total
                    )
                    db.session.add(item)
                    
                    total_amount += subtotal
                    tax_amount += tax
            
            po.total_amount = total_amount
            po.tax_amount = tax_amount
            po.grand_total = total_amount + tax_amount
            
            db.session.commit()
            flash(f'Purchase Order {po_number} created successfully!', 'success')
            return redirect(url_for('main.view_purchase_order', id=po.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    suppliers = Supplier.query.filter_by(is_active=True).all()
    materials = RawMaterial.query.filter_by(is_active=True).all()
    return render_template('purchase_orders/form.html', suppliers=suppliers, materials=materials, po=None)

@main_bp.route('/purchase-orders/<int:id>')
@login_required
def view_purchase_order(id):
    """View purchase order details"""
    po = PurchaseOrder.query.get_or_404(id)
    return render_template('purchase_orders/view.html', po=po)

@main_bp.route('/purchase-orders/<int:id>/send', methods=['POST'])
@login_required
def send_purchase_order(id):
    """Send purchase order to supplier"""
    po = PurchaseOrder.query.get_or_404(id)
    po.status = 'sent'
    db.session.commit()
    flash(f'Purchase Order {po.po_number} sent to supplier!', 'success')
    return redirect(url_for('main.view_purchase_order', id=po.id))

@main_bp.route('/purchase-orders/<int:id>/receive', methods=['GET', 'POST'])
@login_required
def receive_purchase_order(id):
    """Receive goods against purchase order"""
    po = PurchaseOrder.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            grn_number = f"GRN-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            warehouse_id = request.form.get('warehouse_id')
            
            gr = GoodsReceipt(
                grn_number=grn_number,
                po_id=po.id,
                supplier_id=po.supplier_id,
                warehouse_id=int(warehouse_id) if warehouse_id else None,
                status='pending',
                received_by=current_user.username
            )
            db.session.add(gr)
            db.session.flush()
            
            # Add receipt items
            for item in po.items:
                qty_received = float(request.form.get(f'qty_{item.id}', 0))
                qty_accepted = float(request.form.get(f'accepted_{item.id}', qty_received))
                qty_rejected = qty_received - qty_accepted
                batch_number = request.form.get(f'batch_{item.id}', '')
                expiry_date = request.form.get(f'expiry_{item.id}', '')
                
                gri = GoodsReceiptItem(
                    gr_id=gr.id,
                    material_id=item.material_id,
                    quantity_ordered=item.quantity_ordered,
                    quantity_received=qty_received,
                    quantity_accepted=qty_accepted,
                    quantity_rejected=qty_rejected,
                    rejection_reason=request.form.get(f'rejection_reason_{item.id}', ''),
                    batch_number=batch_number,
                    expiry_date=datetime.strptime(expiry_date, '%Y-%m-%d') if expiry_date else None
                )
                db.session.add(gri)
                
                # Update PO item received quantity
                item.quantity_received += qty_received
                
                # Add stock if accepted
                if qty_accepted > 0:
                    material = RawMaterial.query.get(item.material_id)
                    if material:
                        material.current_stock += qty_accepted
                        
                        # Create material batch record
                        batch = MaterialBatch(
                            material_id=item.material_id,
                            batch_number=batch_number or f"BATCH-{datetime.utcnow().strftime('%Y%m%d')}",
                            supplier_id=po.supplier_id,
                            quantity_received=qty_accepted,
                            quantity_remaining=qty_accepted,
                            expiry_date=datetime.strptime(expiry_date, '%Y-%m-%d') if expiry_date else None,
                            warehouse_id=int(warehouse_id) if warehouse_id else None
                        )
                        db.session.add(batch)
            
            gr.total_items = len(po.items)
            gr.accepted_items = sum(1 for item in po.items if float(request.form.get(f'accepted_{item.id}', 0)) > 0)
            gr.rejected_items = gr.total_items - gr.accepted_items
            gr.status = 'accepted' if gr.rejected_items == 0 else 'partial'
            
            # Update PO status
            if all(item.quantity_received >= item.quantity_ordered for item in po.items):
                po.status = 'received'
                po.received_date = datetime.utcnow()
            
            db.session.commit()
            flash(f'Goods Receipt {grn_number} created successfully!', 'success')
            return redirect(url_for('main.view_goods_receipt', id=gr.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('purchase_orders/receive.html', po=po)

@main_bp.route('/goods-receipts')
@login_required
def goods_receipts():
    """List all goods receipts"""
    status = request.args.get('status', 'all')
    query = GoodsReceipt.query
    if status != 'all':
        query = query.filter_by(status=status)
    receipts = query.order_by(GoodsReceipt.created_at.desc()).all()
    return render_template('goods_receipts/index.html', receipts=receipts, current_status=status)

@main_bp.route('/goods-receipts/<int:id>')
@login_required
def view_goods_receipt(id):
    """View goods receipt details"""
    gr = GoodsReceipt.query.get_or_404(id)
    return render_template('goods_receipts/view.html', gr=gr)

@main_bp.route('/warehouses')
@login_required
def warehouses():
    """List all warehouses"""
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    return render_template('warehouses/index.html', warehouses=warehouses)

@main_bp.route('/warehouses/add', methods=['GET', 'POST'])
@login_required
def add_warehouse():
    """Add new warehouse"""
    if request.method == 'POST':
        try:
            warehouse = Warehouse(
                name=request.form.get('name', ''),
                code=request.form.get('code', ''),
                address=request.form.get('address', ''),
                city=request.form.get('city', ''),
                manager=request.form.get('manager', ''),
                phone=request.form.get('phone', '')
            )
            db.session.add(warehouse)
            db.session.commit()
            flash('Warehouse added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.warehouses'))
    
    return render_template('warehouses/form.html', warehouse=None)

@main_bp.route('/warehouses/<int:id>/stock')
@login_required
def warehouse_stock(id):
    """View warehouse stock levels"""
    warehouse = Warehouse.query.get_or_404(id)
    stock_levels = WarehouseStock.query.filter_by(warehouse_id=id).all()
    return render_template('warehouses/stock.html', warehouse=warehouse, stock_levels=stock_levels)

@main_bp.route('/stock-transfers')
@login_required
def stock_transfers():
    """List all stock transfers"""
    status = request.args.get('status', 'all')
    query = StockTransfer.query
    if status != 'all':
        query = query.filter_by(status=status)
    transfers = query.order_by(StockTransfer.created_at.desc()).all()
    return render_template('stock_transfers/index.html', transfers=transfers, current_status=status)

@main_bp.route('/stock-transfers/add', methods=['GET', 'POST'])
@login_required
def add_stock_transfer():
    """Create new stock transfer"""
    if request.method == 'POST':
        try:
            transfer_number = f"ST-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            from_warehouse_id = request.form.get('from_warehouse_id')
            to_warehouse_id = request.form.get('to_warehouse_id')
            
            transfer = StockTransfer(
                transfer_number=transfer_number,
                from_warehouse_id=int(from_warehouse_id),
                to_warehouse_id=int(to_warehouse_id),
                status='pending',
                notes=request.form.get('notes', ''),
                requested_by=current_user.username
            )
            db.session.add(transfer)
            db.session.flush()
            
            # Add items
            material_ids = request.form.getlist('material_id[]')
            quantities = request.form.getlist('quantity[]')
            
            for i, mid in enumerate(material_ids):
                if mid and quantities[i]:
                    item = StockTransferItem(
                        transfer_id=transfer.id,
                        material_id=int(mid),
                        quantity=float(quantities[i])
                    )
                    db.session.add(item)
            
            db.session.commit()
            flash(f'Stock Transfer {transfer_number} created successfully!', 'success')
            return redirect(url_for('main.view_stock_transfer', id=transfer.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    materials = RawMaterial.query.filter_by(is_active=True).all()
    return render_template('stock_transfers/form.html', warehouses=warehouses, materials=materials, transfer=None)

@main_bp.route('/stock-transfers/<int:id>')
@login_required
def view_stock_transfer(id):
    """View stock transfer details"""
    transfer = StockTransfer.query.get_or_404(id)
    return render_template('stock_transfers/view.html', transfer=transfer)

@main_bp.route('/stock-transfers/<int:id>/complete', methods=['POST'])
@login_required
def complete_stock_transfer(id):
    """Complete stock transfer"""
    transfer = StockTransfer.query.get_or_404(id)
    
    try:
        # Deduct from source warehouse
        for item in transfer.items:
            from_stock = WarehouseStock.query.filter_by(
                warehouse_id=transfer.from_warehouse_id,
                material_id=item.material_id
            ).first()
            
            if from_stock:
                from_stock.quantity -= item.quantity
                from_stock.available_quantity = from_stock.quantity - from_stock.reserved_quantity
            
            # Add to destination warehouse
            to_stock = WarehouseStock.query.filter_by(
                warehouse_id=transfer.to_warehouse_id,
                material_id=item.material_id
            ).first()
            
            if to_stock:
                to_stock.quantity += item.quantity
                to_stock.available_quantity = to_stock.quantity - to_stock.reserved_quantity
            else:
                to_stock = WarehouseStock(
                    warehouse_id=transfer.to_warehouse_id,
                    material_id=item.material_id,
                    quantity=item.quantity,
                    available_quantity=item.quantity
                )
                db.session.add(to_stock)
            
            item.quantity_transferred = item.quantity
        
        transfer.status = 'completed'
        transfer.completed_at = datetime.utcnow()
        db.session.commit()
        
        flash('Stock transfer completed successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.view_stock_transfer', id=transfer.id))

@main_bp.route('/material-batches')
@login_required
def material_batches():
    """List all material batches"""
    batches = MaterialBatch.query.order_by(MaterialBatch.received_date.desc()).all()
    expiring_soon = [b for b in batches if b.expiry_date and b.expiry_date <= datetime.utcnow().date() + timedelta(days=30)]
    expired = [b for b in batches if b.expiry_date and b.expiry_date < datetime.utcnow().date()]
    
    return render_template('material_batches/index.html',
                         batches=batches,
                         expiring_soon=expiring_soon,
                         expired=expired)

@main_bp.route('/supplier-invoices')
@login_required
def supplier_invoices():
    """List all supplier invoices"""
    status = request.args.get('status', 'all')
    query = SupplierInvoice.query
    if status != 'all':
        query = query.filter_by(status=status)
    invoices = query.order_by(SupplierInvoice.created_at.desc()).all()
    return render_template('supplier_invoices/index.html', invoices=invoices, current_status=status)

@main_bp.route('/supplier-invoices/add', methods=['GET', 'POST'])
@login_required
def add_supplier_invoice():
    """Add supplier invoice"""
    if request.method == 'POST':
        try:
            invoice = SupplierInvoice(
                invoice_number=request.form.get('invoice_number', ''),
                supplier_id=int(request.form.get('supplier_id', 0)),
                po_id=int(request.form.get('po_id', 0)) if request.form.get('po_id') else None,
                invoice_date=datetime.strptime(request.form.get('invoice_date', ''), '%Y-%m-%d'),
                due_date=datetime.strptime(request.form.get('due_date', ''), '%Y-%m-%d') if request.form.get('due_date') else None,
                total_amount=float(request.form.get('total_amount', 0)),
                notes=request.form.get('notes', '')
            )
            invoice.balance_amount = invoice.total_amount
            db.session.add(invoice)
            db.session.commit()
            flash('Supplier invoice added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.supplier_invoices'))
    
    suppliers = Supplier.query.filter_by(is_active=True).all()
    purchase_orders = PurchaseOrder.query.filter_by(status='received').all()
    return render_template('supplier_invoices/form.html', suppliers=suppliers, purchase_orders=purchase_orders, invoice=None)

    
# ============================================================================
# SOCKET.IO EVENT HANDLERS
# ============================================================================

@socketio.on('connect', namespace='/admin')
def admin_connect():
    if not current_user.is_authenticated:
        from flask_socketio import disconnect
        disconnect()
        return False
    emit('connected', {'message': 'Connected to Ampoulex real-time updates'})

@socketio.on('disconnect', namespace='/admin')
def admin_disconnect():
    username = current_user.username if current_user.is_authenticated else 'anonymous'
    print(f'Admin client disconnected: {username}')