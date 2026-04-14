from datetime import datetime, timedelta
import os
import json
import random
from sqlalchemy import func

from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file, send_from_directory, current_app, session
from flask_login import login_user, logout_user, login_required, current_user
from flask_socketio import emit

from app import db, socketio
from app.models import (
    User, Product, Customer, Inquiry, Order, OrderItem,
    Employee, Expense, Accounting, CompanySetting, CustomerProductPrice,
    TaxSetting, StockAdjustment, Report, InquiryItem,
    Attendance, Timesheet, LeaveRequest, PayrollPayment,
    RawMaterial, BOMItem, MaterialUsage, ProductionBatch,
    StockAlert, OrderApproval,
    Supplier, PurchaseOrder, PurchaseOrderItem,
    GoodsReceipt, GoodsReceiptItem,
    Warehouse, WarehouseStock, StockTransfer, StockTransferItem,
    MaterialBatch, SupplierInvoice,
    QCParameter, QCResult, CustomerComplaint, CAPA, CalibrationRecord,
    StockMovement, InventoryAdjustment, StockCount, StockCountItem,
    FBRInvoice, TaxReturn,
    PaintingServicePrice, PaintingOrder, PaintingOrderItem,
    CustomerPaintingPrice,
    Account, JournalEntry, JournalEntryLine, AccountingPeriod,
    BankAccount, BankReconciliation,
    PaymentVoucher, PaymentVoucherLine,
    ReceiptVoucher, ReceiptVoucherLine,
    AuditLog
)
from app.utils.tax_calculator import calculate_monthly_tax_deduction


main_bp = Blueprint('main', __name__) 

@main_bp.route('/favicon.ico')
def favicon():
    return send_from_directory(
        current_app.static_folder,
        'favicon.ico',
        mimetype='image/vnd.microsoft.icon'
    )
# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def is_accounting_period_open(entry_date):
    """Check if accounting period is open for given date"""
    period = AccountingPeriod.query.filter(
        AccountingPeriod.start_date <= entry_date,
        AccountingPeriod.end_date >= entry_date
    ).first()
    
    if period and period.is_closed:
        return False, period
    return True, None

# ============================================================================
# AUTO-JOURNAL ENTRY HELPER FUNCTIONS
# ============================================================================

def create_journal_entry_for_order(order):
    """Auto-create journal entry when order is completed"""
    try:
        entry_number = f"JE-ORD-{order.order_number}"
        
        # Check if journal already exists
        existing = JournalEntry.query.filter_by(
            reference_type='order',
            reference_id=order.id
        ).first()
        
        if existing:
            return  # Already created
        
        # Get Accounts
        ar_account = Account.query.filter(Account.account_name.like('%Receivable%')).first()
        sales_account = Account.query.filter(Account.account_name.like('%Sales%')).first()
        
        if not ar_account or not sales_account:
            return  # Accounts not configured
        
        # Create Journal Entry
        entry = JournalEntry(
            entry_number=entry_number,
            entry_date=order.created_at.date(),
            description=f"Sales Order {order.order_number}",
            reference=order.order_number,
            reference_type='order',
            reference_id=order.id,
            status='posted',
            created_by='system'
        )
        db.session.add(entry)
        db.session.flush()
        
        # Debit Accounts Receivable
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=ar_account.id,
            description=f'AR for {order.order_number}',
            debit=order.total_amount,
            credit=0
        ))
        
        # Credit Sales Revenue
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=sales_account.id,
            description=f'Sales for {order.order_number}',
            debit=0,
            credit=order.total_amount
        ))
        
        db.session.commit()
        print(f"✅ Auto-journal created for order {order.order_number}")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creating journal for order: {e}")


def create_journal_entry_for_payment(payment):
    """Auto-create journal entry when payment is received"""
    try:
        entry_number = f"JE-PAY-{payment.id}"
        
        # Check if journal already exists
        existing = JournalEntry.query.filter_by(
            reference_type='payment',
            reference_id=payment.id
        ).first()
        
        if existing:
            return
        
        # Get Accounts
        cash_account = Account.query.filter(Account.account_name.like('%Cash%')).first()
        ar_account = Account.query.filter(Account.account_name.like('%Receivable%')).first()
        
        if not cash_account or not ar_account:
            return
        
        # Create Journal Entry
        entry = JournalEntry(
            entry_number=entry_number,
            entry_date=payment.payment_date,
            description=f"Payment received",
            reference=f"PAY-{payment.id}",
            reference_type='payment',
            reference_id=payment.id,
            status='posted',
            created_by='system'
        )
        db.session.add(entry)
        db.session.flush()
        
        # Debit Cash
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=cash_account.id,
            description='Cash received',
            debit=payment.net_pay,
            credit=0
        ))
        
        # Credit Accounts Receivable
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=ar_account.id,
            description='AR reduction',
            debit=0,
            credit=payment.net_pay
        ))
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creating journal for payment: {e}")


def create_journal_entry_for_expense(expense):
    """Auto-create journal entry when expense is paid"""
    try:
        if expense.status != 'paid':
            return
        
        entry_number = f"JE-EXP-{expense.id}"
        
        # Check if journal already exists
        existing = JournalEntry.query.filter_by(
            reference_type='expense',
            reference_id=expense.id
        ).first()
        
        if existing:
            return
        
        # Get Accounts
        cash_account = Account.query.filter(Account.account_name.like('%Cash%')).first()
        expense_account = Account.query.filter(Account.account_name.like('%Expenses%')).first()
        
        if not cash_account or not expense_account:
            return
        
        # Create Journal Entry
        entry = JournalEntry(
            entry_number=entry_number,
            entry_date=expense.date,
            description=f"Expense: {expense.category}",
            reference=expense.reference_number or f"EXP-{expense.id}",
            reference_type='expense',
            reference_id=expense.id,
            status='posted',
            created_by='system'
        )
        db.session.add(entry)
        db.session.flush()
        
        # Debit Expense
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=expense_account.id,
            description=expense.description or expense.category,
            debit=expense.amount,
            credit=0
        ))
        
        # Credit Cash
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=cash_account.id,
            description='Cash paid',
            debit=0,
            credit=expense.amount
        ))
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creating journal for expense: {e}")


def create_journal_entry_for_production_batch(batch):
    """Auto-create journal entry when production batch is completed"""
    try:
        if batch.status != 'completed' or not batch.actual_quantity:
            return
        
        entry_number = f"JE-PROD-{batch.batch_number}"
        
        # Check if journal already exists
        existing = JournalEntry.query.filter_by(
            reference_type='production',
            reference_id=batch.id
        ).first()
        
        if existing:
            return
        
        # Get product and calculate cost
        product = Product.query.get(batch.product_id)
        if not product:
            return
        
        # Get BOM items to calculate material cost
        bom_items = BOMItem.query.filter_by(product_id=batch.product_id, is_active=True).all()
        total_material_cost = 0
        
        for bom_item in bom_items:
            material = RawMaterial.query.get(bom_item.raw_material_id)
            if material:
                total_material_cost += (bom_item.quantity_required * 
                                       (material.cost_per_unit or 0) * 
                                       batch.actual_quantity)
        
        if total_material_cost == 0:
            return
        
        # Get Accounts
        inventory_account = Account.query.filter(Account.account_name.like('%Inventory%')).first()
        cogs_account = Account.query.filter(Account.account_name.like('%Cost of Goods%')).first()
        
        if not inventory_account or not cogs_account:
            return
        
        # Create Journal Entry
        entry = JournalEntry(
            entry_number=entry_number,
            entry_date=batch.end_date.date() if batch.end_date else datetime.utcnow().date(),
            description=f"Production Batch {batch.batch_number}",
            reference=batch.batch_number,
            reference_type='production',
            reference_id=batch.id,
            status='posted',
            created_by='system'
        )
        db.session.add(entry)
        db.session.flush()
        
        # Debit COGS
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=cogs_account.id,
            description=f'Material cost for {batch.actual_quantity} units',
            debit=total_material_cost,
            credit=0
        ))
        
        # Credit Inventory
        db.session.add(JournalEntryLine(
            entry_id=entry.id,
            account_id=inventory_account.id,
            description='Inventory reduction',
            debit=0,
            credit=total_material_cost
        ))
        
        db.session.commit()
        print(f"✅ Auto-journal created for production batch {batch.batch_number}")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creating journal for production: {e}")

def find_or_create_customer(customer_name=None, business_name=None, email=None, phone=None):
    """Find existing customer or create new one automatically"""
    if email and email.strip():
        customer = Customer.query.filter_by(email=email.strip().lower()).first()
        if customer:
            return customer, False
    
    if phone and phone.strip():
        phone_clean = ''.join(filter(str.isdigit, phone))
        customers_by_phone = Customer.query.filter(Customer.phone.like(f'%{phone_clean}%')).all()
        if customers_by_phone:
            return customers_by_phone[0], False
    
    if business_name and business_name.strip():
        customer = Customer.query.filter_by(business_name=business_name.strip()).first()
        if customer:
            return customer, False
    
    if customer_name and customer_name.strip():
        customer = Customer.query.filter_by(name=customer_name.strip()).first()
        if customer:
            return customer, False
    
    new_customer = Customer(
        name=customer_name or 'N/A',
        business_name=business_name or 'N/A',
        email=email.strip().lower() if email else None
    )
    db.session.add(new_customer)
    db.session.commit()
    return new_customer, True


def broadcast_update(event_name, data):
    """Broadcast real-time updates"""
    try:
        socketio.emit(event_name, data, namespace='/admin')
    except Exception as e:
        print(f"Socket.IO error: {e}")


def notify_new_inquiry(inquiry):
    """Notify when new inquiry is created"""
    broadcast_update('new_inquiry', {
        'id': inquiry.id,
        'number': inquiry.inquiry_number,
        'customer': inquiry.customer_name,
        'business': inquiry.business_name,
        'quantity': inquiry.quantity,
        'status': inquiry.status,
        'created_at': inquiry.created_at.isoformat()
    })


def notify_new_order(order):
    """Notify when new order is created"""
    broadcast_update('new_order', {
        'id': order.id,
        'number': order.order_number,
        'customer': order.customer_name_snapshot,
        'total': float(order.total_amount),
        'status': order.status,
        'payment_status': order.payment_status,
        'created_at': order.created_at.isoformat()
    })


def notify_low_stock_alert(product):
    """Notify when product stock is low"""
    broadcast_update('low_stock_alert', {
        'product_id': product.id,
        'product_name': product.name,
        'current_stock': product.stock,
        'threshold': 10000
    })


def group_products_by_base():
    """Group products by base name"""
    products = Product.query.filter_by(is_deleted=False).all()
    grouped = {}
    
    for product in products:
        base_name = product.name.split(' - ')[0] if ' - ' in product.name else product.name
        if base_name not in grouped:
            grouped[base_name] = {
                'name': base_name,
                'specification': product.specification,
                'product_type': product.product_type,
                'variants': []
            }
        grouped[base_name]['variants'].append({
            'id': product.id,
            'color': product.color,
            'price': float(product.base_price) or 0,
            'stock': product.stock or 0
        })
    
    return list(grouped.values())


def notify_attendance_marked(attendance):
    """Notify when attendance is marked"""
    broadcast_update('attendance_updated', {
        'employee': attendance.employee.name,
        'date': attendance.date.isoformat(),
        'status': attendance.status,
        'hours': float(attendance.hours_worked)
    })


def check_and_create_stock_alerts():
    """Check all products and create alerts for low stock"""
    low_stock_products = Product.query.filter(Product.stock < 10000, Product.is_deleted == False).all()
    
    for product in low_stock_products:
        existing_alert = StockAlert.query.filter_by(
            product_id=product.id,
            is_resolved=False
        ).first()
        if not existing_alert:
            alert = StockAlert(
                product_id=product.id,
                threshold=10000,
                current_stock=int(product.stock)
            )
            db.session.add(alert)
            notify_low_stock_alert(product)
    
    db.session.commit()


# ============================================================================
# AUTHENTICATION
# ============================================================================

@main_bp.route('/', methods=['GET', 'POST'])
def index():
    grouped_products = group_products_by_base()
    return render_template('customer-site.html', products=grouped_products)

@main_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            flash('Username and password are required.', 'error')
            return render_template('auth/login.html')

        user = User.query.filter_by(username=username).first()

        # Account lockout check
        if user and user.is_locked():
            from datetime import timezone
            remaining = int((user.locked_until - datetime.utcnow()).total_seconds() / 60) + 1
            flash(f'Account locked due to too many failed attempts. Try again in {remaining} minute(s).', 'error')
            return render_template('auth/login.html')

        if user and user.is_active and user.check_password(password):
            user.record_successful_login()
            db.session.commit()
            remember = request.form.get('remember') == 'on'
            session.permanent = True
            login_user(user, remember=remember)
            if user.must_change_password:
                flash('You must change your password before continuing.', 'warning')
                return redirect(url_for('main.change_password'))
            next_page = request.args.get('next')
            return redirect(next_page or url_for('main.dashboard'))
        else:
            # Record failed attempt if user exists
            if user:
                user.record_failed_login()
                db.session.commit()
                if user.is_locked():
                    flash('Too many failed attempts. Account locked for 15 minutes.', 'error')
                    return render_template('auth/login.html')
            flash('Invalid username or password.', 'error')

    return render_template('auth/login.html')


@main_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))


# ============================================================================
# CHANGE PASSWORD
# ============================================================================

@main_bp.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')

        if not current_user.check_password(current_pw):
            flash('Current password is incorrect.', 'error')
        elif len(new_pw) < 8:
            flash('New password must be at least 8 characters.', 'error')
        elif new_pw != confirm_pw:
            flash('Passwords do not match.', 'error')
        elif new_pw == current_pw:
            flash('New password must be different from the current password.', 'error')
        else:
            current_user.set_password(new_pw)
            current_user.must_change_password = False
            db.session.commit()
            flash('Password changed successfully.', 'success')
            return redirect(url_for('main.dashboard'))

    return render_template('auth/change_password.html')


# ============================================================================
# USER MANAGEMENT (admin only)
# ============================================================================

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated


@main_bp.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=users)


@main_bp.route('/admin/users/create', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_create_user():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', 'user')
        must_change = request.form.get('must_change_password') == 'on'

        errors = []
        if not username:
            errors.append('Username is required.')
        elif User.query.filter_by(username=username).first():
            errors.append('Username already exists.')
        if not email:
            errors.append('Email is required.')
        elif User.query.filter_by(email=email).first():
            errors.append('Email already exists.')
        if len(password) < 8:
            errors.append('Password must be at least 8 characters.')
        if role not in ('admin', 'user', 'manager'):
            errors.append('Invalid role.')

        if errors:
            for e in errors:
                flash(e, 'error')
        else:
            new_user = User(
                username=username,
                email=email,
                role=role,
                is_active=True,
                must_change_password=must_change,
                failed_login_attempts=0,
            )
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            flash(f'User "{username}" created successfully.', 'success')
            return redirect(url_for('main.admin_users'))

    return render_template('admin/create_user.html')


@main_bp.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'error')
    else:
        user.is_active = not user.is_active
        db.session.commit()
        state = 'activated' if user.is_active else 'deactivated'
        flash(f'User "{user.username}" {state}.', 'success')
    return redirect(url_for('main.admin_users'))


@main_bp.route('/admin/users/<int:user_id>/unlock', methods=['POST'])
@login_required
@admin_required
def admin_unlock_user(user_id):
    user = User.query.get_or_404(user_id)
    user.locked_until = None
    user.failed_login_attempts = 0
    db.session.commit()
    flash(f'User "{user.username}" unlocked.', 'success')
    return redirect(url_for('main.admin_users'))


@main_bp.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id):
    user = User.query.get_or_404(user_id)
    new_pw = request.form.get('new_password', '')
    if len(new_pw) < 8:
        flash('Password must be at least 8 characters.', 'error')
    else:
        user.set_password(new_pw)
        user.must_change_password = True
        user.failed_login_attempts = 0
        user.locked_until = None
        db.session.commit()
        flash(f'Password reset for "{user.username}". They must change it on next login.', 'success')
    return redirect(url_for('main.admin_users'))


# ============================================================================
# DASHBOARD
# ============================================================================

# ✅ CORRECT: Only listens to /dashboard
@main_bp.route('/dashboard')
@login_required
def dashboard():
    check_and_create_stock_alerts()
    
    total_revenue = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing'])
    ).scalar() or 0
    
    total_orders = Order.query.count()
    active_inquiries = Inquiry.query.filter(
        Inquiry.status.in_(['new', 'followup']),
        Inquiry.is_deleted == False
    ).count()
    
    total_customers = Customer.query.filter_by(is_active=True, is_deleted=False).count()
    
    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(5).all()
    recent_inquiries = Inquiry.query.filter_by(is_deleted=False).order_by(Inquiry.created_at.desc()).limit(5).all()
    
    return render_template('dashboard.html',
        total_revenue=total_revenue,
        total_orders=total_orders,
        active_inquiries=active_inquiries,
        total_customers=total_customers,
        recent_orders=recent_orders,
        recent_inquiries=recent_inquiries)

# ============================================================================
# PRODUCTS
# ============================================================================

@main_bp.route('/products')
@login_required
def products():
    products = Product.query.filter_by(is_deleted=False).all()
    low_stock_products = [p for p in products if p.stock < 10000]
    return render_template('products/index.html', products=products, low_stock_products=low_stock_products)


@main_bp.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    if request.method == 'POST':
        try:
            base_price = int(request.form.get('base_price', 0) or 0)
            
            product = Product(
                name=request.form.get('name', ''),
                specification=request.form.get('specification', ''),
                
                # ✅ NEW DIMENSION FIELDS:
                body_diameter=float(request.form.get('body_diameter', 0) or 0),
                overall_length=float(request.form.get('overall_length', 0) or 0),
                sealing_point=float(request.form.get('sealing_point', 0) or 0),
                body_length=float(request.form.get('body_length', 0) or 0),
                stem_diameter=float(request.form.get('stem_diameter', 0) or 0),
                wall_thickness=float(request.form.get('wall_thickness', 0) or 0),
                
                # ✅ PRICING: Store as per 1000, calculate per unit
                base_price=base_price,  # Price per 1000 units
                price_per_unit=base_price / 1000,  # Auto-calculated
                
                stock=int(request.form.get('stock', '0') or 0),
                color=request.form.get('color', ''),
                product_type=request.form.get('product_type', 'product'),
                material_type=request.form.get('material_type', ''),
                usp_type=request.form.get('usp_type', ''),
                shape_type=request.form.get('shape_type', ''),
                dimensions=request.form.get('dimensions', ''),
                use_case=request.form.get('use_case', ''),
                
                # ✅ PAINTING SERVICE FIELDS:
                paint_color=request.form.get('paint_color', ''),
                paint_type=request.form.get('paint_type', ''),
                printing_method=request.form.get('printing_method', '')
            )
            db.session.add(product)
            db.session.commit()
            
            # Handle stock adjustment if ...:
            # Note: adj_type and adj_qty typically come from a specific form field if implemented in the add product form.
            if request.form.get('adjustment_type') and request.form.get('quantity'):
                adj_type = request.form.get('adjustment_type')
                adj_qty = request.form.get('quantity')
                adjustment = StockAdjustment(
                    product_id=product.id,
                    adjustment_type=adj_type,
                    quantity=int(adj_qty),
                    reason=request.form.get('adjustment_reason', 'Initial stock'),
                    adjusted_by=current_user.username
                )
                db.session.add(adjustment)
                
                if adj_type == 'add':
                    product.stock += int(adj_qty)
                elif adj_type == 'remove':
                    product.stock -= int(adj_qty)
                elif adj_type == 'correction':
                    product.stock = int(adj_qty)
                
                db.session.commit()
            
            if product.stock < 10000:
                notify_low_stock_alert(product)
            
            flash('Product added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.products'))
    
    return render_template('products/form.html', product=None)


@main_bp.route('/products/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            base_price = int(request.form.get('base_price', 0) or 0)
            
            product.name = request.form.get('name', '')
            product.specification = request.form.get('specification', '')
            
            # ✅ UPDATE DIMENSION FIELDS:
            product.body_diameter = float(request.form.get('body_diameter', 0) or 0)
            product.overall_length = float(request.form.get('overall_length', 0) or 0)
            product.sealing_point = float(request.form.get('sealing_point', 0) or 0)
            product.body_length = float(request.form.get('body_length', 0) or 0)
            product.stem_diameter = float(request.form.get('stem_diameter', 0) or 0)
            product.wall_thickness = float(request.form.get('wall_thickness', 0) or 0)
            
            # ✅ UPDATE PRICING:
            product.base_price = base_price
            product.price_per_unit = base_price / 1000
            
            product.stock = int(request.form.get('stock', '0') or 0)
            product.color = request.form.get('color', '')
            product.product_type = request.form.get('product_type', 'product')
            product.material_type = request.form.get('material_type', '')
            product.usp_type = request.form.get('usp_type', '')
            product.shape_type = request.form.get('shape_type', '')
            product.dimensions = request.form.get('dimensions', '')
            product.use_case = request.form.get('use_case', '')
            
            # ✅ UPDATE PAINTING SERVICE FIELDS:
            product.paint_color = request.form.get('paint_color', '')
            product.paint_type = request.form.get('paint_type', '')
            product.printing_method = request.form.get('printing_method', '')
            
            db.session.commit()
            
            if product.stock < 10000:
                notify_low_stock_alert(product)
            
            flash('Product updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.products'))
    
    return render_template('products/form.html', product=product)

@main_bp.route('/products/delete/<int:id>')
@login_required
def delete_product(id):
    try:
        product = Product.query.get_or_404(id)
        product.is_deleted = True
        db.session.commit()
        flash('Product archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.products'))


@main_bp.route('/products/stock-adjust/<int:id>', methods=['POST'])
@login_required
def adjust_stock(id):
    product = Product.query.get_or_404(id)
    
    try:
        adj_type = request.form.get('adjustment_type', '')
        qty = int(request.form.get('quantity', '0') or 0)
        reason = request.form.get('reason', '')
        
        adjustment = StockAdjustment(
            product_id=product.id,
            adjustment_type=adj_type,
            quantity=qty,
            reason=reason,
            adjusted_by=current_user.username
        )
        db.session.add(adjustment)
        
        if adj_type == 'add':
            product.stock += qty
        elif adj_type == 'remove':
            product.stock -= qty
        elif adj_type == 'correction':
            product.stock = qty
        
        db.session.commit()
        
        if product.stock < 10000:
            notify_low_stock_alert(product)
        
        flash('Stock adjusted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.products'))


# ============================================================================
# INQUIRIES
# ============================================================================

@main_bp.route('/inquiries')
@login_required
def inquiries():
    status = request.args.get('status', 'all')
    show_deleted = request.args.get('show_deleted', 'false') == 'true'
    
    query = Inquiry.query
    if not show_deleted:
        query = query.filter_by(is_deleted=False)
    if status != 'all':
        query = query.filter_by(status=status)
    
    inquiries = query.order_by(Inquiry.created_at.desc()).all()
    return render_template('inquiries/index.html',
                         inquiries=inquiries,
                         current_status=status,
                         show_deleted=show_deleted)


@main_bp.route('/inquiries/add', methods=['GET', 'POST'])
@login_required
def add_inquiry():
    if request.method == 'POST':
        # DEBUG: Print what's being received
        print("="*50)
        print("FORM DATA RECEIVED:")
        print(f"customer_name: {request.form.get('customer_name')}")
        print(f"product_id[]: {request.form.getlist('product_id[]')}")
        print(f"quantity[]: {request.form.getlist('quantity[]')}")
        print("="*50)
        
        try:
            inquiry_number = f"INQ-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            
            # Get or create customer
            customer, is_new = find_or_create_customer(
                customer_name=request.form.get('customer_name', ''),
                business_name=request.form.get('business_name', ''),
                email=request.form.get('email', ''),
                phone=request.form.get('phone', '')
            )
            
            # Create inquiry (without product_id - products will be in InquiryItem)
            inquiry = Inquiry(
                inquiry_number=inquiry_number,
                customer_name=customer.name,
                business_name=customer.business_name,
                email=customer.email,
                phone=customer.phone,
                product_id=None,  # Will be null, products are in InquiryItem
                quantity=0,  # Will be calculated from items
                notes=request.form.get('notes', ''),
                status='new'
            )
            db.session.add(inquiry)
            db.session.flush()  # Get inquiry ID before committing
            
            # ✅ Handle multiple products
            product_ids = request.form.getlist('product_id[]')
            quantities = request.form.getlist('quantity[]')
            
            total_quantity = 0
            
            for i, pid in enumerate(product_ids):
                if pid and quantities[i]:
                    qty = int(quantities[i]) if quantities[i] else 0
                    if qty > 0:
                        # Create InquiryItem for each product
                        item = InquiryItem(
                            inquiry_id=inquiry.id,
                            product_id=int(pid),
                            quantity=qty
                        )
                        db.session.add(item)
                        total_quantity += qty
            
            # Update inquiry total quantity
            inquiry.quantity = total_quantity
            db.session.commit()
            
            notify_new_inquiry(inquiry)
            flash('Inquiry created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.inquiries'))
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('inquiries/form.html', inquiry=None, products=products)

@main_bp.route('/inquiries/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_inquiry(id):
    inquiry = Inquiry.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            inquiry.customer_name = request.form.get('customer_name', '')
            inquiry.business_name = request.form.get('business_name', '')
            inquiry.email = request.form.get('email', '')
            inquiry.phone = request.form.get('phone', '')
            inquiry.status = request.form.get('status', 'new')
            inquiry.notes = request.form.get('notes', '')
            
            for item in inquiry.inquiry_items:
                qty_field = f'qty_{item.id}'
                if qty_field in request.form:
                    new_qty = int(request.form.get(qty_field, item.quantity))
                    if new_qty >= 0:
                        item.quantity = new_qty
            
            inquiry.quantity = sum(item.quantity for item in inquiry.inquiry_items)
            db.session.commit()
            flash('Inquiry updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.edit_inquiry', id=id))
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('inquiries/edit.html', inquiry=inquiry, products=products)


@main_bp.route('/inquiries/delete/<int:id>')
@login_required
def delete_inquiry(id):
    try:
        inquiry = Inquiry.query.get_or_404(id)
        inquiry.is_deleted = True
        inquiry.deleted_at = datetime.utcnow()
        db.session.commit()
        flash('Inquiry archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.inquiries'))


@main_bp.route('/inquiries/process-invoice/<int:id>', methods=['GET', 'POST'])
@login_required
def process_invoice(id):
    """Convert inquiry to order"""
    inquiry = Inquiry.query.get_or_404(id)

    if request.method == 'POST':
        try:
            order_number = f"ORD-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"

            customer, is_new = find_or_create_customer(
                customer_name=inquiry.customer_name,
                business_name=inquiry.business_name,
                email=inquiry.email,
                phone=inquiry.phone
            )

            if is_new:
                flash(f'New customer created: {customer.name}', 'info')

            order = Order(
                order_number=order_number,
                customer_id=customer.id,
                inquiry_id=inquiry.id,
                total_amount=0,  # Will be updated
                tax_amount=0,  # Will be updated
                status='pending',
                payment_status='unpaid',
                customer_name_snapshot=customer.name,
                customer_business_snapshot=customer.business_name,
                customer_phone_snapshot=customer.phone,
                customer_email_snapshot=customer.email
            )
            db.session.add(order)
            db.session.flush()

            total_amount = 0

            for item in inquiry.inquiry_items:
                qty_field = f'qty_{item.id}'
                adjusted_qty = int(request.form.get(qty_field, item.quantity))

                if item.product and adjusted_qty > item.product.stock:
                    flash(f'Stock for {item.product.name} is low. Adjusted quantity to {item.product.stock}.', 'warning')
                    adjusted_qty = item.product.stock

                agreed_price_obj = CustomerProductPrice.query.filter_by(customer_id=customer.id, product_id=item.product_id).first()
                unit_price = agreed_price_obj.agreed_price if agreed_price_obj else (item.product.unit_price if item.product else 0)
                subtotal = unit_price * adjusted_qty
                total_amount += subtotal

                order_item = OrderItem(
                    order_id=order.id,
                    product_id=item.product_id,
                    quantity=adjusted_qty,
                    unit_price=unit_price,
                    subtotal=subtotal
                )
                db.session.add(order_item)

                if item.product:
                    item.product.stock -= adjusted_qty

            tax_rate_setting = CompanySetting.query.filter_by(key='company_gst').first()
            tax_rate = float(tax_rate_setting.value.strip('%')) / 100 if tax_rate_setting and tax_rate_setting.value else 0.17

            order.total_amount = total_amount * (1 + tax_rate)
            order.tax_amount = total_amount * tax_rate

            inquiry.status = 'converted'

            db.session.commit()
            notify_new_order(order)
            flash(f'Order {order.order_number} created from inquiry.', 'success')
            return redirect(url_for('main.view_order', id=order.id))

        except Exception as e:
            db.session.rollback()
            print(f"ERROR: {str(e)}")
            flash(f'Error processing invoice: {str(e)}', 'error')
            return redirect(url_for('main.edit_inquiry', id=id))

    stock_issues = []
    for item in inquiry.inquiry_items:
        if item.product and item.quantity > item.product.stock:
            stock_issues.append({
                'product': item.product.name,
                'requested': item.quantity,
                'available': item.product.stock
            })

    return render_template('inquiries/invoice.html',
                         inquiry=inquiry,
                         stock_issues=stock_issues)


# ============================================================================
# ORDERS
# ============================================================================

@main_bp.route('/orders')
@login_required
def orders():
    status = request.args.get('status', 'all')
    query = Order.query
    if status != 'all':
        query = query.filter_by(status=status)
    orders = query.order_by(Order.created_at.desc()).all()
    return render_template('orders/index.html', orders=orders, current_status=status)


@main_bp.route('/orders/add', methods=['GET', 'POST'])
@login_required
def add_order():
    if request.method == 'POST':
        try:
            order_number = f"ORD-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            unit_price = float(request.form.get('unit_price', 0))
            quantity = int(request.form.get('quantity', '0') or 0)
            subtotal = unit_price * quantity
            tax_amount = subtotal * 0.17
            total_amount = subtotal + tax_amount
            product_id = request.form.get('product_id')
            
            customer_id = request.form.get('customer_id')
            customer = None
            
            if customer_id:
                customer = Customer.query.get(int(customer_id))
            else:
                customer_name = request.form.get('customer_name', '')
                business_name = request.form.get('business_name', '')
                email = request.form.get('email', '')
                phone = request.form.get('phone', '')
                
                if customer_name or business_name:
                    customer, is_new = find_or_create_customer(
                        customer_name=customer_name,
                        business_name=business_name,
                        email=email,
                        phone=phone
                    )
                    customer_id = customer.id
            
            order = Order(
                order_number=order_number,
                customer_id=int(customer_id) if customer_id else None,
                total_amount=total_amount,
                tax_amount=tax_amount,
                status='pending',
                payment_status='unpaid',
                customer_name_snapshot=customer.name if customer else customer_name
            )
            db.session.add(order)
            db.session.flush()

            if product_id:
                order_item = OrderItem(
                    order_id=order.id,
                    product_id=int(product_id),
                    quantity=quantity,
                    unit_price=unit_price,
                    subtotal=subtotal
                )
                db.session.add(order_item)
            
            db.session.commit()
            
            if customer:
                customer.current_balance += total_amount
                db.session.commit()
            
            notify_new_order(order)
            flash('Order created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.orders'))
    
    customers = Customer.query.filter_by(is_active=True).all()
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('orders/form.html', customers=customers, products=products)


@main_bp.route('/orders/view/<int:id>')
@login_required
def view_order(id):
    order = Order.query.get_or_404(id)
    approval_request = OrderApproval.query.filter_by(order_id=id, status='pending').first()
    return render_template('orders/view.html', order=order, approval_request=approval_request)


@main_bp.route('/orders/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_order(id):
    order = Order.query.get_or_404(id)
    if request.method == 'POST':
        try:
            old_status = order.status
            order.status = request.form.get('status', order.status)
            order.payment_status = request.form.get('payment_status', order.payment_status)
            order.payment_method = request.form.get('payment_method', '')
            
            if order.payment_status == 'paid':
                order.paid_amount = order.total_amount
                order.payment_date = datetime.utcnow()
            elif order.payment_status == 'partial':
                paid_amount = float(request.form.get('paid_amount', '0') or 0)
                order.paid_amount = paid_amount
                order.payment_date = datetime.utcnow()
            
            db.session.commit()
            
            # ✅ AUTO-CREATE JOURNAL ENTRY WHEN ORDER IS COMPLETED
            if old_status != 'completed' and order.status == 'completed':
                create_journal_entry_for_order(order)
            
            # ✅ AUTO-CREATE JOURNAL ENTRY WHEN PAYMENT IS RECEIVED
            if order.payment_status == 'paid':
                # Create a pseudo-payment record for journal
                from app.models import PayrollPayment  # Temporary workaround
                # Better: Create a Payment model for customer payments
            
            flash('Order updated successfully!', 'success')
            return redirect(url_for('main.view_order', id=order.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating order: {str(e)}', 'error')
    return render_template('orders/edit.html', order=order)

@main_bp.route('/orders/invoice/<int:id>')
@login_required
def order_invoice(id):
    order = Order.query.get_or_404(id)
    logo_setting = CompanySetting.query.filter_by(key='company_logo').first()
    logo_path = logo_setting.value if logo_setting else None
    return render_template('orders/invoice.html', order=order, logo_path=logo_path)


@main_bp.route('/orders/delete/<int:id>')
@login_required
def delete_order(id):
    try:
        order = Order.query.get_or_404(id)
        db.session.delete(order)
        db.session.commit()
        flash('Order deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.orders'))


# ============================================================================
# CUSTOMERS
# ============================================================================

@main_bp.route('/customers')
@login_required
def customers():
    show_inactive = request.args.get('show_inactive', 'false') == 'true'
    query = Customer.query
    if not show_inactive:
        query = query.filter_by(is_active=True, is_deleted=False)
    customers = query.all()
    
    for customer in customers:
        total_orders = sum(o.total_amount for o in customer.orders if o.status != 'cancelled')
    return render_template('customers/index.html', customers=customers, show_inactive=show_inactive)


@main_bp.route('/customers/add', methods=['GET', 'POST'])
@login_required
def add_customer():
    if request.method == 'POST':
        try:
            customer = Customer(
                name=request.form.get('name', ''),
                business_name=request.form.get('business_name', ''),
                role=request.form.get('role', ''),
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''),
                address=request.form.get('address', ''),
                payment_terms=request.form.get('payment_terms', 'cash'),
                credit_limit=float(request.form.get('credit_limit', 0))
            )
            db.session.add(customer)
            db.session.commit()
            flash('Customer added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.customers'))
    
    return render_template('customers/form.html', customer=None)


@main_bp.route('/customers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            customer.name = request.form.get('name', '')
            customer.business_name = request.form.get('business_name', '')
            customer.role = request.form.get('role', '')
            customer.email = request.form.get('email', '')
            customer.phone = request.form.get('phone', '')
            customer.address = request.form.get('address', '')
            customer.payment_terms = request.form.get('payment_terms', 'cash')
            customer.is_active = request.form.get('is_active') == 'on'
            customer.credit_limit = float(request.form.get('credit_limit', 0))
            
            db.session.commit()
            flash('Customer updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.customers'))
    
    return render_template('customers/form.html', customer=customer)


@main_bp.route('/customers/delete/<int:id>')
@login_required
def delete_customer(id):
    try:
        customer = Customer.query.get_or_404(id)
        customer.is_deleted = True
        customer.is_active = False
        db.session.commit()
        flash('Customer archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.customers'))


@main_bp.route('/customers/painting-pricing/<int:id>', methods=['GET', 'POST'])
@login_required
def customer_painting_pricing(id):
    """Manage customer-specific painting prices"""
    customer = Customer.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            product_ids = request.form.getlist('product_id[]')
            prices = request.form.getlist('painting_price[]')
            setup_charges = request.form.getlist('setup_charge[]')
            payment_terms = request.form.getlist('painting_payment_terms[]')
            
            for i, pid in enumerate(product_ids):
                existing = CustomerPaintingPrice.query.filter_by(
                    customer_id=customer.id,
                    product_id=int(pid)
                ).first()
                
                price_val = float(prices[i]) if i < len(prices) and prices[i] else 0
                setup_charge_val = float(setup_charges[i]) if i < len(setup_charges) and setup_charges[i] else 0
                payment_terms_val = payment_terms[i] if i < len(payment_terms) and payment_terms[i] else None

                if existing:
                    existing.price_per_unit = price_val
                    existing.setup_charge = setup_charge_val
                    existing.payment_terms = payment_terms_val
                else:
                    cpp = CustomerPaintingPrice(
                        customer_id=customer.id,
                        product_id=int(pid),
                        price_per_unit=price_val,
                        setup_charge=setup_charge_val,
                        payment_terms=payment_terms_val
                    )
                    db.session.add(cpp)
            db.session.commit()
            flash('Customer painting prices updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.edit_customer', id=id))
    
    products = Product.query.filter_by(is_deleted=False).all()
    
    painting_prices = {p.product_id: p for p in PaintingServicePrice.query.all()}
    customer_painting_prices = {cp.product_id: cp for cp in CustomerPaintingPrice.query.filter_by(customer_id=id).all()}
    return render_template('customers/painting_pricing.html', 
                         customer=customer, 
                         products=products, 
                         painting_prices=painting_prices,
                         customer_painting_prices=customer_painting_prices)

@main_bp.route('/customers/merge', methods=['GET', 'POST'])
@login_required
def merge_customers():
    """Merge duplicate customers"""
    if request.method == 'POST':
        try:
            primary_id = request.form.get('primary_customer_id')
            merge_ids = request.form.getlist('merge_customer_id[]')
            
            primary_customer = Customer.query.get_or_404(primary_id)
            
            for merge_id in merge_ids:
                merge_customer = Customer.query.get(int(merge_id))
                if merge_customer and merge_customer.id != primary_id:
                    for order in merge_customer.orders:
                        order.customer_id = primary_id
                    
                    primary_customer.current_balance += merge_customer.current_balance
                    merge_customer.is_active = False
                    merge_customer.is_deleted = True
            
            db.session.commit()
            flash(f'Merged {len(merge_ids)} customer(s) into {primary_customer.name}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error merging customers: {str(e)}', 'error')
        
        return redirect(url_for('main.customers'))
    
    potential_duplicates = db.session.query(
        Customer.business_name,
        db.func.count(Customer.id).label('count')
    ).filter(
        Customer.business_name != None,
        Customer.business_name != '',
        Customer.is_deleted == False
    ).group_by(Customer.business_name).having(
        db.func.count(Customer.id) > 1
    ).all()
    
    return render_template('customers/merge.html', potential_duplicates=potential_duplicates)


# ============================================================================
# SUPPLY CHAIN (PHASE 5)
# ============================================================================

@main_bp.route('/suppliers')
@login_required
def suppliers():
    suppliers = Supplier.query.filter_by(is_active=True).all()
    return render_template('suppliers/index.html', suppliers=suppliers)


@main_bp.route('/suppliers/add', methods=['GET', 'POST'])
@login_required
def add_supplier():
    if request.method == 'POST':
        try:
            supplier = Supplier(
                name=request.form.get('name', ''),
                contact_person=request.form.get('contact_person', ''),
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''),
                address=request.form.get('address', ''),
                city=request.form.get('city', ''),
                country=request.form.get('country', 'Pakistan'),
                cnic=request.form.get('cnic', ''),
                ntn=request.form.get('ntn', ''),
                payment_terms=request.form.get('payment_terms', '30 days'),
                credit_limit=float(request.form.get('credit_limit', 0))
            )
            db.session.add(supplier)
            db.session.commit()
            flash('Supplier added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.suppliers'))
    
    return render_template('suppliers/form.html', supplier=None)


@main_bp.route('/suppliers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            supplier.name = request.form.get('name', '')
            supplier.contact_person = request.form.get('contact_person', '')
            supplier.email = request.form.get('email', '')
            supplier.phone = request.form.get('phone', '')
            supplier.address = request.form.get('address', '')
            supplier.city = request.form.get('city', '')
            supplier.country = request.form.get('country', 'Pakistan')
            supplier.cnic = request.form.get('cnic', '')
            supplier.ntn = request.form.get('ntn', '')
            supplier.payment_terms = request.form.get('payment_terms', '30 days')
            supplier.credit_limit = float(request.form.get('credit_limit', 0))
            supplier.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash('Supplier updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.suppliers'))
    
    return render_template('suppliers/form.html', supplier=supplier)


@main_bp.route('/suppliers/delete/<int:id>')
@login_required
def delete_supplier(id):
    try:
        supplier = Supplier.query.get_or_404(id)
        supplier.is_active = False
        db.session.commit()
        flash('Supplier archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.suppliers'))


@main_bp.route('/suppliers/<int:id>/purchase-history')
@login_required
def supplier_purchase_history(id):
    supplier = Supplier.query.get_or_404(id)
    purchase_orders = PurchaseOrder.query.filter_by(supplier_id=id).order_by(PurchaseOrder.order_date.desc()).all()
    
    total_orders = len(purchase_orders)
    total_spent = sum(po.grand_total for po in purchase_orders)
    avg_order_value = total_spent / total_orders if total_orders > 0 else 0
    return render_template('suppliers/history.html', 
                         supplier=supplier, 
                         purchase_orders=purchase_orders,
                         total_spent=total_spent,
                         total_orders=total_orders,
                         avg_order_value=avg_order_value)


@main_bp.route('/purchase-orders')
@login_required
def purchase_orders():
    status = request.args.get('status', 'all')
    query = PurchaseOrder.query
    if status != 'all':
        query = query.filter_by(status=status)
    purchase_orders = query.order_by(PurchaseOrder.order_date.desc()).all()
    return render_template('purchase_orders/index.html', purchase_orders=purchase_orders, current_status=status)


@main_bp.route('/purchase-orders/add', methods=['GET', 'POST'])
@login_required
def add_purchase_order():
    if request.method == 'POST':
        try:
            po_number = f"PO-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            supplier_id = request.form.get('supplier_id')
            
            po = PurchaseOrder(
                po_number=po_number,
                supplier_id=int(supplier_id),
                expected_date=datetime.strptime(request.form.get('expected_date', ''), '%Y-%m-%d') if request.form.get('expected_date') else None,
                status='draft'
            )
            db.session.add(po)
            db.session.flush()
            
            material_ids = request.form.getlist('material_id[]')
            quantities = request.form.getlist('quantity[]')
            unit_prices = request.form.getlist('unit_price[]')
            tax_rates = request.form.getlist('tax_rate[]')

            for i, mid in enumerate(material_ids):
                if mid and quantities[i]:
                    qty = float(quantities[i])
                    price = float(unit_prices[i])
                    tax_rate = float(tax_rates[i]) if tax_rates[i] else 0
                    
                    item = PurchaseOrderItem(
                        po_id=po.id,
                        material_id=int(mid),
                        quantity_ordered=qty,
                        unit_price=price,
                        tax_rate=tax_rate,
                        subtotal=qty * price
                    )
                    db.session.add(item)
            
            db.session.commit()
            flash('Purchase Order created successfully!', 'success')
            return redirect(url_for('main.purchase_orders'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    suppliers = Supplier.query.filter_by(is_active=True).all()
    materials = RawMaterial.query.filter_by(is_active=True).all()
    return render_template('purchase_orders/form.html', suppliers=suppliers, materials=materials, po=None)


@main_bp.route('/purchase-orders/<int:id>')
@login_required
def view_purchase_order(id):
    po = PurchaseOrder.query.get_or_404(id)
    return render_template('purchase_orders/view.html', po=po)


@main_bp.route('/purchase-orders/<int:id>/send', methods=['POST'])
@login_required
def send_purchase_order(id):
    po = PurchaseOrder.query.get_or_404(id)
    po.status = 'sent'
    db.session.commit()
    flash(f'Purchase Order {po.po_number} sent to supplier!', 'success')
    return redirect(url_for('main.view_purchase_order', id=po.id))


@main_bp.route('/purchase-orders/<int:id>/receive', methods=['GET', 'POST'])
@login_required
def receive_purchase_order(id):
    po = PurchaseOrder.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            grn_number = f"GRN-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            warehouse_id = request.form.get('warehouse_id')
            
            gr = GoodsReceipt(
                grn_number=grn_number,
                po_id=po.id,
                supplier_id=po.supplier_id,
                warehouse_id=int(warehouse_id) if warehouse_id else None,
                status='pending'
            )
            db.session.add(gr)
            db.session.flush()
            
            for item in po.items:
                qty_received = float(request.form.get(f'qty_{item.id}', 0))
                qty_accepted = float(request.form.get(f'accepted_{item.id}', qty_received))
                qty_rejected = qty_received - qty_accepted
                batch_number = request.form.get(f'batch_{item.id}', '')
                expiry_date = request.form.get(f'expiry_{item.id}', '')
                
                gri = GoodsReceiptItem(
                    gr_id=gr.id,
                    material_id=item.material_id,
                    quantity_ordered=item.quantity_ordered,
                    quantity_received=qty_received,
                    quantity_accepted=qty_accepted,
                    quantity_rejected=qty_rejected,
                    rejection_reason=request.form.get(f'rejection_reason_{item.id}', ''),
                    batch_number=batch_number,
                    expiry_date=datetime.strptime(expiry_date, '%Y-%m-%d') if expiry_date else None
                )
                db.session.add(gri)
                
                if qty_accepted > 0:
                    material = RawMaterial.query.get(item.material_id)
                    if material:
                        material.current_stock += qty_accepted
                        
                        batch = MaterialBatch(
                            material_id=item.material_id,
                            batch_number=batch_number or f"BATCH-{datetime.utcnow().strftime('%Y%m%d')}",
                            supplier_id=po.supplier_id,
                            quantity_received=qty_accepted,
                            quantity_remaining=qty_accepted,
                            expiry_date=datetime.strptime(expiry_date, '%Y-%m-%d') if expiry_date else None,
                            warehouse_id=int(warehouse_id) if warehouse_id else None
                        )
                        db.session.add(batch)
                        
            if all(item.quantity_received >= item.quantity_ordered for item in po.items):
                po.status = 'received'
                po.received_date = datetime.utcnow()
            
            db.session.commit()
            flash(f'Goods Receipt {grn_number} created successfully!', 'success')
            return redirect(url_for('main.view_goods_receipt', id=gr.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('purchase_orders/receive.html', po=po)


@main_bp.route('/goods-receipts')
@login_required
def goods_receipts():
    status = request.args.get('status', 'all')
    query = GoodsReceipt.query
    if status != 'all':
        query = query.filter_by(status=status)
    receipts = query.order_by(GoodsReceipt.receipt_date.desc()).all()
    return render_template('goods_receipts/index.html', receipts=receipts, current_status=status)


@main_bp.route('/goods-receipts/<int:id>')
@login_required
def view_goods_receipt(id):
    gr = GoodsReceipt.query.get_or_404(id)
    return render_template('goods_receipts/view.html', gr=gr)


@main_bp.route('/warehouses')
@login_required
def warehouses():
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    return render_template('warehouses/index.html', warehouses=warehouses)


@main_bp.route('/warehouses/add', methods=['GET', 'POST'])
@login_required
def add_warehouse():
    if request.method == 'POST':
        try:
            warehouse = Warehouse(
                name=request.form.get('name', ''),
                code=request.form.get('code', ''),
                address=request.form.get('address', ''),
                city=request.form.get('city', ''),
                manager=request.form.get('manager', ''),
                phone=request.form.get('phone', '')
            )
            db.session.add(warehouse)
            db.session.commit()
            flash('Warehouse added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.warehouses'))
    
    return render_template('warehouses/form.html', warehouse=None)


@main_bp.route('/warehouses/<int:id>/stock')
@login_required
def warehouse_stock(id):
    warehouse = Warehouse.query.get_or_404(id)
    stock_levels = WarehouseStock.query.filter_by(warehouse_id=id).all()
    return render_template('warehouses/stock.html', warehouse=warehouse, stock_levels=stock_levels)


@main_bp.route('/stock-transfers')
@login_required
def stock_transfers():
    status = request.args.get('status', 'all')
    query = StockTransfer.query
    if status != 'all':
        query = query.filter_by(status=status)
    transfers = query.order_by(StockTransfer.transfer_date.desc()).all()
    return render_template('stock_transfers/index.html', transfers=transfers, current_status=status)


@main_bp.route('/stock-transfers/add', methods=['GET', 'POST'])
@login_required
def add_stock_transfer():
    if request.method == 'POST':
        try:
            transfer_number = f"ST-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            from_warehouse_id = request.form.get('from_warehouse_id')
            to_warehouse_id = request.form.get('to_warehouse_id')
            
            transfer = StockTransfer(
                transfer_number=transfer_number,
                from_warehouse_id=int(from_warehouse_id),
                to_warehouse_id=int(to_warehouse_id),
                status='pending',
                notes=request.form.get('notes', ''),
                requested_by=current_user.username
            )
            db.session.add(transfer)
            db.session.flush()
            
            material_ids = request.form.getlist('material_id[]')
            quantities = request.form.getlist('quantity[]')
            
            for i, mid in enumerate(material_ids):
                if mid and quantities[i]:
                    item = StockTransferItem(
                        transfer_id=transfer.id,
                        material_id=int(mid),
                        quantity=float(quantities[i])
                    )
                    db.session.add(item)
            
            db.session.commit()
            flash(f'Stock Transfer {transfer_number} created successfully!', 'success')
            return redirect(url_for('main.view_stock_transfer', id=transfer.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    materials = RawMaterial.query.filter_by(is_active=True).all()
    return render_template('stock_transfers/form.html', warehouses=warehouses, materials=materials, transfer=None)


@main_bp.route('/stock-transfers/<int:id>')
@login_required
def view_stock_transfer(id):
    transfer = StockTransfer.query.get_or_404(id)
    return render_template('stock_transfers/view.html', transfer=transfer)


@main_bp.route('/stock-transfers/<int:id>/complete', methods=['POST'])
@login_required
def complete_stock_transfer(id):
    transfer = StockTransfer.query.get_or_404(id)
    
    try:
        for item in transfer.items:
            from_stock = WarehouseStock.query.filter_by(
                warehouse_id=transfer.from_warehouse_id,
                material_id=item.material_id
            ).first()
            
            if from_stock:
                from_stock.quantity -= item.quantity
                from_stock.available_quantity = from_stock.quantity - from_stock.reserved_quantity
            
            to_stock = WarehouseStock.query.filter_by(
                warehouse_id=transfer.to_warehouse_id,
                material_id=item.material_id
            ).first()
            
            if to_stock:
                to_stock.quantity += item.quantity
                to_stock.available_quantity = to_stock.quantity - to_stock.reserved_quantity
            else:
                to_stock = WarehouseStock(
                    warehouse_id=transfer.to_warehouse_id,
                    material_id=item.material_id,
                    quantity=item.quantity,
                    available_quantity=item.quantity
                )
                db.session.add(to_stock)
            
            item.quantity_transferred = item.quantity
        
        transfer.status = 'completed'
        transfer.completed_at = datetime.utcnow()
        db.session.commit()
        flash('Stock transfer completed successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.view_stock_transfer', id=transfer.id))


@main_bp.route('/material-batches')
@login_required
def material_batches():
    batches = MaterialBatch.query.order_by(MaterialBatch.received_date.desc()).all()
    expiring_soon = [b for b in batches if b.expiry_date and b.expiry_date <= datetime.utcnow().date() + timedelta(days=30)]
    expired = [b for b in batches if b.expiry_date and b.expiry_date < datetime.utcnow().date()]
    return render_template('material_batches/index.html', 
                         batches=batches, 
                         expiring_soon=expiring_soon,
                         expired=expired)


@main_bp.route('/supplier-invoices')
@login_required
def supplier_invoices():
    status = request.args.get('status', 'all')
    query = SupplierInvoice.query
    if status != 'all':
        query = query.filter_by(status=status)
    invoices = query.order_by(SupplierInvoice.created_at.desc()).all()
    return render_template('supplier_invoices/index.html', invoices=invoices, current_status=status)


@main_bp.route('/supplier-invoices/add', methods=['GET', 'POST'])
@login_required
def add_supplier_invoice():
    if request.method == 'POST':
        try:
            invoice = SupplierInvoice(
                invoice_number=request.form.get('invoice_number', ''),
                supplier_id=int(request.form.get('supplier_id', '0') or 0),
                po_id=int(request.form.get('po_id', '0') or 0) if request.form.get('po_id') else None,
                invoice_date=datetime.strptime(request.form.get('invoice_date'), '%Y-%m-%d').date(),
                total_amount=float(request.form.get('total_amount', 0))
            )
            db.session.add(invoice)
            db.session.commit()
            flash('Supplier invoice added', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.supplier_invoices'))
    
    suppliers = Supplier.query.filter_by(is_active=True).all()
    purchase_orders = PurchaseOrder.query.filter_by(status='received').all()
    return render_template('supplier_invoices/form.html', suppliers=suppliers, purchase_orders=purchase_orders, invoice=None)


# ============================================================================
# PRODUCTION/MANUFACTURING (PHASE 3)
# ============================================================================

@main_bp.route('/production')
@login_required
def production():
    active_batches = ProductionBatch.query.filter_by(status='in_progress').all()
    planned_batches = ProductionBatch.query.filter_by(status='planned').all()
    completed_today = ProductionBatch.query.filter(
        ProductionBatch.status == 'completed',
        db.func.date(ProductionBatch.end_date) == datetime.utcnow().date()
    ).all()
    today_production = sum(b.actual_quantity or 0 for b in completed_today)
    low_stock_materials = RawMaterial.query.filter(RawMaterial.current_stock < RawMaterial.reorder_level).all()
    return render_template('production/dashboard.html', 
                         active_batches=active_batches,
                         planned_batches=planned_batches,
                         today_production=today_production,
                         low_stock_materials=low_stock_materials)


@main_bp.route('/production/batches')
@login_required
def production_batches():
    status = request.args.get('status', 'all')
    query = ProductionBatch.query
    if status != 'all':
        query = query.filter_by(status=status)
    batches = query.order_by(ProductionBatch.created_at.desc()).all()
    return render_template('production/batches.html', batches=batches, current_status=status)


@main_bp.route('/production/batch/add', methods=['GET', 'POST'])
@login_required
def add_production_batch():
    if request.method == 'POST':
        try:
            batch_number = f"BATCH-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            product_id = request.form.get('product_id')
            planned_quantity = int(request.form.get('planned_quantity', '0') or 0)
            
            batch = ProductionBatch(
                batch_number=batch_number,
                product_id=int(product_id),
                planned_quantity=planned_quantity,
                status='planned',
                created_by=current_user.username,
                notes=request.form.get('notes', '')
            )
            db.session.add(batch)
            db.session.commit()
            flash(f'Production batch {batch_number} created successfully!', 'success')
            return redirect(url_for('main.production_batches'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('production/batch_form.html', batch=None, products=products)


@main_bp.route('/production/batch/<int:id>')
@login_required
def view_production_batch(id):
    batch = ProductionBatch.query.get_or_404(id)
    return render_template('production/batch_view.html', batch=batch)


@main_bp.route('/production/batch/<int:id>/start', methods=['POST'])
@login_required
def start_production_batch(id):
    batch = ProductionBatch.query.get_or_404(id)
    batch.status = 'in_progress'
    batch.start_date = datetime.utcnow()
    db.session.commit()
    flash('Production batch started!', 'success')
    return redirect(url_for('main.view_production_batch', id=batch.id))


@main_bp.route('/production/batch/<int:id>/complete', methods=['GET', 'POST'])
@login_required
def complete_production_batch(id):
    batch = ProductionBatch.query.get_or_404(id)
    if request.method == 'POST':
        try:
            actual_quantity = int(request.form.get('actual_quantity', '0') or 0)
            rejected_quantity = int(request.form.get('rejected_quantity', '0') or 0)
            batch.actual_quantity = actual_quantity
            batch.rejected_quantity = rejected_quantity
            if batch.planned_quantity > 0:
                batch.yield_percentage = (actual_quantity / batch.planned_quantity) * 100
            batch.status = 'completed'
            batch.end_date = datetime.utcnow()
            product = Product.query.get(batch.product_id)
            if product:
                product.stock += actual_quantity
            
            db.session.commit()
            
            # ✅ AUTO-CREATE JOURNAL ENTRY FOR PRODUCTION
            create_journal_entry_for_production_batch(batch)
            
            flash('Production batch completed!', 'success')
            return redirect(url_for('main.production_batches'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    return render_template('production/batch_complete.html', batch=batch)

@main_bp.route('/production/raw-materials')
@login_required
def raw_materials():
    materials = RawMaterial.query.filter_by(is_active=True).all()
    low_stock = [m for m in materials if m.current_stock < m.reorder_level]
    return render_template('materials/index.html', materials=materials, low_stock=low_stock)


@main_bp.route('/raw-materials/add', methods=['GET', 'POST'])
@login_required
def add_raw_material():
    if request.method == 'POST':
        try:
            material = RawMaterial(
                name=request.form.get('name', ''),
                material_type=request.form.get('material_type', ''),
                specification=request.form.get('specification', ''),
                current_stock=float(request.form.get('current_stock', 0)),
                unit=request.form.get('unit', ''),
                reorder_level=float(request.form.get('reorder_level', 0)),
                cost_per_unit=float(request.form.get('cost_per_unit', 0)),
                supplier=request.form.get('supplier', '')
            )
            db.session.add(material)
            db.session.commit()
            flash('Raw material added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.raw_materials'))
    
    return render_template('materials/form.html', material=None)


@main_bp.route('/products/<int:id>/bom', methods=['GET', 'POST'])
@login_required
def manage_product_bom(id):
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            BOMItem.query.filter_by(product_id=product.id).delete()
            
            material_ids = request.form.getlist('material_id[]')
            quantities = request.form.getlist('quantity_required[]')
            waste_percentages = request.form.getlist('waste_percentage[]')
            
            for i, mid in enumerate(material_ids):
                if mid and quantities[i]:
                    bom_item = BOMItem(
                        product_id=product.id,
                        raw_material_id=int(mid),
                        quantity_required=float(quantities[i]),
                        unit='meters',
                        waste_percentage=float(waste_percentages[i]) if waste_percentages[i] else 0
                    )
                    db.session.add(bom_item)
            db.session.commit()
            flash('BOM updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.manage_product_bom', id=product.id))
    
    materials = RawMaterial.query.filter_by(is_active=True).all()
    bom_items = BOMItem.query.filter_by(product_id=product.id).all()
    return render_template('products/bom.html', product=product, materials=materials, bom_items=bom_items)


@main_bp.route('/production/reports')
@login_required
def production_reports():
    total_batches = ProductionBatch.query.count()
    completed_batches = ProductionBatch.query.filter_by(status='completed').count()
    avg_yield = db.session.query(db.func.avg(ProductionBatch.yield_percentage)).filter(
        ProductionBatch.status == 'completed'
    ).scalar() or 0
    
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    production_by_product = db.session.query(
        Product.name,
        db.func.sum(ProductionBatch.actual_quantity)
    ).join(ProductionBatch).filter(
        ProductionBatch.status == 'completed',
        ProductionBatch.end_date >= thirty_days_ago
    ).group_by(Product.name).all()
    
    return render_template('production/reports.html',
                         total_batches=total_batches,
                         completed_batches=completed_batches,
                         avg_yield=avg_yield,
                         production_by_product=production_by_product)


# ============================================================================
# QUALITY MANAGEMENT (PHASE 7)
# ============================================================================

@main_bp.route('/qc/parameters')
@login_required
def qc_parameters():
    product_id = request.args.get('product_id', type=int)
    if product_id:
        parameters = QCParameter.query.filter_by(product_id=product_id, is_active=True).all()
    else:
        parameters = QCParameter.query.filter_by(is_active=True).all()
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('qc/parameters.html', parameters=parameters, products=products, selected_product=product_id)


@main_bp.route('/qc/parameters/add', methods=['GET', 'POST'])
@login_required
def add_qc_parameter():
    if request.method == 'POST':
        try:
            param = QCParameter(
                product_id=int(request.form.get('product_id', 0)),
                parameter_name=request.form.get('parameter_name', ''),
                min_value=float(request.form.get('min_value', '0') or 0),
                max_value=float(request.form.get('max_value', '0') or 0),
                target_value=float(request.form.get('target_value', '0') or 0),
                unit=request.form.get('unit', ''),
                test_method=request.form.get('test_method', ''),
                is_mandatory=request.form.get('is_mandatory') == 'on'
            )
            db.session.add(param)
            db.session.commit()
            flash('QC parameter added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.qc_parameters'))
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('qc/parameter_form.html', products=products, param=None)


@main_bp.route('/qc/results/<int:batch_id>', methods=['GET', 'POST'])
@login_required
def add_qc_results(batch_id):
    batch = ProductionBatch.query.get_or_404(batch_id)
    
    if request.method == 'POST':
        try:
            param_ids = request.form.getlist('parameter_id[]')
            tested_values = request.form.getlist('tested_value[]')
            results = request.form.getlist('result[]')
            remarks_list = request.form.getlist('remarks[]')
            
            for i, pid in enumerate(param_ids):
                qc_result = QCResult(
                    batch_id=batch.id,
                    parameter_id=int(pid),
                    tested_value=float(tested_values[i]) if tested_values[i] else 0,
                    result=results[i],
                    remarks=remarks_list[i]
                )
                db.session.add(qc_result)
            db.session.commit()
            flash('QC results recorded!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.view_production_batch', id=batch.id))
    
    parameters = QCParameter.query.filter_by(is_active=True).all()
    existing_results = {r.parameter_id: r for r in QCResult.query.filter_by(batch_id=batch_id).all()}
    return render_template('qc/results_form.html', batch=batch, parameters=parameters, existing_results=existing_results)


@main_bp.route('/qc/complaints')
@login_required
def customer_complaints():
    status = request.args.get('status', 'all')
    query = CustomerComplaint.query
    if status != 'all':
        query = query.filter_by(status=status)
    complaints = query.order_by(CustomerComplaint.complaint_date.desc()).all()
    return render_template('qc/complaints.html', complaints=complaints, current_status=status)


@main_bp.route('/qc/complaints/add', methods=['GET', 'POST'])
@login_required
def add_customer_complaint():
    if request.method == 'POST':
        try:
            complaint_number = f"COMP-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            
            complaint = CustomerComplaint(
                complaint_number=complaint_number,
                customer_id=int(request.form.get('customer_id', 0)) if request.form.get('customer_id') else None,
                description=request.form.get('description', '')
            )
            db.session.add(complaint)
            db.session.commit()
            flash('Complaint recorded', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.customer_complaints'))
    
    customers = Customer.query.filter_by(is_active=True).all()
    orders = Order.query.order_by(Order.created_at.desc()).limit(50).all()
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('qc/complaint_form.html', customers=customers, orders=orders, products=products, complaint=None)


@main_bp.route('/qc/complaints/<int:id>/resolve', methods=['POST'])
@login_required
def resolve_complaint(id):
    complaint = CustomerComplaint.query.get_or_404(id)
    
    try:
        complaint.status = 'resolved'
        complaint.resolution = request.form.get('resolution', '')
        complaint.resolved_by = current_user.username
        complaint.resolved_date = datetime.utcnow()
        db.session.commit()
        flash('Complaint resolved successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.customer_complaints'))


@main_bp.route('/qc/capa')
@login_required
def capa_records():
    status = request.args.get('status', 'all')
    query = CAPA.query
    if status != 'all':
        query = query.filter_by(status=status)
    capa_list = query.order_by(CAPA.due_date.asc()).all()
    return render_template('qc/capa.html', capa_list=capa_list, current_status=status)


@main_bp.route('/qc/calibration')
@login_required
def calibration_records():
    records = CalibrationRecord.query.filter_by(is_active=True).order_by(CalibrationRecord.next_due_date.asc()).all()
    due_soon = [r for r in records if r.next_due_date <= datetime.utcnow().date() + timedelta(days=30)]
    overdue = [r for r in records if r.next_due_date < datetime.utcnow().date()]
    return render_template('qc/calibration.html', records=records, due_soon=due_soon, overdue=overdue)


@main_bp.route('/qc/coa/<int:batch_id>')
@login_required
def generate_coa(batch_id):
    batch = ProductionBatch.query.get_or_404(batch_id)
    qc_results = QCResult.query.filter_by(batch_id=batch.id).all()
    return render_template('qc/coa.html', batch=batch, qc_results=qc_results)


# ============================================================================
# REPORTS (PHASE 4)
# ============================================================================

@main_bp.route('/reports')
@login_required
def reports_dashboard():
    total_reports = Report.query.count()
    recent_reports = Report.query.order_by(Report.generated_at.desc()).limit(10).all()
    return render_template('reports/dashboard.html',
                         total_reports=total_reports,
                         recent_reports=recent_reports)


@main_bp.route('/reports/generate/<module>/excel', methods=['POST'])
@login_required
def generate_excel_report(module):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        date_from = request.form.get('date_from', '')
        if not date_from:
            date_from = '2000-01-01'
        date_to = request.form.get('date_to', '')
        if not date_to:
            date_to = '2099-12-31'
        
        wb = Workbook()
        ws = wb.active
        if not ws:
            flash('Error creating workbook', 'error')
            return redirect(url_for(f'main.{module}'))
        
        ws.title = f"{module.title()} Report"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:F1')
        ws['A1'] = f"AmpouleX - {module.title()} Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} by {current_user.username}"
        ws['A3'] = f"Period: {date_from} to {date_to}"
        
        row = 5
        headers = []
        data = []
        
        if module == 'inquiries':
            headers = ['Inquiry #', 'Customer', 'Business', 'Product', 'Quantity', 'Status', 'Date']
            query = Inquiry.query.filter(Inquiry.created_at >= date_from, Inquiry.created_at <= date_to)
            for inq in query.all():
                products_str = ", ".join([f"{item.product.name} ({item.quantity})" for item in inq.inquiry_items if item.product])
                data.append([
                    inq.inquiry_number,
                    inq.customer_name,
                    inq.business_name or '-',
                    products_str,
                    inq.quantity,
                    inq.status,
                    inq.created_at.strftime('%Y-%m-%d')
                ])
        elif module == 'orders':
            headers = ['Order #', 'Customer', 'Total (PKR)', 'Status', 'Payment', 'Date']
            query = Order.query.filter(
                Order.created_at >= date_from,
                Order.created_at <= date_to
            )
            for order in query.all():
                data.append([
                    order.order_number,
                    order.customer_name_snapshot or (order.customer.name if order.customer else 'N/A'),
                    f"{order.total_amount:.2f}",
                    order.status,
                    order.payment_status,
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif module == 'products':
            headers = ['ID', 'Name', 'Specification', 'Type', 'Color', 'Price (PKR)', 'Stock']
            for product in Product.query.filter_by(is_deleted=False).all():
                data.append([
                    product.id,
                    product.name,
                    product.specification or '-',
                    product.product_type,
                    product.color,
                    f"{product.base_price:.2f}",
                    product.stock
                ])
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row += 1
        
        # Write data
        for record in data:
            for col, value in enumerate(record, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
        
        # Auto-adjust column widths
        if ws.columns:
            for column in ws.columns:
                if column and len(column) > 0:
                    first_cell = column[0]
                    try:
                        if hasattr(first_cell, 'column_letter') and first_cell.column_letter:
                            column_letter = first_cell.column_letter
                        else:
                            from openpyxl.utils import get_column_letter
                            column_letter = get_column_letter(first_cell.column)
                    except:
                        continue
                    
                    max_length = 0
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        filename = f"{module}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        filepath = os.path.join(reports_dir, filename)
        wb.save(filepath)
        
        report = Report(
            report_type='excel',
            module=module,
            generated_by=current_user.username,
            parameters=json.dumps({'date_from': date_from, 'date_to': date_to}),
            file_path=filepath
        )
        db.session.add(report)
        db.session.commit()
        
        flash(f'Excel report generated: {filename}', 'success')
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')
        return redirect(url_for(f'main.{module}'))


@main_bp.route('/reports/generate/<module>/pdf', methods=['POST'])
@login_required
def generate_pdf_report(module):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        
        date_from = request.form.get('date_from', '')
        if not date_from:
            date_from = '2000-01-01'
        date_to = request.form.get('date_to', '')
        if not date_to:
            date_to = '2099-12-31'
        
        filename = f"{module}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
        reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        filepath = os.path.join(reports_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12
        )
        
        elements.append(Paragraph(f"AmpouleX - {module.title()} Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} by {current_user.username}", styles['Normal']))
        elements.append(Paragraph(f"Period: {date_from} to {date_to}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        data = []
        headers = []
        
        if module == 'inquiries':
            headers = ['Inquiry #', 'Customer', 'Business', 'Product', 'Qty', 'Status', 'Date']
            query = Inquiry.query.filter(Inquiry.created_at >= date_from, Inquiry.created_at <= date_to)
            for inq in query.all():
                products_str = ", ".join([f"{item.product.name} ({item.quantity})" for item in inq.inquiry_items if item.product])
                data.append([
                    inq.inquiry_number,
                    inq.customer_name,
                    inq.business_name or '-',
                    products_str,
                    inq.quantity,
                    inq.status,
                    inq.created_at.strftime('%Y-%m-%d')
                ])
        elif module == 'orders':
            headers = ['Order #', 'Customer', 'Total (PKR)', 'Status', 'Payment', 'Date']
            query = Order.query.filter(
                Order.created_at >= date_from,
                Order.created_at <= date_to
            )
            for order in query.all():
                data.append([
                    order.order_number,
                    order.customer_name_snapshot or (order.customer.name if order.customer else 'N/A'),
                    f"PKR {order.total_amount:.2f}",
                    order.status,
                    order.payment_status,
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        if headers and data:
            table_data = [headers] + data
            table = Table(table_data, colWidths=[1.2*inch] * len(headers))
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            elements.append(table)
        
        doc.build(elements)
        
        report = Report(
            report_type='pdf',
            module=module,
            generated_by=current_user.username,
            parameters=json.dumps({'date_from': date_from, 'date_to': date_to}),
            file_path=filepath
        )
        db.session.add(report)
        db.session.commit()
        
        flash(f'PDF report generated: {filename}', 'success')
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF report: {str(e)}', 'error')
        return redirect(url_for(f'main.{module}'))


@main_bp.route('/reports/sales-analysis')
@login_required
def sales_analysis_report():
    date_from = request.args.get('date_from', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.utcnow().strftime('%Y-%m-%d'))
    group_by = request.args.get('group_by', 'product')
    
    orders = Order.query.filter(
        Order.created_at >= date_from,
        Order.created_at <= date_to,
        Order.status.in_(['completed', 'processing'])
    ).all()
    
    total_revenue = sum(o.total_amount for o in orders)
    total_orders = len(orders)
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0

    if group_by == 'product':
        product_sales = {}
        for order in orders:
            for item in order.order_items:
                if item.product_id not in product_sales:
                    product_sales[item.product_id] = {
                        'name': item.product.name if item.product else 'Unknown',
                        'quantity': 0,
                        'revenue': 0,
                        'orders': 0
                    }
                product_sales[item.product_id]['quantity'] += item.quantity
                product_sales[item.product_id]['revenue'] += item.subtotal
                product_sales[item.product_id]['orders'] += 1
        report_data = sorted(product_sales.values(), key=lambda x: x['revenue'], reverse=True)
    
    elif group_by == 'customer':
        customer_sales = {}
        for order in orders:
            cust_key = order.customer_name_snapshot or 'Walk-in Customer'
            if cust_key not in customer_sales:
                customer_sales[cust_key] = {
                    'name': cust_key,
                    'orders': 0,
                    'revenue': 0,
                    'avg_order': 0
                }
            customer_sales[cust_key]['orders'] += 1
            customer_sales[cust_key]['revenue'] += order.total_amount
        
        for cust in customer_sales.values():
            cust['avg_order'] = cust['revenue'] / cust['orders'] if cust['orders'] > 0 else 0
        report_data = sorted(customer_sales.values(), key=lambda x: x['revenue'], reverse=True)
    
    elif group_by == 'date':
        from collections import defaultdict
        daily_sales = defaultdict(lambda: {'orders': 0, 'revenue': 0})
        for order in orders:
            date_key = order.created_at.strftime('%Y-%m-%d')
            daily_sales[date_key]['orders'] += 1
            daily_sales[date_key]['revenue'] += order.total_amount
        report_data = [{'date': k, 'orders': v['orders'], 'revenue': v['revenue']} for k, v in daily_sales.items()]
        report_data.sort(key=lambda x: x['date'])
        
    return render_template('reports/sales_analysis.html', 
                         report_data=report_data, 
                         total_revenue=total_revenue,
                         total_orders=total_orders,
                         avg_order_value=avg_order_value,
                         date_from=date_from,
                         date_to=date_to,
                         group_by=group_by)


@main_bp.route('/reports/inventory-valuation')
@login_required
def inventory_valuation_report():
    products = Product.query.filter_by(is_deleted=False).all()
    
    total_items = len(products)
    total_stock_value = sum(p.stock * p.base_price for p in products)
    low_stock_items = [p for p in products if p.stock < 10000]
    slow_moving = [p for p in products if p.stock > 0] # Placeholder logic
    
    return render_template('reports/inventory_valuation.html', 
                         products=products, 
                         total_items=total_items,
                         total_stock_value=total_stock_value,
                         low_stock_items=low_stock_items,
                         slow_moving=slow_moving)


@main_bp.route('/reports/customer-purchase-history/<int:customer_id>')
@login_required
def customer_purchase_history(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    orders = Order.query.filter_by(customer_id=customer_id).order_by(Order.created_at.desc()).all()
    
    total_orders = len(orders)
    total_spent = sum(o.total_amount for o in orders)
    total_paid = sum(o.paid_amount for o in orders)
    avg_order_value = total_spent / total_orders if total_orders > 0 else 0
    
    product_purchases = {}
    for order in orders:
        for item in order.order_items:
            if item.product_id not in product_purchases:
                product_purchases[item.product_id] = {
                    'name': item.product.name if item.product else 'Unknown',
                    'quantity': 0,
                    'revenue': 0
                }
            product_purchases[item.product_id]['quantity'] += item.quantity
            product_purchases[item.product_id]['revenue'] += item.subtotal
    
    top_products = sorted(product_purchases.values(), key=lambda x: x['revenue'], reverse=True)[:5]
    
    return render_template('reports/customer_purchase_history.html',
                         customer=customer,
                         orders=orders,
                         total_orders=total_orders,
                         total_spent=total_spent,
                         total_paid=total_paid,
                         avg_order_value=avg_order_value,
                         top_products=top_products)


@main_bp.route('/reports/production-efficiency')
@login_required
def production_efficiency_report():
    date_from = request.args.get('date_from', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.utcnow().strftime('%Y-%m-%d'))
    
    try:
        batches = ProductionBatch.query.filter(
            ProductionBatch.start_date >= date_from,
            ProductionBatch.end_date <= date_to,
            ProductionBatch.status == 'completed'
        ).all()
    except:
        batches = []
    
    total_batches = len(batches)
    avg_yield = sum(b.yield_percentage for b in batches) / total_batches if total_batches > 0 else 0
    return render_template('reports/production_efficiency.html', 
                         batches=batches, 
                         avg_yield=avg_yield,
                         date_from=date_from,
                         date_to=date_to)


@main_bp.route('/reports/material-consumption')
@login_required
def material_consumption_report():
    start_date = request.args.get('start_date', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    consumption = db.session.query(
        RawMaterial.name,
        RawMaterial.unit,
        db.func.sum(MaterialUsage.quantity_used).label('total_used'),
        db.func.count(MaterialUsage.id).label('usage_count')
    ).join(MaterialUsage).filter(
        MaterialUsage.usage_date >= start_date,
        MaterialUsage.usage_date <= end_date
    ).group_by(RawMaterial.id).all()
    
    return render_template('reports/material_consumption.html',
                         consumption=consumption,
                         start_date=start_date,
                         end_date=end_date)


# ============================================================================
# ANALYTICS DASHBOARD
# ============================================================================

@main_bp.route('/analytics/dashboard')
@login_required
def analytics_dashboard():
    today = datetime.utcnow().date()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)
    last_90_days = today - timedelta(days=90)
    
    revenue_today = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        db.func.date(Order.created_at) == today
    ).scalar() or 0
    
    revenue_7_days = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= last_7_days
    ).scalar() or 0
    
    revenue_30_days = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= last_30_days
    ).scalar() or 0
    
    orders_today = Order.query.filter(
        db.func.date(Order.created_at) == today
    ).count()
    
    orders_7_days = Order.query.filter(
        Order.created_at >= last_7_days
    ).count()
    
    orders_30_days = Order.query.filter(
        Order.created_at >= last_30_days
    ).count()
    
    revenue_90_days = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= last_90_days
    ).scalar() or 0

    orders_90_days = Order.query.filter(Order.created_at >= last_90_days).count()
    avg_order_value = revenue_30_days / orders_30_days if orders_30_days > 0 else 0

    # Previous 30-day period for growth rate
    prev_30_start = today - timedelta(days=60)
    prev_30_end = today - timedelta(days=30)
    revenue_prev_30 = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing']),
        Order.created_at >= prev_30_start,
        Order.created_at < prev_30_end
    ).scalar() or 0
    growth_rate = ((revenue_30_days - revenue_prev_30) / revenue_prev_30 * 100) if revenue_prev_30 > 0 else 0

    # Customer metrics
    total_customers = Customer.query.count()
    new_customers_30_days = Customer.query.filter(Customer.created_at >= last_30_days).count() if hasattr(Customer, 'created_at') else 0
    customer_retention_rate = ((total_customers - new_customers_30_days) / total_customers * 100) if total_customers > 0 else 0

    # Total revenue (all time)
    total_revenue = db.session.query(db.func.sum(Order.total_amount)).filter(
        Order.status.in_(['completed', 'processing'])
    ).scalar() or 0

    # Inventory turnover (simple approximation)
    inventory_turnover = (revenue_90_days / (total_revenue / 4)) if total_revenue > 0 else 0

    # Top products by revenue
    top_products_raw = db.session.query(
        Product.name,
        Product.color if hasattr(Product, 'color') else db.literal('N/A'),
        db.func.sum(OrderItem.quantity).label('total_sold'),
        db.func.sum(OrderItem.subtotal).label('total_revenue')
    ).join(OrderItem, OrderItem.product_id == Product.id)\
     .group_by(Product.id, Product.name)\
     .order_by(db.desc('total_revenue'))\
     .limit(10).all()

    class _ProductRow:
        def __init__(self, name, color, total_sold, total_revenue):
            self.name = name
            self.color = color or 'N/A'
            self.total_sold = total_sold or 0
            self.total_revenue = float(total_revenue or 0)

    top_products = [_ProductRow(*r) for r in top_products_raw]

    # Monthly revenue (last 6 months) — list of (label, value) tuples
    monthly_revenue = []
    for i in range(5, -1, -1):
        m_start = today.replace(day=1) - timedelta(days=i * 30)
        m_end = m_start + timedelta(days=30)
        m_rev = db.session.query(db.func.sum(Order.total_amount)).filter(
            Order.status.in_(['completed', 'processing']),
            Order.created_at >= m_start,
            Order.created_at < m_end
        ).scalar() or 0
        monthly_revenue.append((m_start.strftime('%b %Y'), float(m_rev)))

    # Order status distribution
    order_status_dist = db.session.query(
        Order.status, db.func.count(Order.id)
    ).group_by(Order.status).all()

    # Payment status distribution
    payment_status_dist = db.session.query(
        Order.payment_status, db.func.count(Order.id)
    ).group_by(Order.payment_status).all() if hasattr(Order, 'payment_status') else []

    return render_template('analytics/dashboard.html',
                         revenue_today=revenue_today,
                         revenue_7_days=revenue_7_days,
                         revenue_30_days=revenue_30_days,
                         revenue_90_days=revenue_90_days,
                         orders_today=orders_today,
                         orders_7_days=orders_7_days,
                         orders_30_days=orders_30_days,
                         avg_order_value=avg_order_value,
                         growth_rate=growth_rate,
                         total_customers=total_customers,
                         new_customers_30_days=new_customers_30_days,
                         customer_retention_rate=customer_retention_rate,
                         total_revenue=total_revenue,
                         inventory_turnover=inventory_turnover,
                         top_products=top_products,
                         monthly_revenue=monthly_revenue,
                         order_status_dist=order_status_dist,
                         payment_status_dist=payment_status_dist)


# ============================================================================
# BARCODE/QR CODE GENERATION
# ============================================================================

@main_bp.route('/products/<int:id>/barcode')
@login_required
def generate_product_barcode(id):
    try:
        import barcode
        from barcode.writer import ImageWriter
        
        product = Product.query.get_or_404(id)
        barcode_number = f"890{product.id:010d}"
        barcode_class = barcode.get('ean13', barcode_number, writer=ImageWriter())
        
        barcode_dir = os.path.join('static', 'barcodes')
        os.makedirs(barcode_dir, exist_ok=True)
        filename = f"product_{product.id}_barcode"
        filepath = os.path.join(barcode_dir, filename)
        saved_path = barcode_class.save(filepath)
        
        return send_file(f"{saved_path}.png", mimetype='image/png')
    except ImportError:
        flash('Barcode library not installed. Run: pip install python-barcode', 'error')
        return redirect(url_for('main.products'))
    except Exception as e:
        flash(f'Error generating barcode: {str(e)}', 'error')
        return redirect(url_for('main.products'))


@main_bp.route('/products/<int:id>/qrcode')
@login_required
def generate_product_qrcode(id):
    try:
        import qrcode
        
        product = Product.query.get_or_404(id)
        qr_data = f"Product: {product.name}\nID: {product.id}\nPrice: PKR {product.base_price}\nStock: {product.stock}"
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        qrcode_dir = os.path.join('static', 'qrcodes')
        os.makedirs(qrcode_dir, exist_ok=True)
        filename = f"product_{product.id}_qrcode.png"
        filepath = os.path.join(qrcode_dir, filename)
        
        try:
            img.save(filepath)
        except Exception as e:
            with open(filepath, 'wb') as f:
                img.save(f, format='PNG')
        
        return send_file(filepath, mimetype='image/png')
    except ImportError:
        flash('QR code library not installed. Run: pip install qrcode[pil]', 'error')
        return redirect(url_for('main.products'))
    except Exception as e:
        flash(f'Error generating QR code: {str(e)}', 'error')
        return redirect(url_for('main.products'))


@main_bp.route('/products/<int:id>/labels')
@login_required
def generate_product_labels(id):
    product = Product.query.get_or_404(id)
    return render_template('products/labels.html', product=product)


# ============================================================================
# SETTINGS
# ============================================================================

@main_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        try:
            settings_data = {
                'company_name': request.form.get('company_name', ''),
                'company_ntn': request.form.get('company_ntn', ''),
                'company_gst': request.form.get('company_gst', ''),
                'company_secp': request.form.get('company_secp', ''),
                'company_drap': request.form.get('company_drap', ''),
                'company_iso': request.form.get('company_iso', ''),
                'company_address': request.form.get('company_address', ''),
                'company_phone': request.form.get('company_phone', ''),
                'company_email': request.form.get('company_email', ''),
                'company_website': request.form.get('company_website', ''),
                'invoice_footer': request.form.get('invoice_footer', '')
            }
            
            for key, value in settings_data.items():
                setting = CompanySetting.query.filter_by(key=key).first()
                if setting:
                    setting.value = value
                else:
                    setting = CompanySetting(key=key, value=value)
                    db.session.add(setting)
            
            db.session.commit()
            flash('Settings updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.settings'))
    
    settings = {s.key: s.value for s in CompanySetting.query.all()}
    return render_template('settings/index.html', settings=settings)


@main_bp.route('/settings/upload-logo', methods=['POST'])
@login_required
def upload_logo():
    if 'logo' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('main.settings'))
    
    file = request.files['logo']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('main.settings'))
    
    if file:
        filename = "company_logo.png"
        filepath = os.path.join('static', 'logos', filename)
        file.save(filepath)
        
        setting = CompanySetting.query.filter_by(key='company_logo').first()
        if setting:
            setting.value = filepath
        else:
            setting = CompanySetting(key='company_logo', value=filepath)
            db.session.add(setting)
        
        db.session.commit()
        flash('Logo uploaded successfully!', 'success')
        return redirect(url_for('main.settings'))


# ============================================================================
# CUSTOMER-FACING WEBSITE
# ============================================================================

@main_bp.route('/customer-site')
def customer_site():
    grouped_products = group_products_by_base()
    return render_template('customer-site.html', products=grouped_products)


@main_bp.route('/submit-inquiry', methods=['GET', 'POST'])
def submit_inquiry():
    try:
        inquiry_number = f"INQ-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        product_ids = request.form.getlist('product_ids')
        customer_name = request.form.get('customer_name', '')
        business_name = request.form.get('business_name', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        
        customer, is_new = find_or_create_customer(
            customer_name=customer_name,
            business_name=business_name,
            email=email,
            phone=phone
        )
        
        inquiry = Inquiry(
            inquiry_number=inquiry_number,
            customer_name=customer_name,
            business_name=business_name,
            email=email,
            phone=phone,
            product_id=None,
            quantity=0,
            notes=request.form.get('notes', ''),
            status='new'
        )
        db.session.add(inquiry)
        db.session.flush()
        
        total_quantity = 0
        products_list = []
        
        # ✅ CREATE InquiryItem RECORDS
        for pid in product_ids:
            qty = request.form.get(f'qty_{pid}', '1')
            qty = int(qty) if qty else 1
            
            if qty > 0:
                # Create InquiryItem
                item = InquiryItem(
                    inquiry_id=inquiry.id,
                    product_id=int(pid),
                    quantity=qty
                )
                db.session.add(item)
                
                total_quantity += qty
                product = Product.query.get(int(pid))
                if product:
                    products_list.append(f"{product.name} ({product.color})")
        
        inquiry.quantity = total_quantity
        db.session.commit()  # ✅ Commit AFTER creating items
        
        try:
            socketio.emit('new_inquiry', {
                'id': inquiry.id,
                'number': inquiry.inquiry_number,
                'customer': customer_name,
                'business': business_name,
                'products': ', '.join(products_list),
                'quantity': total_quantity,
                'status': inquiry.status,
                'created_at': inquiry.created_at.isoformat(),
                'customer_id': customer.id,
                'is_new_customer': is_new
            }, namespace='/admin')
        except Exception as e:
            print(f"Socket.IO error: {e}")
        
        flash(f'Thank you! Your inquiry #{inquiry_number} has been submitted.', 'success')
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {str(e)}")
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.customer_site'))

# ============================================================================
# API ENDPOINTS
# ============================================================================

@main_bp.route('/api/check-accounting-period')
@login_required
def check_accounting_period():
    """Check if accounting period is open for given date"""
    date_str = request.args.get('date', '')
    
    if not date_str:
        return jsonify({'error': 'Date required'}), 400
    
    try:
        entry_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        period = AccountingPeriod.query.filter(
            AccountingPeriod.start_date <= entry_date,
            AccountingPeriod.end_date >= entry_date
        ).first()
        
        if period:
            return jsonify({
                'is_open': not period.is_closed,
                'period_name': period.period_name,
                'start_date': period.start_date.strftime('%Y-%m-%d'),
                'end_date': period.end_date.strftime('%Y-%m-%d'),
                'is_closed': period.is_closed
            })
        else:
            return jsonify({
                'is_open': True,
                'period_name': 'No period defined',
                'warning': 'No accounting period defined for this date'
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main_bp.route('/api/inquiry-count')
@login_required
def inquiry_count():
    count = Inquiry.query.filter(
        Inquiry.status.in_(['new', 'followup']),
        Inquiry.is_deleted == False
    ).count()
    return jsonify({'count': count})


@main_bp.route('/api/revenue-data')
@login_required
def revenue_data():
    six_months_ago = datetime.utcnow() - timedelta(days=180)
    monthly_revenue = db.session.query(
        db.func.to_char(Order.created_at, 'YYYY-MM'),
        db.func.sum(Order.total_amount)
    ).filter(
        Order.created_at >= six_months_ago,
        Order.status.in_(['completed', 'processing'])
    ).group_by(db.func.to_char(Order.created_at, 'YYYY-MM')).all()
    
    return jsonify({
        'labels': [r[0] for r in monthly_revenue],
        'data': [float(r[1]) for r in monthly_revenue]
    })


@main_bp.route('/api/low-stock-count')
@login_required
def low_stock_count():
    count = Product.query.filter(
        Product.stock < 10000,
        Product.is_deleted == False
    ).count()
    return jsonify({'count': count})


@main_bp.route('/api/overdue-orders-count')
@login_required
def overdue_orders_count():
    count = Order.query.filter(
        Order.payment_status == 'unpaid',
        Order.created_at < datetime.utcnow() - timedelta(days=30)
    ).count()
    return jsonify({'count': count})


@main_bp.route('/api/pending-orders-count')
@login_required
def pending_orders_count():
    count = Order.query.filter_by(status='pending').count()
    return jsonify({'count': count})


@main_bp.route('/api/order-status-data')
@login_required
def order_status_data():
    status_counts = db.session.query(
        Order.status,
        db.func.count(Order.id)
    ).group_by(Order.status).all()
    
    labels = [s[0] or 'Unknown' for s in status_counts]
    data = [s[1] for s in status_counts]
    
    return jsonify({
        'labels': labels,
        'data': data
    })


# ============================================================================
# PAYROLL/EMPLOYEES
# ============================================================================

@main_bp.route('/payroll')
@login_required
def payroll():
    employees = Employee.query.all()
    return render_template('payroll/index.html', employees=employees)


@main_bp.route('/employees/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if request.method == 'POST':
        try:
            annual_salary = float(request.form.get('base_salary', 0)) * 12
            monthly_tax, _ = calculate_monthly_tax_deduction(annual_salary)
            
            employee = Employee(
                name=request.form.get('name', ''),
                role=request.form.get('role', ''),
                department=request.form.get('department', ''),
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''),
                cnic=request.form.get('cnic', ''),
                address=request.form.get('address', ''),
                emergency_contact=request.form.get('emergency_contact', ''),
                emergency_phone=request.form.get('emergency_phone', ''),
                base_salary=float(request.form.get('base_salary', 0)),
                payment_frequency=request.form.get('payment_frequency', 'monthly'),
                bank_account=request.form.get('bank_account', ''),
                bank_name=request.form.get('bank_name', ''),
                outstanding_loan=float(request.form.get('outstanding_loan', 0))
            )
            db.session.add(employee)
            db.session.commit()
            flash('Employee added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.payroll'))
    
    return render_template('payroll/form.html', employee=None)


@main_bp.route('/employees/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_employee(id):
    employee = Employee.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            employee.name = request.form.get('name', '')
            employee.role = request.form.get('role', '')
            employee.department = request.form.get('department', '')
            employee.email = request.form.get('email', '')
            employee.phone = request.form.get('phone', '')
            employee.cnic = request.form.get('cnic', '')
            employee.address = request.form.get('address', '')
            employee.emergency_contact = request.form.get('emergency_contact', '')
            employee.emergency_phone = request.form.get('emergency_phone', '')
            employee.base_salary = float(request.form.get('base_salary', 0))
            employee.payment_frequency = request.form.get('payment_frequency', 'monthly')
            employee.bank_account = request.form.get('bank_account', '')
            employee.bank_name = request.form.get('bank_name', '')
            employee.outstanding_loan = float(request.form.get('outstanding_loan', 0))
            employee.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash('Employee updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.payroll'))
    
    return render_template('payroll/form.html', employee=employee)


@main_bp.route('/employees/delete/<int:id>')
@login_required
def delete_employee(id):
    try:
        employee = Employee.query.get_or_404(id)
        employee.is_active = False
        db.session.commit()
        flash('Employee deactivated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.payroll'))


# ============================================================================
# EXPENSES
# ============================================================================

@main_bp.route('/expenses')
@login_required
def expenses():
    expenses = Expense.query.order_by(Expense.date.desc()).all()
    total_expenses = db.session.query(db.func.sum(Expense.amount)).scalar() or 0
    pending_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.status == 'pending'
    ).scalar() or 0
    paid_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.status == 'paid'
    ).scalar() or 0
    
    return render_template('expenses/index.html',
                         expenses=expenses,
                         total_expenses=total_expenses,
                         pending_expenses=pending_expenses,
                         paid_expenses=paid_expenses)


@main_bp.route('/expenses/add', methods=['GET', 'POST'])
@login_required
def add_expense():
    if request.method == 'POST':
        try:
            date_str = request.form.get('date')
            if date_str:
                expense_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                expense_date = datetime.utcnow().date()
            
            expense = Expense(
                category=request.form.get('category', ''),
                description=request.form.get('description', ''),
                amount=float(request.form.get('amount', 0)),
                date=expense_date,
                status=request.form.get('status', 'pending'),
                paid_by=request.form.get('paid_by', ''),
                payment_method=request.form.get('payment_method', ''),
                reference_number=request.form.get('reference_number', ''),
                approved_by=request.form.get('approved_by', '')
            )
            db.session.add(expense)
            db.session.commit()
            flash('Expense added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.expenses'))
    
    return render_template('expenses/form.html', expense=None)


@main_bp.route('/expenses/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_expense(id):
    expense = Expense.query.get_or_404(id)
    if request.method == 'POST':
        try:
            old_status = expense.status
            expense.category = request.form.get('category', '')
            expense.description = request.form.get('description', '')
            expense.amount = float(request.form.get('amount', 0))
            date_str = request.form.get('date')
            if date_str:
                expense.date = datetime.strptime(date_str, '%Y-%m-%d').date()
            expense.status = request.form.get('status', 'pending')
            expense.paid_by = request.form.get('paid_by', '')
            expense.payment_method = request.form.get('payment_method', '')
            expense.reference_number = request.form.get('reference_number', '')
            expense.approved_by = request.form.get('approved_by', '')
            db.session.commit()
            
            # ✅ AUTO-CREATE JOURNAL ENTRY WHEN EXPENSE IS PAID
            if old_status != 'paid' and expense.status == 'paid':
                create_journal_entry_for_expense(expense)
            
            flash('Expense updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('main.expenses'))
    return render_template('expenses/form.html', expense=expense)

@main_bp.route('/expenses/delete/<int:id>')
@login_required
def delete_expense(id):
    try:
        expense = Expense.query.get_or_404(id)
        db.session.delete(expense)
        db.session.commit()
        flash('Expense deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.expenses'))


# ============================================================================
# ACCOUNTING
# ============================================================================

@main_bp.route('/accounting')
@login_required
def accounting():
    show_inactive = request.args.get('show_inactive', 'false') == 'true'
    

    petty_cash_accounts = Account.query.filter_by(account_type='Asset', account_name='Petty Cash').all()
    petty_cash_account_ids = [a.id for a in petty_cash_accounts]
    bank = Accounting.query.filter_by(account_type='bank').all()
    receivables = Accounting.query.filter_by(account_type='receivable').all()
    payables = Accounting.query.filter_by(account_type='payable').all()

    # ✅ Calculate petty cash balance
    petty_cash_balance = sum(
        a.amount if a.transaction_type == 'credit' else -a.amount 
        for a in Accounting.query.filter_by(account_type='petty_cash').all()
    )

    # ✅ Calculate bank balance
    bank_balance = sum(
        a.amount if a.transaction_type == 'credit' else -a.amount 
        for a in bank
    )
    
    # ✅ Calculate receivables balance (what customers owe us)
    receivables_balance = sum(
        a.amount if a.transaction_type == 'debit' else -a.amount 
        for a in receivables
    )
    
    # ✅ Calculate payables balance (what we owe suppliers)
    payables_balance = sum(
        a.amount if a.transaction_type == 'credit' else -a.amount 
        for a in payables
    )
    
    employee_loans = Employee.query.filter(Employee.outstanding_loan > 0).all()
    
    return render_template('accounting/index.html',
                         petty_cash_balance=petty_cash_balance,
                         bank_balance=bank_balance,  # ✅ Added
                         receivables_balance=receivables_balance,  # ✅ Added
                         payables_balance=payables_balance,  # ✅ Added
                         bank=bank,
                         receivables=receivables,
                         payables=payables,
                         employee_loans=employee_loans)


# ============================================================================
# FINANCIAL STATEMENTS
# ============================================================================
@main_bp.route('/financials/profit-loss')
@login_required
def profit_loss():
    start_date = request.args.get('start_date', datetime.utcnow().replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    # ✅ USE JOURNAL ENTRIES FOR REVENUE
    income_account_ids = [a.id for a in Account.query.filter_by(account_type='Income').all()]
    total_revenue = db.session.query(
        db.func.sum(JournalEntryLine.credit)
    ).join(JournalEntry).filter(
        JournalEntryLine.account_id.in_(income_account_ids),
        JournalEntry.status == 'posted',
        JournalEntry.entry_date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
        JournalEntry.entry_date <= datetime.strptime(end_date, '%Y-%m-%d').date()
    ).scalar() or 0
    
    # ✅ USE JOURNAL ENTRIES FOR COGS
    cogs_account_ids = [a.id for a in Account.query.filter(
        Account.account_type == 'Expense',
        Account.account_name.like('%Cost of Goods%')
    ).all()]
    cogs = db.session.query(
        db.func.sum(JournalEntryLine.debit)
    ).join(JournalEntry).filter(
        JournalEntryLine.account_id.in_(cogs_account_ids),
        JournalEntry.status == 'posted',
        JournalEntry.entry_date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
        JournalEntry.entry_date <= datetime.strptime(end_date, '%Y-%m-%d').date()
    ).scalar() or 0
    
    gross_profit = total_revenue - cogs
    
    # ✅ USE JOURNAL ENTRIES FOR EXPENSES
    expense_account_ids = [a.id for a in Account.query.filter_by(account_type='Expense').all()]
    total_expenses = db.session.query(
        db.func.sum(JournalEntryLine.debit)
    ).join(JournalEntry).filter(
        JournalEntryLine.account_id.in_(expense_account_ids),
        JournalEntry.status == 'posted',
        JournalEntry.entry_date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
        JournalEntry.entry_date <= datetime.strptime(end_date, '%Y-%m-%d').date()
    ).scalar() or 0
    
    operating_profit = gross_profit - total_expenses
    
    # Payroll (keep existing logic for now)
    total_payroll = db.session.query(db.func.sum(PayrollPayment.net_pay)).filter(
        PayrollPayment.payment_date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
        PayrollPayment.payment_date <= datetime.strptime(end_date, '%Y-%m-%d').date()
    ).scalar() or 0
    
    net_profit = operating_profit - total_payroll
    
    # Expenses by category (from JournalEntry)
    expenses_by_category = db.session.query(
        Account.account_name,
        db.func.sum(JournalEntryLine.debit)
    ).join(JournalEntryLine).join(JournalEntry).filter(
        JournalEntryLine.account_id.in_(expense_account_ids),
        JournalEntry.status == 'posted',
        JournalEntry.entry_date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
        JournalEntry.entry_date <= datetime.strptime(end_date, '%Y-%m-%d').date()
    ).group_by(Account.account_name).all()
    
    # Six months trend
    six_months = []
    for i in range(6):
        month_end = datetime.utcnow() - timedelta(days=30*i)
        month_start = month_end - timedelta(days=30)
        
        month_revenue = db.session.query(
            db.func.sum(JournalEntryLine.credit)
        ).join(JournalEntry).filter(
            JournalEntryLine.account_id.in_(income_account_ids),
            JournalEntry.status == 'posted',
            JournalEntry.entry_date >= month_start.date(),
            JournalEntry.entry_date <= month_end.date()
        ).scalar() or 0
        
        six_months.append({
            'month': month_end.strftime('%b %Y'),
            'revenue': month_revenue
        })
    
    return render_template('financials/profit_loss.html',
                         start_date=start_date,
                         end_date=end_date,
                         total_revenue=total_revenue,
                         cogs=cogs,
                         gross_profit=gross_profit,
                         total_expenses=total_expenses,
                         operating_profit=operating_profit,
                         total_payroll=total_payroll,
                         net_profit=net_profit,
                         expenses_by_category=expenses_by_category,
                         six_months=six_months)

@main_bp.route('/financials/balance-sheet')
@login_required
def balance_sheet():
    # ✅ USE JOURNAL ENTRIES FOR CASH BALANCES
    cash_account_ids = [a.id for a in Account.query.filter(
        Account.account_type == 'Asset',
        (Account.account_name.like('%Cash%')) | (Account.account_name.like('%Bank%'))
    ).all()]
    
    total_cash = 0
    bank_balance = 0  # ✅ ADD THIS
    
    for account_id in cash_account_ids:
        debits = db.session.query(db.func.sum(JournalEntryLine.debit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account_id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        credits = db.session.query(db.func.sum(JournalEntryLine.credit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account_id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        
        account = Account.query.get(account_id)
        if account and 'Bank' in account.account_name:
            bank_balance += (debits - credits)
        else:
            total_cash += (debits - credits)  # Asset accounts: Debit - Credit
    
    # ✅ Calculate petty cash balance
    petty_cash_balance = sum(
        a.amount if a.transaction_type == 'credit' else -a.amount 
        for a in Accounting.query.filter_by(account_type='petty_cash').all()
    )

    # Accounts Receivable (from Customer balances for now)
    accounts_receivable = db.session.query(db.func.sum(Customer.current_balance)).filter(
        Customer.current_balance > 0
    ).scalar() or 0
    
    # Inventory Value (from Product stock)
    inventory_value = sum(p.stock * p.unit_price for p in Product.query.filter(
        Product.stock > 0,
        Product.is_deleted == False
    ).all())
    
    # Raw Material Value (from RawMaterial stock)
    raw_material_value = sum(m.current_stock * (m.cost_per_unit or 0) for m in RawMaterial.query.filter(
        RawMaterial.current_stock > 0
    ).all())

    # Liabilities (from JournalEntry)
    liability_account_ids = [a.id for a in Account.query.filter_by(account_type='Liability').all()]

    # Accounts Payable (from Supplier balances or Liability accounts)
    accounts_payable = db.session.query(db.func.sum(SupplierInvoice.total_amount)).filter(
        SupplierInvoice.status == 'pending'
    ).scalar() or 0

    # Tax Payable (from Order tax_amount for unpaid/partial orders)
    tax_payable = db.session.query(db.func.sum(Order.tax_amount)).filter(
        Order.payment_status.in_(['unpaid', 'partial']),
        Order.status != 'cancelled'
    ).scalar() or 0

    total_current_assets = total_cash + accounts_receivable + inventory_value
    fixed_assets = 0  # Placeholder for fixed assets (machinery, equipment, etc.)
    total_current_liabilities = accounts_payable + tax_payable
    total_assets = total_current_assets

    total_liabilities = 0
    for account_id in liability_account_ids:
        debits = db.session.query(db.func.sum(JournalEntryLine.debit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account_id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        credits = db.session.query(db.func.sum(JournalEntryLine.credit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account_id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        total_liabilities += (credits - debits)  # Liability: Credit - Debit
    
    # Employee loans payable
    employee_loans_payable = db.session.query(db.func.sum(Employee.outstanding_loan)).scalar() or 0
    total_liabilities += employee_loans_payable
    
    # Equity
    equity_account_ids = [a.id for a in Account.query.filter_by(account_type='Equity').all()]
    
    equity = 0
    for account_id in equity_account_ids:
        debits = db.session.query(db.func.sum(JournalEntryLine.debit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account_id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        credits = db.session.query(db.func.sum(JournalEntryLine.credit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account_id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        equity += (credits - debits)  # Equity: Credit - Debit
    
    # Retained earnings (Net profit from all time)
    income_account_ids = [a.id for a in Account.query.filter_by(account_type='Income').all()]
    expense_account_ids = [a.id for a in Account.query.filter_by(account_type='Expense').all()]
    
    total_revenue = db.session.query(db.func.sum(JournalEntryLine.credit)).join(JournalEntry).filter(
        JournalEntryLine.account_id.in_(income_account_ids),
        JournalEntry.status == 'posted'
    ).scalar() or 0
    
    total_expenses = db.session.query(db.func.sum(JournalEntryLine.debit)).join(JournalEntry).filter(
        JournalEntryLine.account_id.in_(expense_account_ids),
        JournalEntry.status == 'posted'
    ).scalar() or 0
    
    retained_earnings = total_revenue - total_expenses
    equity += retained_earnings
    
    long_term_liabilities = max(0, total_liabilities - total_current_liabilities)

    return render_template('financials/balance_sheet.html',
                         bank_balance=bank_balance,
                         petty_cash_balance=petty_cash_balance,
                         total_cash=total_cash,
                         accounts_receivable=accounts_receivable,
                         inventory_value=inventory_value,
                         tax_payable=tax_payable,
                         total_employee_loans=employee_loans_payable,
                         raw_material_value=raw_material_value,
                         fixed_assets=fixed_assets,
                         accounts_payable=accounts_payable,
                         total_current_liabilities=total_current_liabilities,
                         long_term_liabilities=long_term_liabilities,
                         total_current_assets=total_current_assets,
                         total_assets=total_assets,
                         total_liabilities=total_liabilities,
                         equity=equity)


@main_bp.route('/financials/cash-flow')
@login_required
def cash_flow():
    start_date = request.args.get('start_date', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.utcnow().strftime('%Y-%m-%d'))
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')

    opening_bank_balance = db.session.query(db.func.sum(db.case((Accounting.transaction_type == 'credit', Accounting.amount), else_=-Accounting.amount))).filter(Accounting.account_type == 'bank', Accounting.date < start_date_obj).scalar() or 0
    opening_petty_cash = db.session.query(db.func.sum(db.case((Accounting.transaction_type == 'credit', Accounting.amount), else_=-Accounting.amount))).filter(Accounting.account_type == 'petty_cash', Accounting.date < start_date_obj).scalar() or 0
    opening_cash = opening_bank_balance + opening_petty_cash

    cash_from_sales = db.session.query(db.func.sum(Order.paid_amount)).filter(
        Order.payment_status.in_(['paid', 'partial']),
        Order.payment_date >= start_date,
        Order.payment_date <= end_date
    ).scalar() or 0

    cash_for_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.status == 'paid',
        Expense.date >= start_date,
        Expense.date <= end_date
    ).scalar() or 0

    cash_for_payroll = db.session.query(db.func.sum(PayrollPayment.net_pay)).filter(
        PayrollPayment.status == 'paid',
        PayrollPayment.payment_date >= start_date,
        PayrollPayment.payment_date <= end_date
    ).scalar() or 0

    net_cash_from_operations = cash_from_sales - cash_for_expenses - cash_for_payroll
    cash_from_investing = 0
    net_cash_from_investing = -cash_from_investing

    cash_from_loans = 0
    cash_to_loans = db.session.query(db.func.sum(PayrollPayment.loan_deduction)).filter(
        PayrollPayment.payment_date >= start_date,
        PayrollPayment.payment_date <= end_date
    ).scalar() or 0

    net_cash_from_financing = cash_from_loans - cash_to_loans
    net_change_in_cash = net_cash_from_operations + net_cash_from_investing + net_cash_from_financing
    closing_cash = opening_cash + net_change_in_cash

    return render_template('financials/cash_flow.html',
                         start_date=start_date,
                         end_date=end_date,
                         opening_cash=opening_cash,
                         cash_from_sales=cash_from_sales,
                         cash_from_expenses=cash_for_expenses,
                         cash_from_payroll=cash_for_payroll,
                         cash_from_investing=cash_from_investing,
                         cash_from_loans=cash_from_loans,
                         cash_to_loans=cash_to_loans,
                         net_cash_from_operations=net_cash_from_operations,
                         net_cash_from_investing=net_cash_from_investing,
                         net_cash_from_financing=net_cash_from_financing,
                         net_change_in_cash=net_change_in_cash,
                         closing_cash=closing_cash)


# ============================================================================
# ATTENDANCE TRACKING
# ============================================================================

@main_bp.route('/attendance')
@login_required
def attendance():
    date = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
    employees = Employee.query.filter_by(is_active=True).all()
    attendance_records = {a.employee_id: a for a in Attendance.query.filter_by(date=date).all()}
    return render_template('payroll/attendance.html', 
                         employees=employees, 
                         attendance_records=attendance_records,
                         selected_date=date)


@main_bp.route('/attendance/mark', methods=['POST'])
@login_required
def mark_attendance():
    date_str = request.form.get('date', '')
    try:
        employee_id = int(request.form.get('employee_id', '0') or 0)
        date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.utcnow().date()
        status = request.form.get('status', 'absent')
        check_in_str = request.form.get('check_in', '')
        check_out_str = request.form.get('check_out', '')

        check_in = datetime.strptime(check_in_str, '%H:%M') if check_in_str else None
        check_out = datetime.strptime(check_out_str, '%H:%M') if check_out_str else None

        hours_worked = 0
        overtime_hours = 0

        if check_in and check_out:
            check_in_dt = datetime.combine(date, check_in.time())
            check_out_dt = datetime.combine(date, check_out.time())
            if check_out_dt < check_in_dt:
                check_out_dt += timedelta(days=1)

            duration_hours = (check_out_dt - check_in_dt).total_seconds() / 3600
            if duration_hours > 6:
                duration_hours -= 1

            hours_worked = min(duration_hours, 8)
            overtime_hours = max(0, duration_hours - 8)

        attendance = Attendance.query.filter_by(employee_id=employee_id, date=date).first()
        if not attendance:
            attendance = Attendance(employee_id=employee_id, date=date)
            db.session.add(attendance)
        
        attendance.status = status
        attendance.check_in = check_in
        attendance.check_out = check_out_dt if check_out else None
        attendance.hours_worked = float(hours_worked)
        attendance.overtime_hours = overtime_hours
        attendance.notes = request.form.get('notes', '')
        
        db.session.commit()
        notify_attendance_marked(attendance)
        flash('Attendance marked successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.attendance', date=date_str))


@main_bp.route('/attendance/history')
@login_required
def attendance_history():
    employee_id = request.args.get('employee_id', type=int)
    date_from = request.args.get('date_from', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.utcnow().strftime('%Y-%m-%d'))
    
    query = Attendance.query.filter(
        Attendance.date >= date_from,
        Attendance.date <= date_to
    )
    if employee_id:
        query = query.filter_by(employee_id=employee_id)
    
    attendance_records = query.order_by(Attendance.date.desc()).all()
    employees = Employee.query.filter_by(is_active=True).all()
    
    return render_template('payroll/attendance_history.html',
                         attendance_records=attendance_records,
                         employees=employees,
                         selected_employee=employee_id,
                         date_from=date_from,
                         date_to=date_to)


# ============================================================================
# TIMESHEET MANAGEMENT
# ============================================================================

@main_bp.route('/timesheets')
@login_required
def timesheets():
    status = request.args.get('status', 'all')
    query = Timesheet.query
    if status != 'all':
        query = query.filter_by(status=status)
    timesheets_list = query.order_by(Timesheet.created_at.desc()).all()
    return render_template('payroll/timesheets.html', timesheets=timesheets_list, current_status=status)


@main_bp.route('/timesheets/generate', methods=['GET', 'POST'])
@login_required
def generate_timesheet():
    if request.method == 'POST':
        try:
            employee_id = int(request.form.get('employee_id', '0') or 0)
            period_start_str = request.form.get('period_start', '')
            period_end_str = request.form.get('period_end', '')
            
            if not all([employee_id, period_start_str, period_end_str]):
                flash('Employee, start date, and end date are required.', 'error')
                return redirect(url_for('main.generate_timesheet'))

            period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date()
            period_end = datetime.strptime(period_end_str, '%Y-%m-%d').date()

            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee_id,
                Attendance.date >= period_start,
                Attendance.date <= period_end
            ).all()

            regular_hours = sum(a.hours_worked for a in attendance_records)
            overtime_hours = sum(a.overtime_hours for a in attendance_records)

            timesheet = Timesheet(
                employee_id=employee_id,
                period_start=period_start,
                period_end=period_end,
                regular_hours=regular_hours,
                overtime_hours=overtime_hours,
                total_hours=regular_hours + overtime_hours,
                status='draft'
            )
            db.session.add(timesheet)
            db.session.commit()
            flash('Timesheet generated successfully!', 'success')
            return redirect(url_for('main.edit_timesheet', id=timesheet.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    employees = Employee.query.filter_by(is_active=True).all()
    return render_template('payroll/timesheet_form.html', employees=employees, timesheet=None)


@main_bp.route('/timesheets/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_timesheet(id):
    timesheet = Timesheet.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            timesheet.regular_hours = float(request.form.get('regular_hours', 0))
            timesheet.overtime_hours = float(request.form.get('overtime_hours', 0))
            timesheet.total_hours = timesheet.regular_hours + timesheet.overtime_hours
            timesheet.leave_days = float(request.form.get('leave_days', 0))
            timesheet.sick_days = float(request.form.get('sick_days', 0))
            timesheet.notes = request.form.get('notes', '')
            
            employee = Employee.query.get(timesheet.employee_id)
            hourly_rate = employee.base_salary / 176 if employee else 0
            # Further calculation logic...
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.timesheets'))
    
    return render_template('payroll/timesheet_form.html', employees=Employee.query.all(), timesheet=timesheet)


@main_bp.route('/timesheets/approve/<int:id>')
@login_required
def approve_timesheet(id):
    timesheet = Timesheet.query.get_or_404(id)
    timesheet.status = 'approved'
    timesheet.approved_by = current_user.username
    timesheet.approved_at = datetime.utcnow()
    db.session.commit()
    flash('Timesheet approved!', 'success')
    return redirect(url_for('main.timesheets'))


@main_bp.route('/timesheets/delete/<int:id>')
@login_required
def delete_timesheet(id):
    timesheet = Timesheet.query.get_or_404(id)
    db.session.delete(timesheet)
    db.session.commit()
    flash('Timesheet deleted!', 'success')
    return redirect(url_for('main.timesheets'))


# ============================================================================
# LEAVE MANAGEMENT
# ============================================================================

@main_bp.route('/leave-requests')
@login_required
def leave_requests():
    status = request.args.get('status', 'all')
    query = LeaveRequest.query
    if status != 'all':
        query = query.filter_by(status=status)
    leave_requests_list = query.order_by(LeaveRequest.created_at.desc()).all()
    return render_template('payroll/leave_requests.html', leave_requests=leave_requests_list, current_status=status)


@main_bp.route('/leave-requests/add', methods=['GET', 'POST'])
@login_required
def add_leave_request():
    if request.method == 'POST':
        try:
            employee_id = int(request.form.get('employee_id', '0') or 0)
            start_date_str = request.form.get('start_date', '')
            end_date_str = request.form.get('end_date', '')
            
            if not all([employee_id, start_date_str, end_date_str]):
                flash('Employee, start date, and end date are required.', 'error')
                return redirect(url_for('main.add_leave_request'))

            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            total_days = (end_date - start_date).days + 1

            leave_request = LeaveRequest(
                employee_id=employee_id,
                leave_type=request.form.get('leave_type', 'annual'),
                start_date=start_date,
                end_date=end_date,
                total_days=total_days,
                reason=request.form.get('reason', ''),
                status='pending'
            )
            db.session.add(leave_request)
            db.session.commit()
            flash('Leave request submitted', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.leave_requests'))
    
    employees = Employee.query.filter_by(is_active=True).all()
    return render_template('payroll/leave_form.html', employees=employees, leave_request=None)


@main_bp.route('/leave-requests/approve/<int:id>')
@login_required
def approve_leave(id):
    leave_request = LeaveRequest.query.get_or_404(id)
    leave_request.status = 'approved'
    leave_request.approved_by = current_user.username
    leave_request.approved_at = datetime.utcnow()
    db.session.commit()
    flash('Leave request approved!', 'success')
    return redirect(url_for('main.leave_requests'))


@main_bp.route('/leave-requests/reject/<int:id>', methods=['POST'])
@login_required
def reject_leave(id):
    leave_request = LeaveRequest.query.get_or_404(id)
    leave_request.status = 'rejected'
    leave_request.rejection_reason = request.form.get('reason', '')
    leave_request.approved_by = current_user.username
    leave_request.approved_at = datetime.utcnow()
    db.session.commit()
    flash('Leave request rejected!', 'warning')
    return redirect(url_for('main.leave_requests'))


# ============================================================================
# PAYROLL PAYMENTS
# ============================================================================

@main_bp.route('/payroll-payments')
@login_required
def payroll_payments():
    status = request.args.get('status', 'all')
    query = PayrollPayment.query
    if status != 'all':
        query = query.filter_by(status=status)
    payments = query.order_by(PayrollPayment.created_at.desc()).all()
    return render_template('payroll/payments.html', payments=payments, current_status=status)


@main_bp.route('/payroll-payments/process', methods=['GET', 'POST'])
@login_required
def process_payment():
    if request.method == 'POST':
        try:
            timesheet_id = int(request.form.get('timesheet_id', '0') or 0)
            timesheet = Timesheet.query.get_or_404(timesheet_id)
            employee = Employee.query.get(timesheet.employee_id)
            
            annual_salary = employee.base_salary * 12
            monthly_tax, _ = calculate_monthly_tax_deduction(annual_salary)
            
            loan_deduction = min(employee.outstanding_loan, timesheet.net_pay * 0.1)
            total_deductions = monthly_tax + loan_deduction
            net_pay = timesheet.net_pay - total_deductions
            
            payment = PayrollPayment(
                employee_id=employee.id,
                timesheet_id=timesheet_id,
                period_start=timesheet.period_start,
                period_end=timesheet.period_end,
                base_salary=employee.base_salary,
                hourly_wages=timesheet.regular_pay,
                overtime_pay=timesheet.overtime_pay,
                bonuses=float(request.form.get('bonuses', 0)),
                allowances=float(request.form.get('allowances', 0)),
                total_earnings=timesheet.net_pay + float(request.form.get('bonuses', 0)) + float(request.form.get('allowances', 0)),
                income_tax=monthly_tax,
                loan_deduction=loan_deduction,
                other_deductions=float(request.form.get('other_deductions', 0)),
                total_deductions=total_deductions + float(request.form.get('other_deductions', 0)),
                net_pay=net_pay,
                payment_method=request.form.get('payment_method', 'bank_transfer'),
                payment_date=datetime.utcnow().date(),
                bank_name=employee.bank_name,
                bank_account=employee.bank_account,
                status='processed',
                processed_by=current_user.username,
                processed_at=datetime.utcnow()
            )
            db.session.add(payment)
            
            if loan_deduction > 0:
                employee.outstanding_loan -= loan_deduction
            
            timesheet.status = 'paid'
            timesheet.paid_at = datetime.utcnow()
            
            db.session.commit()
            flash('Payment processed successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.payroll_payments'))
    
    timesheets = Timesheet.query.filter_by(status='approved').all()
    return render_template('payroll/payment_form.html', timesheets=timesheets)


# ============================================================================
# FBR TAX INTEGRATION (PHASE 9)
# ============================================================================

@main_bp.route('/tax/fbr-invoices')
@login_required
def fbr_invoices():
    status = request.args.get('status', 'all')
    query = FBRInvoice.query
    if status != 'all':
        query = query.filter_by(status=status)
    invoices = query.order_by(FBRInvoice.invoice_date.desc()).all()
    return render_template('tax/fbr_invoices.html', invoices=invoices, current_status=status)


@main_bp.route('/tax/fbr-invoices/generate/<int:order_id>', methods=['POST'])
@login_required
def generate_fbr_invoice(order_id):
    order = Order.query.get_or_404(order_id)
    
    try:
        invoice_number = f"FBR-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        taxable_amount = order.total_amount / 1.17
        gst_amount = order.total_amount - taxable_amount
        
        invoice = FBRInvoice(
            order_id=order.id,
            invoice_number=invoice_number,
            total_amount=order.total_amount,
            taxable_amount=taxable_amount,
            gst_amount=gst_amount
        )
        db.session.add(invoice)
        db.session.commit()
        
        flash(f'FBR Invoice {invoice_number} generated!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.view_order', id=order.id))


@main_bp.route('/tax/returns')
@login_required
def tax_returns():
    returns = TaxReturn.query.order_by(TaxReturn.return_period.desc()).all()
    return render_template('tax/returns.html', returns=returns)


@main_bp.route('/tax/returns/generate', methods=['GET', 'POST'])
@login_required
def generate_tax_return():
    if request.method == 'POST':
        try:
            return_period = request.form.get('return_period', '')
            
            month_start = datetime.strptime(return_period + '-01', '%Y-%m-%d')
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
            
            orders = Order.query.filter(
                Order.created_at >= month_start,
                Order.created_at < month_end,
                Order.status.in_(['completed', 'processing'])
            ).all()
            
            sales_tax = sum(o.tax_amount for o in orders)
            # Tax return creation logic...
            flash('Tax return generated', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.tax_returns'))
    
    return render_template('tax/return_form.html')


@main_bp.route('/tax/reports/sales-tax')
@login_required
def sales_tax_report():
    date_from = request.args.get('date_from', (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.utcnow().strftime('%Y-%m-%d'))
    
    orders = Order.query.filter(
        Order.created_at >= date_from,
        Order.created_at <= date_to,
        Order.status.in_(['completed', 'processing'])
    ).all()
    
    total_sales = sum(o.total_amount for o in orders)
    return render_template('tax/reports/sales_tax.html', 
                         orders=orders, 
                         total_sales=total_sales,
                         date_from=date_from,
                         date_to=date_to)

# ============================================================================
# PAINTING SERVICE (ENHANCED WITH PRODUCT-SPECIFIC PRICING)
# ============================================================================

@main_bp.route('/painting')
@login_required
def painting_dashboard():
    """Painting service dashboard"""
    active_orders = PaintingOrder.query.filter_by(status='in_progress').all()
    pending_orders = PaintingOrder.query.filter_by(status='pending').all()
    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    completed_this_month = PaintingOrder.query.filter(
        PaintingOrder.status == 'completed',
        PaintingOrder.order_date >= month_start
    ).all()
    
    total_revenue = sum(o.total_amount for o in completed_this_month)
    return render_template('painting/dashboard.html',
                         active_orders=active_orders,
                         pending_orders=pending_orders,
                         completed_this_month=completed_this_month,
                         total_revenue=total_revenue)


@main_bp.route('/painting/prices')
@login_required
def painting_prices():
    """Manage painting service prices per product"""
    prices = PaintingServicePrice.query.all()  # ✅ Remove is_active filter
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('painting/prices.html', prices=prices, products=products)


@main_bp.route('/painting/prices/add', methods=['GET', 'POST'])
@login_required
def add_painting_price():
    """Add or update painting price for a product"""
    if request.method == 'POST':
        try:
            product_id = int(request.form.get('product_id', 0))
            if not product_id:
                flash('Product is required.', 'error')
                return redirect(url_for('main.painting_prices'))

            existing = PaintingServicePrice.query.filter_by(product_id=product_id).first()
            
            if existing:
                existing.price_per_unit = float(request.form.get('price_per_unit', 0))
                existing.minimum_quantity = int(request.form.get('minimum_quantity', 1000))
                existing.setup_charge = float(request.form.get('setup_charge', 0))
                existing.is_active = request.form.get('is_active') == 'on'
                flash('Painting price updated!', 'success')
            else:
                # Create new
                price = PaintingServicePrice(
                    product_id=product_id,
                    price_per_unit=float(request.form.get('price_per_unit', 0)),
                    minimum_quantity=int(request.form.get('minimum_quantity', 1000)),
                    setup_charge=float(request.form.get('setup_charge', 0)),
                    is_active=True
                )
                db.session.add(price)
                flash('Painting price added successfully!', 'success')
            
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.painting_prices'))
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('painting/price_form.html', products=products, price=None)


@main_bp.route('/painting/prices/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_painting_price(id):
    """Edit painting price"""
    price = PaintingServicePrice.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            price.price_per_unit = float(request.form.get('price_per_unit', 0))
            price.minimum_quantity = int(request.form.get('minimum_quantity', 1000))
            price.setup_charge = float(request.form.get('setup_charge', 0))
            price.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash('Painting price updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('main.painting_prices'))
    
    products = Product.query.filter_by(is_deleted=False).all()
    return render_template('painting/price_form.html', products=products, price=price)


@main_bp.route('/painting/prices/delete/<int:id>')
@login_required
def delete_painting_price(id):
    """Delete painting price"""
    try:
        price = PaintingServicePrice.query.get_or_404(id)
        price.is_active = False
        db.session.commit()
        flash('Painting price archived successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.painting_prices'))


@main_bp.route('/painting/orders')
@login_required
def painting_orders():
    """List all painting orders"""
    status = request.args.get('status', 'all')
    query = PaintingOrder.query
    if status != 'all':
        query = query.filter_by(status=status)
    orders = query.order_by(PaintingOrder.order_date.desc()).all()
    return render_template('painting/orders.html', orders=orders, current_status=status)


@main_bp.route('/painting/orders/add', methods=['GET', 'POST'])
@login_required
def add_painting_order():
    """Create new painting order"""
    if request.method == 'POST':
        try:
            order_number = f"PAINT-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            
            # Get or create customer
            customer_id = request.form.get('customer_id')
            customer = None
            
            if customer_id:
                customer = Customer.query.get(int(customer_id))
            else:
                customer_name = request.form.get('customer_name', '')
                customer_phone = request.form.get('customer_phone', '')
                
                if customer_name:
                    customer, is_new = find_or_create_customer(
                        customer_name=customer_name,
                        phone=customer_phone
                    )
            
            order = PaintingOrder(
                order_number=order_number,
                customer_id=customer.id if customer else None,
                customer_name=customer.name if customer else request.form.get('customer_name', ''),
                customer_phone=customer.phone if customer else request.form.get('customer_phone', ''),
                notes=request.form.get('notes', ''),
                created_by=current_user.username
            )
            db.session.add(order)
            db.session.flush()
            
            product_ids = request.form.getlist('product_id[]')
            quantities = request.form.getlist('quantity[]')
            
            for i, pid in enumerate(product_ids):
                if pid and quantities[i]:
                    product = Product.query.get(int(pid))
                    
                    customer_painting_price = None
                    if customer:
                        customer_painting_price = CustomerPaintingPrice.query.filter_by(
                            customer_id=customer.id,
                            product_id=int(pid),
                            is_active=True
                        ).first()
                    
                    if customer_painting_price:
                        price_per_unit = customer_painting_price.price_per_unit
                        setup_charge = customer_painting_price.setup_charge
                    else:
                        painting_price = PaintingServicePrice.query.filter_by(
                            product_id=int(pid),
                            is_active=True
                        ).first()
                        price_per_unit = painting_price.price_per_unit if painting_price else 0
                        setup_charge = painting_price.setup_charge if painting_price else 0
                    
                    subtotal = (price_per_unit * int(quantities[i])) + setup_charge
                    order.total_amount += subtotal
                    
                    item = PaintingOrderItem(
                        order_id=order.id,
                        product_id=int(pid),
                        quantity=int(quantities[i]),
                        price_per_unit=price_per_unit,
                        setup_charge=setup_charge,
                        subtotal=subtotal,
                        color_specification=request.form.get(f'color_{pid}', ''),
                        special_instructions=request.form.get(f'instructions_{pid}', '')
                    )
                    db.session.add(item)
            
            db.session.commit()
            flash('Painting order created successfully!', 'success')
            return redirect(url_for('main.painting_orders'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    # GET request handling
    customers = Customer.query.filter_by(is_active=True).all()
    products = Product.query.filter_by(is_deleted=False).all()
    painting_prices = {p.product_id: p for p in PaintingServicePrice.query.all()}
    return render_template('painting/order_form.html',
                         customers=customers,
                         products=products,
                         painting_prices=painting_prices,
                         order=None)


@main_bp.route('/painting/orders/<int:id>')
@login_required
def view_painting_order(id):
    """View painting order details"""
    order = PaintingOrder.query.get_or_404(id)
    return render_template('painting/order_view.html', order=order)


@main_bp.route('/painting/orders/<int:id>/start', methods=['POST'])
@login_required
def start_painting_order(id):
    """Start painting order"""
    order = PaintingOrder.query.get_or_404(id)
    order.status = 'in_progress'
    db.session.commit()
    flash('Painting order started!', 'success')
    return redirect(url_for('main.view_painting_order', id=order.id))


@main_bp.route('/painting/orders/<int:id>/complete', methods=['POST'])
@login_required
def complete_painting_order(id):
    """Complete painting order"""
    order = PaintingOrder.query.get_or_404(id)
    order.status = 'completed'
    order.delivery_date = datetime.utcnow()
    db.session.commit()
    flash('Painting order completed!', 'success')
    return redirect(url_for('main.view_painting_order', id=order.id))


@main_bp.route('/painting/orders/<int:id>/invoice')
@login_required
def painting_order_invoice(id):
    """Generate painting order invoice"""
    order = PaintingOrder.query.get_or_404(id)
    return render_template('painting/invoice.html', order=order)

# ============================================================================
# ACCOUNTING MODULE ROUTES
# ============================================================================

@main_bp.route('/accounting')
@login_required
def accounting_dashboard():
    """Accounting Dashboard"""
    # Get account balances by type
    accounts = Account.query.filter_by(is_active=True).all()
    
    # Calculate balances from journal entries
    account_balances = {}
    for account in accounts:
        debits = db.session.query(db.func.sum(JournalEntryLine.debit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account.id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        
        credits = db.session.query(db.func.sum(JournalEntryLine.credit)).join(JournalEntry).filter(
            JournalEntryLine.account_id == account.id,
            JournalEntry.status == 'posted'
        ).scalar() or 0
        
        if account.account_type in ['Asset', 'Expense']:
            balance = debits - credits
        else:
            balance = credits - debits
        
        account_balances[account.id] = balance
    
    # Recent journal entries
    recent_entries = JournalEntry.query.filter_by(status='posted').order_by(
        JournalEntry.created_at.desc()
    ).limit(10).all()
    
    # Pending vouchers
    pending_payments = PaymentVoucher.query.filter_by(status='draft').count()
    pending_receipts = ReceiptVoucher.query.filter_by(status='draft').count()
    
    # Bank accounts
    bank_accounts = BankAccount.query.filter_by(is_active=True).all()
    
    return render_template('accounting/dashboard.html',
                         accounts=accounts,
                         account_balances=account_balances,
                         recent_entries=recent_entries,
                         pending_payments=pending_payments,
                         pending_receipts=pending_receipts,
                         bank_accounts=bank_accounts)


@main_bp.route('/accounting/chart-of-accounts')
@login_required
def chart_of_accounts():
    """Chart of Accounts Management"""
    accounts = Account.query.filter_by(is_active=True).order_by(Account.account_code).all()
    return render_template('accounting/chart_of_accounts.html', accounts=accounts)


@main_bp.route('/accounting/accounts/add', methods=['GET', 'POST'])
@login_required
def add_account():
    """Add New Account"""
    if request.method == 'POST':
        try:
            account = Account(
                account_code=request.form.get('account_code', ''),
                account_name=request.form.get('account_name', ''),
                account_type=request.form.get('account_type', ''),
                parent_account_id=int(request.form.get('parent_account_id', 0)) or None,
                is_active=request.form.get('is_active') == 'on'
            )
            db.session.add(account)
            db.session.commit()
            
            # Audit log
            audit = AuditLog(
                table_name='account',
                record_id=account.id,
                action='CREATE',
                new_values=json.dumps({
                    'account_code': account.account_code,
                    'account_name': account.account_name
                }),
                username=current_user.username
            )
            db.session.add(audit)
            db.session.commit()
            
            flash('Account added successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('main.chart_of_accounts'))
    
    parent_accounts = Account.query.filter_by(is_active=True).all()
    return render_template('accounting/account_form.html', account=None, parent_accounts=parent_accounts)


@main_bp.route('/accounting/accounts/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_account(id):
    """Edit Account"""
    account = Account.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            account.account_code = request.form.get('account_code', '')
            account.account_name = request.form.get('account_name', '')
            account.account_type = request.form.get('account_type', '')
            account.parent_account_id = int(request.form.get('parent_account_id', 0)) or None
            account.is_active = request.form.get('is_active') == 'on'
            db.session.commit()
            flash('Account updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('main.chart_of_accounts'))
    
    parent_accounts = Account.query.filter_by(is_active=True).all()
    return render_template('accounting/account_form.html', account=account, parent_accounts=parent_accounts)


@main_bp.route('/accounting/accounts/delete/<int:id>', methods=['POST'])
@login_required
def delete_account(id):
    """Deactivate Account"""
    account = Account.query.get_or_404(id)
    
    try:
        account.is_active = False
        db.session.commit()
        flash('Account deactivated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.chart_of_accounts'))


@main_bp.route('/accounting/journal-entries')
@login_required
def journal_entries():
    """Journal Entries List"""
    status = request.args.get('status', 'all')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = JournalEntry.query
    
    if status != 'all':
        query = query.filter_by(status=status)
    
    if date_from:
        query = query.filter(JournalEntry.entry_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    
    if date_to:
        query = query.filter(JournalEntry.entry_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    entries = query.order_by(JournalEntry.entry_date.desc()).all()
    
    return render_template('accounting/journal_entries.html',
                         entries=entries,
                         current_status=status,
                         date_from=date_from,
                         date_to=date_to)


@main_bp.route('/accounting/journal-entries/add', methods=['GET', 'POST'])
@login_required
def add_journal_entry():
    """Create New Journal Entry with Period Validation"""
    if request.method == 'POST':
        try:
            entry_number = f"JE-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            entry_date_str = request.form.get('entry_date', '')
            entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else datetime.utcnow().date()
            
            # ✅ STEP 7: CHECK IF ACCOUNTING PERIOD IS CLOSED
            period = AccountingPeriod.query.filter(
                AccountingPeriod.start_date <= entry_date,
                AccountingPeriod.end_date >= entry_date
            ).first()
            
            if period and period.is_closed:
                flash(f'Cannot post entries to closed period: {period.period_name} ({period.start_date} to {period.end_date})', 'error')
                return redirect(url_for('main.add_journal_entry'))
            
            # Create journal entry
            entry = JournalEntry(
                entry_number=entry_number,
                entry_date=entry_date,
                description=request.form.get('description', ''),
                reference=request.form.get('reference', ''),
                reference_type=request.form.get('reference_type', ''),
                status='draft',
                created_by=current_user.username
            )
            db.session.add(entry)
            db.session.flush()
            
            # Add lines
            account_ids = request.form.getlist('account_id[]')
            descriptions = request.form.getlist('line_description[]')
            debits = request.form.getlist('debit[]')
            credits = request.form.getlist('credit[]')
            
            total_debit = 0
            total_credit = 0
            
            for i, aid in enumerate(account_ids):
                if aid:
                    debit = float(debits[i]) if debits[i] else 0
                    credit = float(credits[i]) if credits[i] else 0
                    
                    line = JournalEntryLine(
                        entry_id=entry.id,
                        account_id=int(aid),
                        description=descriptions[i] if i < len(descriptions) else '',
                        debit=debit,
                        credit=credit
                    )
                    db.session.add(line)
                    total_debit += debit
                    total_credit += credit
            
            # Validate debits = credits
            if abs(total_debit - total_credit) > 0.01:
                db.session.rollback()
                flash(f'Error: Debits ({total_debit:.2f}) must equal Credits ({total_credit:.2f})', 'error')
                return redirect(url_for('main.add_journal_entry'))
            
            db.session.commit()
            flash(f'Journal Entry {entry_number} created successfully!', 'success')
            return redirect(url_for('main.journal_entries'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    accounts = Account.query.filter_by(is_active=True).all()
    
    # ✅ Get open accounting periods for dropdown
    open_periods = AccountingPeriod.query.filter_by(is_closed=False).order_by(
        AccountingPeriod.start_date.desc()
    ).all()
    
    return render_template('accounting/journal_entry_form.html', 
                         accounts=accounts, 
                         entry=None,
                         open_periods=open_periods)

@main_bp.route('/accounting/journal-entries/<int:id>/post', methods=['POST'])
@login_required
def post_journal_entry(id):
    """Post Journal Entry with Period Validation"""
    entry = JournalEntry.query.get_or_404(id)
    
    if entry.status != 'draft':
        flash('Only draft entries can be posted', 'error')
        return redirect(url_for('main.journal_entries'))
    
    # ✅ STEP 7: CHECK IF PERIOD IS CLOSED BEFORE POSTING
    period = AccountingPeriod.query.filter(
        AccountingPeriod.start_date <= entry.entry_date,
        AccountingPeriod.end_date >= entry.entry_date
    ).first()
    
    if period and period.is_closed:
        flash(f'Cannot post entries to closed period: {period.period_name}', 'error')
        return redirect(url_for('main.journal_entries'))
    
    try:
        entry.status = 'posted'
        entry.posted_by = current_user.username
        entry.posted_at = datetime.utcnow()
        db.session.commit()
        
        # Audit log
        audit = AuditLog(
            table_name='journal_entry',
            record_id=entry.id,
            action='UPDATE',
            old_values=json.dumps({'status': 'draft'}),
            new_values=json.dumps({'status': 'posted'}),
            username=current_user.username
        )
        db.session.add(audit)
        db.session.commit()
        
        flash('Journal entry posted!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.journal_entries'))

@main_bp.route('/accounting/journal-entries/<int:id>')
@login_required
def view_journal_entry(id):
    """View Journal Entry Details"""
    entry = JournalEntry.query.get_or_404(id)
    return render_template('accounting/journal_entry_view.html', entry=entry)


@main_bp.route('/accounting/general-ledger')
@login_required
def general_ledger():
    """General Ledger View"""
    account_id = request.args.get('account_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    accounts = Account.query.filter_by(is_active=True).order_by(Account.account_code).all()
    
    if account_id:
        account = Account.query.get_or_404(account_id)
        
        query = JournalEntryLine.query.join(JournalEntry).filter(
            JournalEntryLine.account_id == account_id,
            JournalEntry.status == 'posted'
        )
        
        if date_from:
            query = query.filter(JournalEntry.entry_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        
        if date_to:
            query = query.filter(JournalEntry.entry_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        
        lines = query.order_by(JournalEntry.entry_date).all()
        
        # Calculate running balance
        running_balance = 0
        for line in lines:
            if account.account_type in ['Asset', 'Expense']:
                running_balance += line.debit - line.credit
            else:
                running_balance += line.credit - line.debit
            line.running_balance = running_balance
    else:
        account = None
        lines = []
    
    return render_template('accounting/general_ledger.html',
                         accounts=accounts,
                         account=account,
                         lines=lines,
                         date_from=date_from,
                         date_to=date_to)


@main_bp.route('/accounting/trial-balance')
@login_required
def trial_balance():
    """Trial Balance Report"""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    accounts = Account.query.filter_by(is_active=True).order_by(Account.account_code).all()
    
    trial_data = []
    total_debits = 0
    total_credits = 0
    
    for account in accounts:
        query = JournalEntryLine.query.join(JournalEntry).filter(
            JournalEntryLine.account_id == account.id,
            JournalEntry.status == 'posted'
        )
        
        if date_from:
            query = query.filter(JournalEntry.entry_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        
        if date_to:
            query = query.filter(JournalEntry.entry_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        
        debits = db.session.query(db.func.sum(JournalEntryLine.debit)).filter(
            JournalEntryLine.account_id == account.id
        ).scalar() or 0
        
        credits = db.session.query(db.func.sum(JournalEntryLine.credit)).filter(
            JournalEntryLine.account_id == account.id
        ).scalar() or 0
        
        if debits > 0 or credits > 0:
            trial_data.append({
                'account': account,
                'debit': debits,
                'credit': credits
            })
            total_debits += debits
            total_credits += credits
    
    return render_template('accounting/trial_balance.html',
                         trial_data=trial_data,
                         total_debits=total_debits,
                         total_credits=total_credits,
                         date_from=date_from,
                         date_to=date_to)


@main_bp.route('/accounting/payment-vouchers')
@login_required
def payment_vouchers():
    """Payment Vouchers List"""
    status = request.args.get('status', 'all')
    query = PaymentVoucher.query
    
    if status != 'all':
        query = query.filter_by(status=status)
    
    vouchers = query.order_by(PaymentVoucher.created_at.desc()).all()
    
    return render_template('accounting/payment_vouchers.html',
                         vouchers=vouchers,
                         current_status=status)


@main_bp.route('/accounting/payment-vouchers/add', methods=['GET', 'POST'])
@login_required
def add_payment_voucher():
    """Create Payment Voucher"""
    if request.method == 'POST':
        try:
            voucher_number = f"PV-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            voucher_date_str = request.form.get('voucher_date', '')
            voucher_date = datetime.strptime(voucher_date_str, '%Y-%m-%d').date() if voucher_date_str else datetime.utcnow().date()
            
            voucher = PaymentVoucher(
                voucher_number=voucher_number,
                voucher_date=voucher_date,
                payee_name=request.form.get('payee_name', ''),
                payee_type=request.form.get('payee_type', ''),
                payment_method=request.form.get('payment_method', 'cash'),
                bank_account_id=int(request.form.get('bank_account_id', 0)) or None,
                cheque_number=request.form.get('cheque_number', ''),
                total_amount=0,
                description=request.form.get('description', ''),
                status='draft',
                created_by=current_user.username
            )
            db.session.add(voucher)
            db.session.flush()
            
            # Add lines
            account_ids = request.form.getlist('account_id[]')
            descriptions = request.form.getlist('line_description[]')
            amounts = request.form.getlist('amount[]')
            
            total_amount = 0
            
            for i, aid in enumerate(account_ids):
                if aid and amounts[i]:
                    amount = float(amounts[i])
                    
                    line = PaymentVoucherLine(
                        voucher_id=voucher.id,
                        account_id=int(aid),
                        description=descriptions[i] if i < len(descriptions) else '',
                        amount=amount
                    )
                    db.session.add(line)
                    total_amount += amount
            
            voucher.total_amount = total_amount
            db.session.commit()
            
            flash(f'Payment Voucher {voucher_number} created!', 'success')
            return redirect(url_for('main.payment_vouchers'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    accounts = Account.query.filter_by(is_active=True).all()
    bank_accounts = BankAccount.query.filter_by(is_active=True).all()
    return render_template('accounting/payment_voucher_form.html',
                         voucher=None,
                         accounts=accounts,
                         bank_accounts=bank_accounts)


@main_bp.route('/accounting/payment-vouchers/<int:id>')
@login_required
def view_payment_voucher(id):
    """View Payment Voucher Details"""
    voucher = PaymentVoucher.query.get_or_404(id)
    return render_template('accounting/payment_voucher_view.html', voucher=voucher)


@main_bp.route('/accounting/receipt-vouchers')
@login_required
def receipt_vouchers():
    """Receipt Vouchers List"""
    status = request.args.get('status', 'all')
    query = ReceiptVoucher.query
    
    if status != 'all':
        query = query.filter_by(status=status)
    
    vouchers = query.order_by(ReceiptVoucher.created_at.desc()).all()
    
    return render_template('accounting/receipt_vouchers.html',
                         vouchers=vouchers,
                         current_status=status)


@main_bp.route('/accounting/receipt-vouchers/add', methods=['GET', 'POST'])
@login_required
def add_receipt_voucher():
    """Create Receipt Voucher"""
    if request.method == 'POST':
        try:
            voucher_number = f"RV-{datetime.utcnow().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            voucher_date_str = request.form.get('voucher_date', '')
            voucher_date = datetime.strptime(voucher_date_str, '%Y-%m-%d').date() if voucher_date_str else datetime.utcnow().date()
            
            voucher = ReceiptVoucher(
                voucher_number=voucher_number,
                voucher_date=voucher_date,
                payer_name=request.form.get('payer_name', ''),
                payer_type=request.form.get('payer_type', ''),
                payer_id=int(request.form.get('customer_id', 0)) or None,
                payment_method=request.form.get('payment_method', 'cash'),
                bank_account_id=int(request.form.get('bank_account_id', 0)) or None,
                cheque_number=request.form.get('cheque_number', ''),
                total_amount=0,
                description=request.form.get('description', ''),
                status='draft',
                created_by=current_user.username
            )
            db.session.add(voucher)
            db.session.flush()
            
            # Add lines
            account_ids = request.form.getlist('account_id[]')
            descriptions = request.form.getlist('line_description[]')
            amounts = request.form.getlist('amount[]')
            
            total_amount = 0
            
            for i, aid in enumerate(account_ids):
                if aid and amounts[i]:
                    amount = float(amounts[i])
                    
                    line = ReceiptVoucherLine(
                        voucher_id=voucher.id,
                        account_id=int(aid),
                        description=descriptions[i] if i < len(descriptions) else '',
                        amount=amount
                    )
                    db.session.add(line)
                    total_amount += amount
            
            voucher.total_amount = total_amount
            db.session.commit()
            
            flash(f'Receipt Voucher {voucher_number} created!', 'success')
            return redirect(url_for('main.receipt_vouchers'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    accounts = Account.query.filter_by(is_active=True).all()
    bank_accounts = BankAccount.query.filter_by(is_active=True).all()
    customers = Customer.query.filter_by(is_active=True, is_deleted=False).all()
    return render_template('accounting/receipt_voucher_form.html',
                         voucher=None,
                         accounts=accounts,
                         bank_accounts=bank_accounts,
                         customers=customers)


@main_bp.route('/accounting/receipt-vouchers/<int:id>')
@login_required
def view_receipt_voucher(id):
    """View Receipt Voucher Details"""
    voucher = ReceiptVoucher.query.get_or_404(id)
    return render_template('accounting/receipt_voucher_view.html', voucher=voucher)


@main_bp.route('/accounting/bank-accounts')
@login_required
def bank_accounts():
    """Bank Accounts Management"""
    accounts = BankAccount.query.filter_by(is_active=True).all()
    return render_template('accounting/bank_accounts.html', accounts=accounts)


@main_bp.route('/accounting/bank-reconciliation')
@login_required
def bank_reconciliation():
    """Bank Reconciliation"""
    bank_account_id = request.args.get('bank_account_id', type=int)
    
    bank_accounts = BankAccount.query.filter_by(is_active=True).all()
    reconciliations = []
    
    if bank_account_id:
        reconciliations = BankReconciliation.query.filter_by(
            bank_account_id=bank_account_id
        ).order_by(BankReconciliation.reconciliation_date.desc()).all()
    
    return render_template('accounting/bank_reconciliation.html',
                         bank_accounts=bank_accounts,
                         reconciliations=reconciliations,
                         selected_account=bank_account_id)


@main_bp.route('/accounting/audit-log')
@login_required
def audit_log():
    """Audit Trail Log"""
    table_name = request.args.get('table_name', '')
    username = request.args.get('username', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = AuditLog.query
    
    if table_name:
        query = query.filter_by(table_name=table_name)
    
    if username:
        query = query.filter_by(username=username)
    
    if date_from:
        query = query.filter(AuditLog.timestamp >= datetime.strptime(date_from, '%Y-%m-%d'))
    
    if date_to:
        query = query.filter(AuditLog.timestamp <= datetime.strptime(date_to, '%Y-%m-%d'))
    
    logs = query.order_by(AuditLog.timestamp.desc()).limit(100).all()
    
    return render_template('accounting/audit_log.html',
                         logs=logs,
                         table_name=table_name,
                         username=username,
                         date_from=date_from,
                         date_to=date_to)


# ============================================================================
# ACCOUNTING PERIOD MANAGEMENT
# ============================================================================

@main_bp.route('/accounting/periods')
@login_required
def accounting_periods():
    """List Accounting Periods"""
    periods = AccountingPeriod.query.order_by(AccountingPeriod.start_date.desc()).all()
    return render_template('accounting/periods.html', periods=periods)


@main_bp.route('/accounting/periods/add', methods=['GET', 'POST'])
@login_required
def add_accounting_period():
    """Add New Accounting Period"""
    if request.method == 'POST':
        try:
            period_name = request.form.get('period_name', '')
            start_date = datetime.strptime(request.form.get('start_date', ''), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.form.get('end_date', ''), '%Y-%m-%d').date()
            
            # Check for overlapping periods
            existing = AccountingPeriod.query.filter(
                ((AccountingPeriod.start_date <= start_date) & (AccountingPeriod.end_date >= start_date)) |
                ((AccountingPeriod.start_date <= end_date) & (AccountingPeriod.end_date >= end_date))
            ).first()
            
            if existing:
                flash(f'Overlapping period exists: {existing.period_name}', 'error')
                return redirect(url_for('main.add_accounting_period'))
            
            period = AccountingPeriod(
                period_name=period_name,
                start_date=start_date,
                end_date=end_date,
                is_closed=False
            )
            db.session.add(period)
            db.session.commit()
            
            flash(f'Accounting period {period_name} created!', 'success')
            return redirect(url_for('main.accounting_periods'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('accounting/period_form.html', period=None)


@main_bp.route('/accounting/periods/<int:id>/close', methods=['POST'])
@login_required
def close_accounting_period(id):
    """Close Accounting Period"""
    period = AccountingPeriod.query.get_or_404(id)
    
    try:
        period.is_closed = True
        period.closed_by = current_user.username
        period.closed_at = datetime.utcnow()
        db.session.commit()
        
        # Audit log
        audit = AuditLog(
            table_name='accounting_period',
            record_id=period.id,
            action='UPDATE',
            old_values=json.dumps({'is_closed': False}),
            new_values=json.dumps({'is_closed': True}),
            username=current_user.username
        )
        db.session.add(audit)
        db.session.commit()
        
        flash(f'Period {period.period_name} closed!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.accounting_periods'))


@main_bp.route('/accounting/periods/<int:id>/reopen', methods=['POST'])
@login_required
def reopen_accounting_period(id):
    """Reopen Accounting Period"""
    period = AccountingPeriod.query.get_or_404(id)
    
    try:
        period.is_closed = False
        period.closed_by = None
        period.closed_at = None
        db.session.commit()
        
        flash(f'Period {period.period_name} reopened!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.accounting_periods'))


# ============================================================================
# EXPENSES - BULK UPLOAD
# ============================================================================
@main_bp.route('/expenses/bulk-upload', methods=['GET', 'POST'])
@login_required
def expenses_bulk_upload():
    """Bulk upload expenses from Excel file"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                import pandas as pd
                import io
                
                # Read Excel file
                df = pd.read_excel(io.BytesIO(file.read()))
                
                # Validate required columns
                required_cols = ['category', 'amount', 'date']
                if not all(col in df.columns for col in required_cols):
                    flash(f'Excel must have columns: {", ".join(required_cols)}', 'error')
                    return redirect(request.url)
                
                # Process each row
                uploaded_count = 0
                errors = []
                
                for idx, row in df.iterrows():
                    try:
                        # Parse date
                        if isinstance(row['date'], str):
                            expense_date = datetime.strptime(row['date'], '%Y-%m-%d').date()
                        else:
                            expense_date = row['date'].date() if hasattr(row['date'], 'date') else datetime.utcnow().date()
                        
                        expense = Expense(
                            category=str(row['category']).strip(),
                            description=str(row.get('description', '')).strip(),
                            amount=float(row['amount']),
                            date=expense_date,
                            status=str(row.get('status', 'pending')).strip(),
                            paid_by=str(row.get('paid_by', current_user.username)).strip(),
                            payment_method=str(row.get('payment_method', 'cash')).strip(),
                            reference_number=str(row.get('reference_number', '')).strip(),
                            approved_by=str(row.get('approved_by', '')).strip()
                        )
                        db.session.add(expense)
                        uploaded_count += 1
                    except Exception as e:
                        errors.append(f'Row {idx + 2}: {str(e)}')
                
                db.session.commit()
                
                if errors:
                    flash(f'Uploaded {uploaded_count} expenses. {len(errors)} errors: {" | ".join(errors[:3])}', 'warning')
                else:
                    flash(f'Successfully uploaded {uploaded_count} expenses!', 'success')
                
                return redirect(url_for('main.expenses'))
                
            except ImportError:
                flash('pandas library not installed. Run: pip install pandas openpyxl', 'error')
                return redirect(url_for('main.expenses'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error processing file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
    
    return render_template('expenses/bulk_upload.html')

# ============================================================================
# PRODUCTS - BULK UPLOAD
# ============================================================================
@main_bp.route('/products/bulk-upload', methods=['GET', 'POST'])
@login_required
def products_bulk_upload():
    """Bulk upload products from Excel file"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                import pandas as pd
                import io
                
                df = pd.read_excel(io.BytesIO(file.read()))
                
                # Validate required columns
                required_cols = ['name', 'base_price']
                if not all(col in df.columns for col in required_cols):
                    flash(f'Excel must have columns: {", ".join(required_cols)}', 'error')
                    return redirect(request.url)
                
                uploaded_count = 0
                errors = []
                
                for idx, row in df.iterrows():
                    try:
                        base_price = float(row.get('base_price', 0))
                        product = Product(
                            name=str(row['name']).strip(),
                            specification=str(row.get('specification', '')).strip(),
                            base_name=str(row.get('base_name', '')).strip(),
                            volume_cc=float(row.get('volume_cc', 0)) if row.get('volume_cc') else None,
                            glass_type=str(row.get('glass_type', '')).strip(),
                            neck_finish=str(row.get('neck_finish', '')).strip(),
                            sku=str(row.get('sku', '')).strip(),
                            color=str(row.get('color', '')).strip(),
                            product_type=str(row.get('product_type', 'product')).strip(),
                            base_price=base_price,
                            price_per_unit=base_price / 1000 if base_price > 0 else 0,
                            stock=int(row.get('stock', 0)),
                            is_active=row.get('is_active', 'true').lower() == 'true'
                        )
                        db.session.add(product)
                        uploaded_count += 1
                    except Exception as e:
                        errors.append(f'Row {idx + 2}: {str(e)}')
                
                db.session.commit()
                
                if errors:
                    flash(f'Uploaded {uploaded_count} products. {len(errors)} errors: {" | ".join(errors[:3])}', 'warning')
                else:
                    flash(f'Successfully uploaded {uploaded_count} products!', 'success')
                
                return redirect(url_for('main.products'))
                
            except ImportError:
                flash('pandas library not installed. Run: pip install pandas openpyxl', 'error')
                return redirect(url_for('main.products'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error processing file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
    
    return render_template('products/bulk_upload.html')

# ============================================================================
# CUSTOMERS - BULK UPLOAD
# ============================================================================
@main_bp.route('/customers/bulk-upload', methods=['GET', 'POST'])
@login_required
def customers_bulk_upload():
    """Bulk upload customers from Excel file"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                import pandas as pd
                import io
                
                df = pd.read_excel(io.BytesIO(file.read()))
                
                required_cols = ['name']
                if not all(col in df.columns for col in required_cols):
                    flash(f'Excel must have columns: {", ".join(required_cols)}', 'error')
                    return redirect(request.url)
                
                uploaded_count = 0
                errors = []
                
                for idx, row in df.iterrows():
                    try:
                        # Check for duplicate email
                        email = str(row.get('email', '')).strip().lower()
                        if email:
                            existing = Customer.query.filter_by(email=email).first()
                            if existing:
                                errors.append(f'Row {idx + 2}: Email {email} already exists')
                                continue
                        
                        customer = Customer(
                            name=str(row['name']).strip(),
                            business_name=str(row.get('business_name', '')).strip(),
                            role=str(row.get('role', 'customer')).strip(),
                            email=email if email else None,
                            phone=str(row.get('phone', '')).strip(),
                            address=str(row.get('address', '')).strip(),
                            payment_terms=str(row.get('payment_terms', 'cash')).strip(),
                            credit_limit=float(row.get('credit_limit', 0)),
                            is_active=row.get('is_active', 'true').lower() == 'true'
                        )
                        db.session.add(customer)
                        uploaded_count += 1
                    except Exception as e:
                        errors.append(f'Row {idx + 2}: {str(e)}')
                
                db.session.commit()
                
                if errors:
                    flash(f'Uploaded {uploaded_count} customers. {len(errors)} errors: {" | ".join(errors[:3])}', 'warning')
                else:
                    flash(f'Successfully uploaded {uploaded_count} customers!', 'success')
                
                return redirect(url_for('main.customers'))
                
            except ImportError:
                flash('pandas library not installed. Run: pip install pandas openpyxl', 'error')
                return redirect(url_for('main.customers'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error processing file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
    
    return render_template('customers/bulk_upload.html')


# ============================================================================
# TEMPLATE DOWNLOAD ROUTES
# ============================================================================
@main_bp.route('/expenses/download-template')
@login_required
def download_expense_template():
    """Download sample Excel template for expenses"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Create sample dataframe
        data = {
            'category': ['Utilities', 'Supplies', 'Maintenance'],
            'amount': [5000, 2500, 10000],
            'date': ['2026-03-01', '2026-03-05', '2026-03-10'],
            'description': ['Electricity bill', 'Office supplies', 'Machine repair'],
            'status': ['pending', 'pending', 'pending'],
            'paid_by': ['', '', ''],
            'payment_method': ['cash', 'bank_transfer', 'cash'],
            'reference_number': ['', '', ''],
            'approved_by': ['', '', '']
        }
        df = pd.DataFrame(data)
        
        # Write to BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Expenses')
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='expense_upload_template.xlsx'
        )
    except ImportError:
        flash('pandas or openpyxl not installed', 'error')
        return redirect(url_for('main.expenses'))

@main_bp.route('/products/download-template')
@login_required
def download_product_template():
    """Download sample Excel template for products"""
    try:
        import pandas as pd
        from io import BytesIO
        
        data = {
            'name': ['5cc Amber Ampoule', '10cc Transparent Ampoule'],
            'base_price': [150, 200],
            'specification': ['Type I glass', 'Type I glass'],
            'base_name': ['Ampoule', 'Ampoule'],
            'volume_cc': [5, 10],
            'glass_type': ['Amber', 'Transparent'],
            'neck_finish': ['Crimp', 'Crimp'],
            'sku': ['AMP-5CC-AMB', 'AMP-10CC-TRN'],
            'color': ['Amber', 'Clear'],
            'product_type': ['product', 'product'],
            'stock': [50000, 30000],
            'is_active': ['true', 'true']
        }
        df = pd.DataFrame(data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Products')
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='product_upload_template.xlsx'
        )
    except ImportError:
        flash('pandas or openpyxl not installed', 'error')
        return redirect(url_for('main.products'))            

# ============================================================================
# SOCKET.IO EVENT HANDLERS
# ============================================================================

@socketio.on('connect', namespace='/admin')
def admin_connect():
    if not current_user.is_authenticated:
        from flask_socketio import disconnect
        disconnect()
        return False
    emit('connected', {'message': 'Connected to AmpouleX real-time updates'})


@socketio.on('disconnect', namespace='/admin')
def admin_disconnect():
    username = current_user.username if current_user.is_authenticated else 'Anonymous'
    print(f'Admin client disconnected: {username}')